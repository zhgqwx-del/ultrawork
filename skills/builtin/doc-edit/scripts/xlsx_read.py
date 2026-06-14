#!/usr/bin/env python3
"""Read a .xlsx: dump sheet values. ultrawork doc-edit skill."""
from __future__ import annotations
import argparse, json, sys

def _require():
    try:
        import openpyxl  # noqa: F401
        return openpyxl
    except ImportError:
        print("Missing dependency: openpyxl (pip install openpyxl)", file=sys.stderr)
        raise SystemExit(1)

def main(argv):
    ap = argparse.ArgumentParser(description="Read values from a .xlsx")
    ap.add_argument("file")
    ap.add_argument("--sheet", help="sheet name (default: all sheets)")
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args(argv)
    openpyxl = _require()
    try:
        wb = openpyxl.load_workbook(args.file, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        print(f"Error opening {args.file}: {exc}", file=sys.stderr)
        return 1
    names = [args.sheet] if args.sheet else wb.sheetnames
    out = {}
    for name in names:
        if name not in wb.sheetnames:
            print(f"Sheet not found: {name} (have: {', '.join(wb.sheetnames)})", file=sys.stderr)
            return 1
        ws = wb[name]
        out[name] = [["" if c is None else c for c in row] for row in ws.iter_rows(values_only=True)]
    if args.json:
        print(json.dumps(out, ensure_ascii=False, indent=2, default=str))
    else:
        for name, rows in out.items():
            print(f"=== {name} ===")
            for row in rows:
                print(",".join(str(c) for c in row))
    return 0

if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
