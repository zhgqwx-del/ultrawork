#!/usr/bin/env python3
"""X4 — recalculate a workbook, with two engines checking each other.

Why this exists at all: **openpyxl never writes cached values.** Every workbook a
library produces stores `<f>` and no `<v>`, so anything that reads values rather
than formulas — this skill's own reader included — sees an empty cell where a
number belongs. Recalculating and writing the results back is what makes such a
file readable by something that is not a spreadsheet application.

Two engines, on purpose (059 §7):

    soffice  LibreOffice, the authority. Converts the workbook to .xlsx, which
             forces a full recalculation and writes cached values for EVERY sheet
             (the csv route only ever covers the first one).
    python   office/evaluate.py — small, deliberately incomplete, and refuses
             anything it does not implement rather than guessing.

    --engine both (default) runs both and REPORTS EVERY DISAGREEMENT. It does not
    pick a winner: two independent engines differing on a number means one of them
    is wrong, and which one is not something this script can know.

    python3 xlsx_recalc.py --in book.xlsx --out calculated.xlsx --report r.json
    python3 xlsx_recalc.py --in book.xlsx --engine python      # no LibreOffice
    python3 xlsx_recalc.py --in book.xlsx --fail-on disagreement,unsupported

Formulas the python engine cannot do are listed with **their formula text**, never
with a number. That is the one behaviour this file must never lose: a wrong value
that looks calculated is worse than an honest gap.
"""
from __future__ import annotations

import argparse
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from office.evaluate import Evaluator, ExcelError, SUPPORTED, Unsupported  # noqa: E402
from office.package import Package  # noqa: E402
from office.sheet import Workbook  # noqa: E402
from office.soffice import convert, find_soffice  # noqa: E402
from xlsxcommon import XlsxError, emit, ensure_distinct, fail, run  # noqa: E402

TOLERANCE = 1e-9
ENGINES = ("both", "soffice", "python")
FAIL_CLASSES = ("disagreement", "unsupported", "error")


def numeric(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def same(a, b) -> bool:
    """Do two engines agree? Blank and 0 are the same cell state in Excel."""
    if a is None:
        a = 0 if numeric(b) else a
    if b is None:
        b = 0 if numeric(a) else b
    if numeric(a) and numeric(b):
        scale = max(1.0, abs(float(a)), abs(float(b)))
        return abs(float(a) - float(b)) <= TOLERANCE * scale
    if isinstance(a, bool) or isinstance(b, bool):
        return bool(a) == bool(b)
    return (a if a is not None else "") == (b if b is not None else "")


def read_cells(path: Path) -> dict[str, dict[str, object]]:
    """{sheet: {ref: formula-or-constant}} — the input to the python engine."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        return {ws.title: {c.coordinate: c.value for row in ws.iter_rows()
                           for c in row if c.value is not None}
                for ws in wb.worksheets}
    finally:
        wb.close()


def soffice_values(path: Path, timeout: int) -> tuple[dict, str]:
    """Recalculated values for every sheet, via a LibreOffice xlsx round trip."""
    import openpyxl
    with tempfile.TemporaryDirectory(prefix="xlsx-recalc-") as td:
        out, err = convert(path, "xlsx", Path(td), timeout=timeout)
        if out is None:
            return {}, err
        wb = openpyxl.load_workbook(out, data_only=True)
        try:
            return ({ws.title: {c.coordinate: c.value for row in ws.iter_rows()
                                for c in row} for ws in wb.worksheets}, "")
        finally:
            wb.close()


def python_values(cells: dict, formulas: dict) -> tuple[dict, list[dict]]:
    gaps: list[dict] = []
    values: dict[str, dict[str, object]] = {}
    for sheet, refs in formulas.items():
        ev = Evaluator(cells, sheet)
        out: dict[str, object] = {}
        for ref, formula in refs.items():
            try:
                out[ref] = ev.value_of(sheet, ref)
            except Unsupported as e:
                gaps.append({"class": "unsupported", "cell": f"{sheet}!{ref}",
                             "formula": formula, "why": str(e)})
            except ExcelError as e:
                out[ref] = e.token
            except Exception as e:  # noqa: BLE001 - an engine bug must not be a number
                gaps.append({"class": "unsupported", "cell": f"{sheet}!{ref}",
                             "formula": formula,
                             "why": f"evaluator failed: {type(e).__name__}: {e}"})
        values[sheet] = out
    return values, gaps


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in", dest="src", required=True, type=Path)
    ap.add_argument("--out", type=Path, help="write the workbook with cached values")
    ap.add_argument("--engine", choices=ENGINES, default="both")
    ap.add_argument("--timeout", type=int, default=180)
    ap.add_argument("--report", type=Path)
    ap.add_argument("--fail-on", default="",
                    help=f"comma-separated classes that make this exit 1 "
                         f"({', '.join(FAIL_CLASSES)})")
    args = ap.parse_args()

    state: dict = {}

    def entry():
        wanted = [c.strip() for c in args.fail_on.split(",") if c.strip()]
        unknown = [c for c in wanted if c not in FAIL_CLASSES]
        if unknown:
            fail(f"--fail-on: unknown class(es) {', '.join(unknown)}; "
                 f"choose from {', '.join(FAIL_CLASSES)}")
        if not args.src.is_file():
            fail(f"no such file: {args.src}")
        if args.out:
            ensure_distinct(args.src, args.out)

        pkg = Package.open(args.src)
        wb = Workbook(pkg)
        formulas = {name: wb.sheet(name).formula_cells() for name in wb.sheets}
        total = sum(len(f) for f in formulas.values())

        cells = read_cells(args.src)
        findings: list[dict] = []
        engines: dict[str, dict] = {}
        notes: list[str] = []

        if args.engine in ("both", "soffice"):
            if not find_soffice():
                if args.engine == "soffice":
                    fail("--engine soffice was requested but LibreOffice is not "
                         "installed; install it or use --engine python")
                notes.append("LibreOffice is not installed, so only the python "
                             "engine ran — nothing cross-checked it")
            else:
                vals, err = soffice_values(args.src, args.timeout)
                if err:
                    if args.engine == "soffice":
                        fail(err)
                    notes.append(f"LibreOffice failed ({err}); only the python "
                                 f"engine ran — nothing cross-checked it")
                else:
                    engines["soffice"] = vals

        if args.engine in ("both", "python"):
            vals, gaps = python_values(cells, formulas)
            engines["python"] = vals
            findings += gaps

        if not engines:
            fail("no engine produced a result")

        # Cross-check. Only cells BOTH engines produced a value for can disagree;
        # everything else is already reported as unsupported.
        agreed = disagreed = 0
        if len(engines) == 2:
            for sheet, refs in formulas.items():
                for ref in refs:
                    a = engines["soffice"].get(sheet, {}).get(ref)
                    b = engines["python"].get(sheet, {}).get(ref)
                    if sheet not in engines["python"] or ref not in engines["python"][sheet]:
                        continue
                    if same(a, b):
                        agreed += 1
                    else:
                        disagreed += 1
                        findings.append({
                            "class": "disagreement", "cell": f"{sheet}!{ref}",
                            "formula": refs[ref], "soffice": a, "python": b,
                            "why": "two independent engines computed different "
                                   "results; this script does not pick a winner"})

        # Which engine's numbers get written. LibreOffice is the authority when it
        # ran; python fills in only where it did not.
        chosen = "soffice" if "soffice" in engines else "python"
        written: list[dict] = []
        if args.out:
            for sheet, refs in formulas.items():
                ws = wb.sheet(sheet)
                for ref in refs:
                    v = engines[chosen].get(sheet, {}).get(ref)
                    if v is None and chosen == "soffice" and "python" in engines:
                        v = engines["python"].get(sheet, {}).get(ref)
                    if v is None:
                        continue
                    r = ws.set_cached(ref, v)
                    if r.get("written"):
                        written.append({"cell": f"{sheet}!{ref}", **r})
            wb.save(args.out)

        for sheet, refs in formulas.items():
            for ref in refs:
                v = engines[chosen].get(sheet, {}).get(ref)
                if isinstance(v, str) and v.startswith("#"):
                    findings.append({"class": "error", "cell": f"{sheet}!{ref}",
                                     "formula": refs[ref], "value": v,
                                     "why": "the formula evaluates to an error"})

        by_class = {c: sum(1 for f in findings if f["class"] == c)
                    for c in FAIL_CLASSES}
        report = {
            "file": args.src.name,
            "engines": sorted(engines),
            "values_written_from": chosen if args.out else None,
            "formulas": total,
            "cross_checked": agreed + disagreed,
            "agreed": agreed, "disagreed": disagreed,
            "supported_functions": sorted(SUPPORTED),
            "by_class": by_class,
            "findings": findings,
            "cells_written": len(written),
            # Said out loud rather than inferred from a missing key: a run with one
            # engine is not a cross-checked run, and the two must never look alike.
            "cross_checked_by_two_engines": len(engines) == 2,
            "notes": notes,
        }
        emit(report, args.report, "findings", "supported_functions")
        state["hit"] = sum(by_class[c] for c in wanted)

    rc = run(entry)
    return 1 if rc == 0 and state.get("hit") else rc


if __name__ == "__main__":
    raise SystemExit(main())
