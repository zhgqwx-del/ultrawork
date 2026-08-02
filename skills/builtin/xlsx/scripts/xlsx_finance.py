#!/usr/bin/env python3
"""X14 — the financial modelling colour convention: blue in, black out, green across.

    blue   a hard-coded number somebody typed
    black  a formula computed on this sheet
    green  a formula that reaches ANOTHER sheet

The point is auditability at a glance: in a model that follows it, every assumption
is visibly blue, and a blue cell where a formula belongs is a hard-coded override
somebody should know about.

    python3 xlsx_finance.py --in model.xlsx --check
    python3 xlsx_finance.py --in model.xlsx --out coloured.xlsx --apply

**This is opt-in and stays opt-in.** It is a convention of financial modelling, not
a property of spreadsheets — the repo's own L2 gate learned that the hard way, where
turning the same check on by default immediately reddened an ordinary test fixture
that was doing nothing wrong. So this is a separate command nobody runs by accident,
and `--check` on a workbook that never claimed to follow the convention will report
a great many findings, correctly.

`--apply` only ever changes the font COLOUR. Size, face, bold and every other
attribute are carried across, because a cell's typography is not this convention's
business.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from office.formula import references  # noqa: E402
from office.rebuild import RebuildError, rebuild  # noqa: E402
from xlsxcommon import XlsxError, emit, ensure_distinct, fail, run  # noqa: E402

BLUE, BLACK, GREEN = "FF0000FF", "FF000000", "FF008000"
ROLE_COLOUR = {"input": BLUE, "formula": BLACK, "link": GREEN}


def role_of(value, sheet: str) -> str | None:
    """input / formula / link, or None when the convention says nothing.

    Labels, dates and blanks are deliberately outside it: colouring them would make
    the signal noisier, not stronger.
    """
    if isinstance(value, str) and value.startswith("="):
        # A cross-sheet reference is `Sheet!A1`, not "contains a bang" — `=#REF!*2`
        # carries one too, and calling that a link is how the repo's L2 gate first
        # got this wrong.
        for ref in references(value):
            if ref.sheet and ref.sheet != sheet:
                return "link"
        return "formula"
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return "input"
    return None


def rgb_of(cell) -> str:
    c = cell.font.color if cell.font else None
    if c is None or c.type != "rgb" or not isinstance(c.rgb, str):
        return BLACK          # theme/indexed colours carry no rgb; default is black
    return c.rgb[-6:].upper().rjust(8, "F")[-8:] if len(c.rgb) >= 6 else BLACK


def scan(wb, only_sheet: str | None) -> list[dict]:
    out: list[dict] = []
    for ws in wb.worksheets:
        if only_sheet and ws.title != only_sheet:
            continue
        for row in ws.iter_rows():
            for cell in row:
                role = role_of(cell.value, ws.title)
                if role is None:
                    continue
                want = ROLE_COLOUR[role]
                got = rgb_of(cell)
                out.append({"cell": f"{ws.title}!{cell.coordinate}", "role": role,
                            "want": want, "got": got, "ok": got[-6:] == want[-6:]})
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in", dest="src", required=True, type=Path)
    ap.add_argument("--out", type=Path)
    ap.add_argument("--sheet", help="restrict to one sheet")
    ap.add_argument("--apply", action="store_true", help="recolour to the convention")
    ap.add_argument("--check", action="store_true", help="report violations only")
    ap.add_argument("--fail-on-violation", action="store_true",
                    help="exit 1 when --check finds any")
    ap.add_argument("--report", type=Path)
    args = ap.parse_args()

    state: dict = {}

    def entry():
        import openpyxl
        from contextlib import closing
        if not args.src.is_file():
            fail(f"no such file: {args.src}")
        if args.apply == args.check:
            fail("pass exactly one of --apply or --check")
        if args.apply and not args.out:
            fail("--apply needs --out")

        if args.check:
            wb = openpyxl.load_workbook(args.src)
            with closing(wb):
                cells = scan(wb, args.sheet)
            bad = [c for c in cells if not c["ok"]]
            by_role = {r: sum(1 for c in bad if c["role"] == r)
                       for r in ROLE_COLOUR}
            emit({"file": args.src.name, "convention": "blue input / black formula / "
                  "green cross-sheet link", "cells_in_scope": len(cells),
                  "violations": len(bad), "by_role": by_role,
                  # An empty scope is not a clean bill of health, and the two must
                  # not print the same way.
                  "note": ("no cell falls under this convention — nothing was checked"
                           if not cells else None),
                  "findings": bad}, args.report, "findings")
            state["bad"] = len(bad) if args.fail_on_violation else 0
            return

        ensure_distinct(args.src, args.out)
        changed: list[dict] = []

        def mutate(book):
            from openpyxl.styles import Font
            for ws in book.worksheets:
                if args.sheet and ws.title != args.sheet:
                    continue
                for row in ws.iter_rows():
                    for cell in row:
                        role = role_of(cell.value, ws.title)
                        if role is None:
                            continue
                        want = ROLE_COLOUR[role]
                        if rgb_of(cell)[-6:] == want[-6:]:
                            continue
                        base = cell.font
                        # Only the colour. Carrying the rest across is the whole
                        # reason this does not just assign a fresh Font().
                        cell.font = Font(name=base.name, size=base.size,
                                         bold=base.bold, italic=base.italic,
                                         underline=base.underline,
                                         strike=base.strike, vertAlign=base.vertAlign,
                                         color=want)
                        changed.append({"cell": f"{ws.title}!{cell.coordinate}",
                                        "role": role, "colour": want})
            return {"recoloured": changed}

        try:
            report = rebuild(args.src, args.out, mutate)
        except RebuildError as e:
            fail(str(e))
        emit({"in": args.src.name, "out": args.out.name, **report},
             args.report, "recoloured", "grafted")

    rc = run(entry)
    return 1 if rc == 0 and state.get("bad") else rc


if __name__ == "__main__":
    raise SystemExit(main())
