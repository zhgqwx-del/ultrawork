#!/usr/bin/env python3
"""Rebuild the xlsx skill's sample workbooks.

⚠️ Only run this when a fixture's CONTENT is meant to change. These files live under
`skills/builtin/`, whose hash is the sentinel every desktop client compares against
to decide whether to reinstall the builtin skills — a byte that moves for no reason
makes every installed client redownload.

Which is why the output is byte-reproducible, and it takes work to be so: openpyxl
writes every zip entry with the current timestamp AND stamps the save time into
`docProps/core.xml` over whatever `wb.properties.modified` was set to, so two runs
a second apart produce different bytes for identical content. `normalize()` below
pins all three sources of drift.

    python3 fixtures/make_fixtures.py
"""
from __future__ import annotations

import datetime as dt
import re
import shutil
import zipfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
EPOCH = dt.datetime(2026, 8, 1, 0, 0, 0)
ZIP_DATE = (2026, 8, 1, 0, 0, 0)

# A part class openpyxl has no model for. Real Office files are full of these —
# customXml (content-control data), xl/metadata.xml (dynamic arrays), threaded
# comments, pivot caches — and they are what a load→save round trip silently loses.
CUSTOM_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<contract xmlns="urn:ultrawork:fixture"><id>Q3-2026-0417</id><owner>财务部</owner>\
<note>这是 openpyxl 不认识的自定义 part，round-trip 会把它整个丢掉</note></contract>
"""
CUSTOM_XML_PROPS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<ds:datastoreItem xmlns:ds="http://schemas.openxmlformats.org/officeDocument/2006/\
customXml" ds:itemID="{6C4A9F1E-0000-4E2B-9A11-0D4E2C7B8A31}">\
<ds:schemaRefs><ds:schemaRef ds:uri="urn:ultrawork:fixture"/></ds:schemaRefs>\
</ds:datastoreItem>
"""
CUSTOM_XML_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">\
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/\
relationships/customXmlProps" Target="itemProps1.xml"/></Relationships>
"""

DATA_FONT = "宋体"
DATA_FONT_SIZE = 12

INCOME_ROWS = [
    ("营业收入", 1240, 1103),
    ("营业成本", 769, 702),
]


def build_book(path: Path) -> None:
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "利润表"
    ws["A1"] = "季度经营分析报告"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center")
    for col, head in zip("ABCD", ("科目", "本季度", "上年同期", "同比")):
        ws[f"{col}2"] = head
        ws[f"{col}2"].font = Font(bold=True)
    for i, (name, now, prev) in enumerate(INCOME_ROWS, start=3):
        ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"] = name, now, prev
        ws[f"D{i}"] = f"=B{i}/C{i}-1"
        ws[f"D{i}"].number_format = "0.0%"
        # A DELIBERATELY non-default face and size on the number cells. Without it,
        # "changing the colour resets the font" is invisible: assigning a fresh
        # Font() lands on Calibri 11, which is what the cell already had, and the
        # negative control for that defect cannot fire. Asserted by V0.
        for col in "BC":
            ws[f"{col}{i}"].font = Font(name=DATA_FONT, size=DATA_FONT_SIZE)
    ws["A5"] = "毛利"
    ws["B5"], ws["C5"] = "=B3-B4", "=C3-C4"
    ws["D5"] = "=B5/C5-1"
    ws["D5"].number_format = "0.0%"

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:D5"
    ws.conditional_formatting.add(
        "D3:D5", CellIsRule(operator="lessThan", formula=["0"],
                            font=Font(color="FF9C0006")))
    ws.add_data_validation(dv := DataValidation(type="decimal", operator="greaterThan",
                                                formula1="0"))
    dv.add("B3:C4")
    # Explicit widths, counted the way X15 counts them: 营业收入 is 4 CJK characters
    # = 8 display units, not 4.
    for col, width in (("A", 14), ("B", 12), ("C", 12), ("D", 10)):
        ws.column_dimensions[col].width = width

    summary = wb.create_sheet("汇总")
    summary["A1"] = "指标"
    summary["B1"] = "取值"
    summary["A2"] = "毛利"
    summary["B2"] = "=利润表!B5"          # cross-sheet reference (X10)
    summary["A3"] = "毛利率"
    summary["B3"] = "=B2/利润表!B3"
    summary["B3"].number_format = "0.0%"
    for col, width in (("A", 14), ("B", 12)):
        summary.column_dimensions[col].width = width

    chart = BarChart()
    chart.title = "本季度 vs 上年同期"
    chart.add_data(Reference(ws, min_col=2, max_col=3, min_row=2, max_row=4),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=3, max_row=4))
    summary.add_chart(chart, "D2")

    wb.properties.created = EPOCH
    wb.properties.modified = EPOCH
    wb.properties.creator = "ultrawork xlsx skill fixtures"
    wb.properties.lastModifiedBy = "ultrawork xlsx skill fixtures"
    wb.save(path)
    wb.close()
    add_custom_xml(path)


def add_custom_xml(path: Path) -> None:
    """Wire a customXml pair into an existing package.

    Written by hand rather than by a library on purpose: the point of these parts is
    that no library in this stack models them, so a library cannot be asked to
    produce one.
    """
    import sys
    sys.path.insert(0, str(HERE.parent / "scripts"))
    from office.package import Package

    pkg = Package.open(path)
    pkg.write("customXml/item1.xml", CUSTOM_XML.encode("utf-8"))
    pkg.write("customXml/itemProps1.xml", CUSTOM_XML_PROPS.encode("utf-8"))
    pkg.write("customXml/_rels/item1.xml.rels", CUSTOM_XML_RELS.encode("utf-8"))
    pkg.set_override("customXml/itemProps1.xml",
                     "application/vnd.openxmlformats-officedocument.customXmlProperties+xml")
    pkg.add_relationship(
        "_rels/.rels",
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXml",
        "customXml/item1.xml")
    pkg.save(path)


def build_narrow(src: Path, dst: Path) -> None:
    """The same workbook with every explicit column width removed.

    This is X15's input and it is not a strawman: it is the state every workbook
    written by a library that does not think about width is in, including the ones
    the old doc-edit skill produced. 营业收入合计 in an 8.43-unit column is cut off
    on screen while being perfectly present in the file.
    """
    import sys
    sys.path.insert(0, str(HERE.parent / "scripts"))
    from office.package import Package
    from office.xmlorder import q

    pkg = Package.open(src)
    for name in [n for n in pkg.names() if n.startswith("xl/worksheets/sheet")]:
        root = pkg.tree(name)
        cols = root.find(q("cols"))
        if cols is not None:
            root.remove(cols)
        pkg.put_tree(name, root)
    pkg.save(dst)
    # One label long enough that the 8.43 default is visibly not enough for it.
    from office.sheet import Workbook
    wb = Workbook(Package.open(dst))
    wb.sheet("利润表").set_cell("A6", "营业收入合计（含其他业务）")
    wb.save(dst)


MODIFIED = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def normalize(path: Path) -> None:
    """Make the bytes a function of the content only.

    Three sources of drift, all found by running this twice and diffing rather than
    by reasoning about it:

      * zip entry timestamps — openpyxl writes "now" for every entry
      * the compression level
      * `dcterms:modified` — openpyxl stamps the save time over whatever
        `wb.properties.modified` was set to, so pinning the property is not enough

    Re-running the script without changing anything must leave the files
    byte-identical, or the builtin-skills sentinel moves and every installed client
    reinstalls.
    """
    with zipfile.ZipFile(path) as z:
        entries = [(i.filename, z.read(i.filename)) for i in z.infolist()
                   if not i.filename.endswith("/")]
    stamp = EPOCH.strftime("%Y-%m-%dT%H:%M:%SZ").encode()
    entries = [(n, MODIFIED.sub(rb"\g<1>" + stamp + rb"\g<2>", d)
                if n == "docProps/core.xml" else d) for n, d in entries]
    tmp = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as z:
        for name, data in entries:
            info = zipfile.ZipInfo(name, date_time=ZIP_DATE)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            # `ZipInfo.__init__` sets create_system = 0 on Windows and 3 elsewhere,
            # and that byte lands in the central directory — so the same generator
            # emits different bytes there. Found by the docx fixtures' F0 assertion
            # on CI; this generator has the same gap and no assertion that would
            # ever fire on it, which is the more dangerous of the two. Pinned to 3,
            # what macOS and Linux already write, so committed bytes do not change.
            info.create_system = 3
            z.writestr(info, data)
    shutil.move(str(tmp), str(path))


def main() -> None:
    book = HERE / "book.xlsx"
    narrow = HERE / "narrow.xlsx"
    build_book(book)
    build_narrow(book, narrow)
    for f in (book, narrow):
        normalize(f)
        print(f"wrote {f.name} ({f.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
