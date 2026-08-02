#!/usr/bin/env python3
"""X8 — add a chart to a worksheet.

Same path as `xlsx_format.py`: openpyxl builds the chart, drawing and rels graph,
and `office/rebuild.py` puts back the parts the rebuild dropped.

    python3 xlsx_chart.py --in book.xlsx --out out.xlsx --sheet 汇总 \
            --type bar --data 利润表!B2:C4 --categories 利润表!A3:A4 \
            --anchor D18 --title "本季度 vs 上年同期"

`--data` includes the HEADER row by default (`--no-headers` turns that off), because
a chart whose series are called "Series1" and "Series2" is a chart nobody can read,
and the header is where the real names already are.

Ranges are `Sheet!A1:B2`. The sheet holding the DATA and the sheet holding the CHART
are deliberately separate arguments — putting a summary chart on a different sheet
from its numbers is the normal case, not the exotic one.
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from office.formula import quote_sheet, unquote_sheet  # noqa: E402
from office.rebuild import RebuildError, rebuild  # noqa: E402
from xlsxcommon import XlsxError, emit, ensure_distinct, fail, run  # noqa: E402

CHART_TYPES = ("bar", "column", "line", "pie", "scatter", "area")
RANGE = re.compile(r"^(?:(?P<sheet>'(?:[^']|'')+'|[^!]+)!)?"
                   r"(?P<a>\$?[A-Za-z]{1,3}\$?[0-9]+):(?P<b>\$?[A-Za-z]{1,3}\$?[0-9]+)$")


def split_range(spec: str, default_sheet: str, what: str) -> tuple[str, str]:
    m = RANGE.match(spec.strip())
    if not m:
        fail(f"{what} {spec!r} is not a range (expected e.g. 利润表!B2:C4)")
    sheet = unquote_sheet(m.group("sheet")) or default_sheet
    return sheet, f"{m.group('a')}:{m.group('b')}"


def boundaries(area: str) -> tuple[int, int, int, int]:
    from openpyxl.utils.cell import range_boundaries
    c0, r0, c1, r1 = range_boundaries(area)
    return c0, r0, c1, r1


def make_chart(kind: str, title: str | None):
    from openpyxl.chart import (
        AreaChart, BarChart, LineChart, PieChart, ScatterChart,
    )
    if kind in ("bar", "column"):
        chart = BarChart()
        # openpyxl's "bar" is horizontal and "col" vertical; the word people mean by
        # "bar chart" is almost always the vertical one, so `column` is explicit and
        # `bar` follows the common expectation rather than the library's naming.
        chart.type = "col" if kind == "column" else "col"
    elif kind == "line":
        chart = LineChart()
    elif kind == "pie":
        chart = PieChart()
    elif kind == "scatter":
        chart = ScatterChart()
    else:
        chart = AreaChart()
    if title:
        chart.title = title
    return chart


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in", dest="src", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    ap.add_argument("--sheet", required=True, help="sheet the chart is placed on")
    ap.add_argument("--type", dest="kind", choices=CHART_TYPES, default="bar")
    ap.add_argument("--data", required=True, metavar="利润表!B2:C4")
    ap.add_argument("--categories", metavar="利润表!A3:A4")
    ap.add_argument("--anchor", default="H2", metavar="H2")
    ap.add_argument("--title")
    ap.add_argument("--no-headers", action="store_true",
                    help="the first row of --data is data, not series names")
    ap.add_argument("--width", type=float, default=15.0)
    ap.add_argument("--height", type=float, default=7.5)
    ap.add_argument("--report", type=Path)
    args = ap.parse_args()

    def entry():
        from openpyxl.chart import Reference
        ensure_distinct(args.src, args.out)
        info: dict = {}

        def mutate(wb):
            if args.sheet not in wb.sheetnames:
                fail(f"no sheet named {args.sheet!r} (have: {', '.join(wb.sheetnames)})")
            target = wb[args.sheet]
            data_sheet, data_area = split_range(args.data, args.sheet, "--data")
            if data_sheet not in wb.sheetnames:
                fail(f"--data names sheet {data_sheet!r}, which this workbook does "
                     f"not have (sheets: {', '.join(wb.sheetnames)})")
            src_ws = wb[data_sheet]
            c0, r0, c1, r1 = boundaries(data_area)
            chart = make_chart(args.kind, args.title)
            chart.width, chart.height = args.width, args.height
            chart.add_data(Reference(src_ws, min_col=c0, min_row=r0,
                                     max_col=c1, max_row=r1),
                           titles_from_data=not args.no_headers)
            cats = None
            if args.categories:
                cat_sheet, cat_area = split_range(args.categories, args.sheet,
                                                  "--categories")
                if cat_sheet not in wb.sheetnames:
                    fail(f"--categories names sheet {cat_sheet!r}, which this "
                         f"workbook does not have")
                k0, q0, k1, q1 = boundaries(cat_area)
                cats = Reference(wb[cat_sheet], min_col=k0, min_row=q0,
                                 max_col=k1, max_row=q1)
                chart.set_categories(cats)
            before = len(target._charts)
            target.add_chart(chart, args.anchor)
            info.update({
                "chart": {"type": args.kind, "on": args.sheet, "anchor": args.anchor,
                          "data": f"{quote_sheet(data_sheet)}!{data_area}",
                          "categories": args.categories,
                          "series_named_from_header": not args.no_headers,
                          "charts_on_sheet_before": before,
                          "charts_on_sheet_after": before + 1},
            })
            return info

        try:
            report = rebuild(args.src, args.out, mutate)
        except RebuildError as e:
            fail(str(e))
        emit({"in": args.src.name, "out": args.out.name, **report},
             args.report, "grafted")

    return run(entry)


if __name__ == "__main__":
    raise SystemExit(main())
