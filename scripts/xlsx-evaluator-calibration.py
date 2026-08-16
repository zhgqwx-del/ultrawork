#!/usr/bin/env python3
"""The measured coverage boundary of the pure-Python formula evaluator (059 §7 X4).

Every row below is GROUND TRUTH taken from LibreOffice, not from anyone's memory of
what Excel does. Six of them contradicted the first implementation, and each would
have shipped as a confident wrong number:

    -2^2        4  (unary minus binds TIGHTER than ^)      first draft: -4
    2^3^2      64  (^ is LEFT-associative)                 first draft: 512
    MOD(-7,3)   2  (sign of the DIVISOR, i.e. Python %)    first draft: -1  (fmod)
    IF(TRUE,,5) 0  (an empty argument is blank)            first draft: parse error
    SUM()       0  (no arguments at all)                   first draft: parse error
    AVERAGE(1,"x",2)  #VALUE!  (direct text ≠ text in a range)   first draft: 1.5

    python3 scripts/xlsx-evaluator-calibration.py            # check both engines
    python3 scripts/xlsx-evaluator-calibration.py --emit     # re-measure via soffice

`--emit` reprints this table from a live LibreOffice run. Use it when adding a
function; paste the output back here. Never hand-edit an expected value — that is
exactly how the L2 gate carried a wrong X3 expectation for a month, unnoticed
because the only machine that could disprove it had no LibreOffice.

This module is imported by scripts/test-xlsx-skill.py, which asserts:
  * the python engine matches every row (runs everywhere, including CI)
  * LibreOffice matches every row too, when present (catches a wrong pin)
  * every function named in the evaluator's SUPPORTED set is exercised here
  * a function NOT in SUPPORTED is refused, never silently miscomputed
"""
from __future__ import annotations

import argparse
import subprocess
import sys
import tempfile
from pathlib import Path

# ── stdout must be UTF-8 on every platform, and on Windows it is not ──────────
# This gate prints ✅/❌ and Chinese. Windows encodes a CAPTURED stdout in the
# machine's ANSI code page and Python only defaults to UTF-8 from 3.15 (PEP 686);
# CI pins 3.11. Measured on CI: this script died with
# `UnicodeEncodeError: 'charmap' codec can't encode character '\u2705'` inside its
# own `print(json.dumps(...))`. The skills were fixed for this first and the GATES
# were missed — the same defect has two homes, and only one of them was product.
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        try:
            _stream.reconfigure(encoding="utf-8")
        except (ValueError, OSError):        # already detached / not reconfigurable
            pass


REPO = Path(__file__).resolve().parent.parent
SKILL = REPO / "skills" / "builtin" / "xlsx"

# The inputs every probe below references.
DATA = {"H1": 10, "H2": 20, "H3": 30, "H5": "abc", "H6": 0,
        "I1": 1.5, "I2": -2.5, "I3": 2.5, "J1": "  x  y  ", "J2": True}

# (formula, LibreOffice's answer). Measured 2026-08-02, LibreOffice 26.2.5.2.
CALIBRATION = [
    ("=SUM(H1:H3)", 60), ("=SUM(H1,H2)", 30), ("=SUM()", 0),
    ("=PRODUCT(H1:H3)", 6000), ("=AVERAGE(H1:H3)", 20),
    ("=MIN(H1:H3)", 10), ("=MAX(H1:H3)", 30),
    ("=COUNT(H1:H5)", 3), ("=COUNTA(H1:H5)", 4),
    ("=ABS(I2)", 2.5), ("=INT(I1)", 1), ("=INT(I2)", -3), ("=SQRT(H1*10)", 10),
    ("=ROUND(I3,0)", 3), ("=ROUND(I2,0)", -3), ("=ROUND(1.2345,2)", 1.23),
    ("=ROUNDUP(1.234,1)", 1.3), ("=ROUNDDOWN(1.789,1)", 1.7),
    ("=ROUNDUP(-1.234,1)", -1.3), ("=POWER(2,10)", 1024),
    ("=MOD(7,3)", 1), ("=MOD(-7,3)", 2), ("=MOD(7,-3)", -2),
    ('=IF(H1>5,"big","small")', "big"), ("=IF(H1>50,1,2)", 2), ("=IF(TRUE,,5)", 0),
    ("=AND(H1>5,H2>5)", True), ("=OR(H1>50,H2>5)", True), ("=NOT(H1>50)", True),
    ("=LEN(H5)", 3), ("=UPPER(H5)", "ABC"), ('=LOWER("ABC")', "abc"),
    ("=TRIM(J1)", "x y"), ("=LEFT(H5,2)", "ab"), ("=RIGHT(H5,2)", "bc"),
    ("=MID(H5,2,2)", "bc"), ("=CONCATENATE(H5,H1)", "abc10"),
    ('=H5&"-"&H1', "abc-10"), ("=H1/H6", "#DIV/0!"),
    ('=IFERROR(H1/H6,"na")', "na"), ("=IFERROR(H1/H2,0)", 0.5),
    ("=-2^2", 4), ("=2^3^2", 64), ("=2^-1", 0.5), ("=10%", 0.1), ("=1+2%", 1.02),
    ("=3-2^2", -1), ("=H1=10", True), ("=H1<>10", False), ("=H1>=10", True),
    ("=H1<H2", True), ('="a"<"b"', True), ("=H1&H2", "1020"),
    ("=SUM(H1:H3)/COUNT(H1:H3)", 20), ("=AVERAGE(H4:H4)", "#DIV/0!"),
]

# Constructs the evaluator must REFUSE. Being on this list is a claim that the
# engine says "I cannot" — not that it returns something plausible. Anything here
# that starts quietly producing a number is a regression, and the gate says so.
MUST_REFUSE = [
    ("=VLOOKUP(H1,H1:I3,2,0)", "the lookup family is not implemented"),
    ("=INDEX(H1:H3,2)", "the lookup family is not implemented"),
    ("=TODAY()", "volatile: a cross-check against a moving value is meaningless"),
    ("=SUMIF(H1:H3,\">15\")", "criteria syntax is not implemented"),
    ('=AVERAGE(1,"x",2)', "direct text argument: Excel's coercion differs from ranges"),
    ('=COUNT(1,"2","x")', "direct text argument: Excel's coercion differs from ranges"),
    ("=SQRT(-1)", "Excel says #NUM!, LibreOffice says #VALUE! — no authority to follow"),
    ("=H1:H3", "a bare range as the whole result"),
    ("=MyNamedRange*2", "defined names are not resolved"),
]


def unexercised() -> list[str]:
    """Functions the evaluator CLAIMS to support that this corpus never runs.

    Without this, SUPPORTED is a wish list: a name can sit in it, be exercised by
    nothing, and still read as "measured". Same shape as L1's C5 — declared but
    unbacked.
    """
    sys.path.insert(0, str(SKILL / "scripts"))
    from office.evaluate import SUPPORTED
    corpus = " ".join(f for f, _ in CALIBRATION) + " " + \
             " ".join(f for f, _ in MUST_REFUSE)
    return sorted(name for name in SUPPORTED if f"{name}(" not in corpus.upper())


def norm(v):
    """Compare the way a spreadsheet does: blank is 0, ints and floats are equal."""
    if v is None:
        return 0
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return round(float(v), 9)
    return v


def python_results() -> dict[str, object]:
    sys.path.insert(0, str(SKILL / "scripts"))
    from office.evaluate import Evaluator, ExcelError, Unsupported
    ev = Evaluator({"C": dict(DATA)}, "C")
    out: dict[str, object] = {}
    for formula, _ in CALIBRATION:
        try:
            out[formula] = ev.evaluate(formula)
        except Unsupported as e:
            out[formula] = f"UNSUPPORTED: {e}"
        except ExcelError as e:
            out[formula] = e.token
        except Exception as e:  # noqa: BLE001
            out[formula] = f"CRASH: {type(e).__name__}: {e}"
    return out


def python_refusals() -> dict[str, object]:
    sys.path.insert(0, str(SKILL / "scripts"))
    from office.evaluate import Evaluator, ExcelError, Unsupported
    ev = Evaluator({"C": dict(DATA)}, "C")
    out: dict[str, object] = {}
    for formula, _ in MUST_REFUSE:
        try:
            out[formula] = ("VALUE", ev.evaluate(formula))
        except Unsupported as e:
            out[formula] = ("REFUSED", str(e))
        except ExcelError as e:
            out[formula] = ("ERROR", e.token)
        except Exception as e:  # noqa: BLE001
            out[formula] = ("CRASH", f"{type(e).__name__}: {e}")
    return out


def soffice_results(timeout: int = 300) -> tuple[dict[str, object], str]:
    """Ask LibreOffice for the same table. Returns ({}, reason) when unavailable."""
    sys.path.insert(0, str(SKILL / "scripts"))
    from office.soffice import convert, find_soffice
    if not find_soffice():
        return {}, "LibreOffice is not installed on this host"
    import openpyxl
    with tempfile.TemporaryDirectory(prefix="calib-") as td:
        book = Path(td) / "calib.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "C"
        for ref, v in DATA.items():
            ws[ref] = v
        for i, (formula, _) in enumerate(CALIBRATION, 1):
            ws.cell(row=i, column=2, value=formula)
        wb.save(book)
        wb.close()
        out, err = convert(book, "xlsx", Path(td) / "out", timeout=timeout)
        if out is None:
            return {}, err
        got = openpyxl.load_workbook(out, data_only=True)["C"]
        return ({formula: got.cell(row=i, column=2).value
                 for i, (formula, _) in enumerate(CALIBRATION, 1)}, "")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--emit", action="store_true",
                    help="reprint CALIBRATION from a live LibreOffice run")
    args = ap.parse_args()

    if args.emit:
        vals, err = soffice_results()
        if err:
            print(f"[error] {err}", file=sys.stderr)
            return 2
        print("CALIBRATION = [")
        for formula, _ in CALIBRATION:
            print(f"    ({formula!r}, {vals[formula]!r}),")
        print("]")
        return 0

    py = python_results()
    bad = [(f, want, py[f]) for f, want in CALIBRATION if norm(py[f]) != norm(want)]
    for f, want, got in bad:
        print(f"FAIL  {f:<30} pinned={want!r}  python={got!r}")
    refusals = python_refusals()
    leaked = [(f, why, refusals[f]) for f, why in MUST_REFUSE
              if refusals[f][0] != "REFUSED"]
    for f, why, got in leaked:
        print(f"FAIL  {f:<30} must be refused ({why}) — got {got!r}")

    missing = unexercised()
    if missing:
        print(f"FAIL  the evaluator claims to support {', '.join(missing)}, and this "
              f"corpus never exercises them — SUPPORTED would be a wish list")

    lo, err = soffice_results()
    drift = []
    if err:
        print(f"\n[skipped] LibreOffice half: {err}")
        print("  Skipped is not green — nothing checked the pinned values this run.")
    else:
        drift = [(f, want, lo[f]) for f, want in CALIBRATION
                 if norm(lo[f]) != norm(want)]
        for f, want, got in drift:
            print(f"FAIL  {f:<30} pinned={want!r}  LibreOffice={got!r}  <- the PIN is wrong")

    total = len(CALIBRATION) + len(MUST_REFUSE)
    failed = len(bad) + len(leaked) + len(drift) + len(missing)
    print(f"\n[calibration] {total - failed} passed, {failed} failed "
          f"({len(CALIBRATION)} pinned values, {len(MUST_REFUSE)} must-refuse)"
          + ("" if err else " · LibreOffice agreed with every pin"))
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
