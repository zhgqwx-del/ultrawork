#!/usr/bin/env python3
"""X13 — render a workbook to PDF (and optionally to page images).

This is not a nicety. **A .xlsx cannot be previewed inside ultrawork** — the
artifact panel renders PDF and shows a binary info card for everything else — so
converting to PDF is the only way a workbook this skill produced becomes something
the user can actually look at without leaving the app (059 §7).

    python3 xlsx_pdf.py --in book.xlsx --out book.pdf
    python3 xlsx_pdf.py --in book.xlsx --out book.pdf --png ./pages --dpi 150
    python3 xlsx_pdf.py --in book.xlsx --out book.pdf --sheet 利润表

Requires LibreOffice, which is a declared dependency of this skill. There is no
pure-Python fallback and pretending otherwise would mean inventing a layout engine:
column widths, page breaks, print areas and number formats all have to be resolved
the way a spreadsheet application resolves them.

Two things it refuses rather than papers over:

  * **a PDF with no ink on the first page.** LibreOffice exits 0 for an empty
    sheet and produces a blank page; handing that back as "your preview" is worse
    than an error, because it looks like the data is gone.
  * **a workbook whose formulas have never been calculated.** Those cells render
    EMPTY. The file looks wrong and the cause is invisible in the picture.
"""
from __future__ import annotations

import argparse
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from office.soffice import convert, find_soffice  # noqa: E402
from xlsxcommon import (  # noqa: E402
    XlsxError, display_width, displayed_text, emit, ensure_distinct, fail, run,
)

# Share of dark pixels below which a page counts as blank. Calibrated against the
# same raster the repo's L2 gate uses, which was itself standardised after a
# synthetic-only threshold misfired on real output (gotchas §10⑭).
BLANK_INK = 0.0005
BLANK_DPI = 100


def hidden_sheets(path: Path) -> list[str]:
    import openpyxl
    from contextlib import closing
    wb = openpyxl.load_workbook(path, read_only=True)
    with closing(wb):
        return [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]


def uncalculated(path: Path) -> int:
    import openpyxl
    from contextlib import closing
    f = openpyxl.load_workbook(path, read_only=True, data_only=False)
    v = openpyxl.load_workbook(path, read_only=True, data_only=True)
    n = 0
    with closing(f), closing(v):
        for name in f.sheetnames:
            for frow, vrow in zip(f[name].iter_rows(values_only=True),
                                  v[name].iter_rows(values_only=True)):
                for a, b in zip(frow, vrow):
                    if isinstance(a, str) and a.startswith("=") and b is None:
                        n += 1
    return n


def only_sheet(src: Path, name: str, workdir: Path) -> Path:
    """A copy with every other sheet removed, so --sheet means one sheet."""
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from office.package import Package
    import openpyxl
    from contextlib import closing
    wb = openpyxl.load_workbook(src)
    with closing(wb):
        if name not in wb.sheetnames:
            fail(f"no sheet named {name!r} (have: {', '.join(wb.sheetnames)})")
        for other in [s for s in wb.sheetnames if s != name]:
            del wb[other]
        trimmed = workdir / src.name
        wb.save(trimmed)
    return trimmed


def page_ink(page) -> float:
    """Share of dark pixels on one page, rendered grey at BLANK_DPI.

    `stride` is the row length in bytes and is NOT always the width — PDFium pads
    rows — so the rows are counted individually. Treating the buffer as one flat
    w*h run divides by a length that includes padding bytes and quietly reports
    less ink than there is, which on this threshold is the difference between
    "blank" and "fine".
    """
    bitmap = page.render(scale=BLANK_DPI / 72.0, grayscale=True)
    buf, stride, width, height = (bitmap.buffer, bitmap.stride, bitmap.width,
                                  bitmap.height)
    if not width or not height:
        return 0.0
    dark = bytes(1 if v < 140 else 0 for v in range(256))
    hits = sum(bytes(buf[r * stride:r * stride + width]).translate(dark).count(1)
               for r in range(height))
    return hits / (width * height)


def header_labels(src: Path, sheet: str | None) -> list[str]:
    """The text in row 1 of the sheet being rendered — the column labels.

    Only meaningful for a single sheet with a text header row; anything else returns
    nothing and the split check below stays silent rather than guessing.
    """
    import openpyxl
    from contextlib import closing
    book = openpyxl.load_workbook(src, read_only=True, data_only=True)
    with closing(book):
        names = [sheet] if sheet else [ws.title for ws in book.worksheets
                                       if ws.sheet_state == "visible"]
        if len(names) != 1:
            return []
        rows = book[names[0]].iter_rows(min_row=1, max_row=1, values_only=True)
        first = next(iter(rows), ()) or ()
        return [v for v in first if isinstance(v, str) and v.strip()]


def split_columns(doc, src: Path, sheet: str | None) -> list[str]:
    """Column labels that did NOT make it onto page 1 — i.e. the table was cut.

    Row 1 is always on the first page vertically, so a header that turns up later
    can only mean the sheet was too wide for the paper and the columns to its right
    were moved to a page of their own. Measured 2026-08-16 (059 §三十二): widening
    five columns by hand did exactly that to 利润表 — page 2 held one column of
    eight percentages with no row labels beside them — and nothing said so.
    """
    labels = header_labels(src, sheet)
    if len(doc) < 2 or not labels:
        return []
    front = doc[0].get_textpage().get_text_range()
    return [label for label in labels if label not in front]


def too_narrow_columns(src: Path, sheet: str | None) -> list[str]:
    """Columns whose widest displayed value does not fit the width in the file.

    Derived, not observed — `###` on the page is the observation. This names the
    likely culprits so the answer is actionable, using the same display-width and
    number-format rendering `--autofit` uses, so the two can never disagree.

    The `+ 1` is measured, not chosen: rendering `108.8%` (six display units) at
    column widths 6, 7 and 8 gives ###, fine, fine (2026-08-16, LibreOffice). One
    unit of inset is what the cell takes; `--autofit` asks for two, and that extra
    unit is deliberate slack rather than a disagreement with this.
    """
    inset = 1
    import openpyxl
    from contextlib import closing
    from office.sheet import index_to_col
    book = openpyxl.load_workbook(src, data_only=False)
    values = openpyxl.load_workbook(src, data_only=True)
    guilty: list[str] = []
    with closing(book), closing(values):
        for ws in book.worksheets:
            if sheet and ws.title != sheet:
                continue
            widths = {c: d.width for c, d in ws.column_dimensions.items() if d.width}
            worst: dict[int, float] = {}
            for row, vrow in zip(ws.iter_rows(),
                                 values[ws.title].iter_rows(values_only=True)):
                for cell, cached in zip(row, vrow):
                    v = cell.value
                    if isinstance(v, str) and not v.startswith("="):
                        continue          # text spills into the next cell, not ###
                    shown = displayed_text(cached, cell.number_format)
                    if not shown:
                        continue
                    worst[cell.column] = max(worst.get(cell.column, 0.0),
                                             display_width(shown))
            for col, need in sorted(worst.items()):
                letter = index_to_col(col)
                have = widths.get(letter)
                if have is not None and have + 1e-6 < need + inset:
                    guilty.append(f"{ws.title}!{letter}")
    return guilty


def hash_marks(doc, src: Path, sheet: str | None) -> dict:
    """Did a number come out as `###` — the width failure that looks like content?

    `blank_pages` catches a page with nothing on it and `columns_off_first_page`
    catches a table cut in half, but a single cell replaced by hash marks passes
    both while hiding the number the reader came for. Measured 2026-08-16 (059
    §三十三): the B9 deliverable rendered 营业利润's 同比 — 108.8% — as `###`, and
    every field in the report said the preview was fine.
    """
    text = "\n".join(doc[i].get_textpage().get_text_range() for i in range(len(doc)))
    if "###" not in text:
        return {"hash_marked_cells": False}
    import openpyxl
    from contextlib import closing
    book = openpyxl.load_workbook(src, read_only=True, data_only=True)
    with closing(book):
        literal = any("###" in v for ws in book.worksheets
                      for row in ws.iter_rows(values_only=True) for v in row
                      if isinstance(v, str))
    if literal:
        # A sheet that really contains "###" makes the page ambiguous, and guessing
        # which one this is would be worse than saying it cannot be told.
        return {"hash_marked_cells": None,
                "hash_marked_note": "the page contains ### and so does the sheet as "
                                    "literal text — this check cannot tell a "
                                    "truncated number from real content here"}
    return {"hash_marked_cells": True,
            "hash_marked_columns": too_narrow_columns(src, sheet)}


def inspect_pdf(pdf: Path, png_dir: Path | None, dpi: int,
                src: Path | None = None, sheet: str | None = None) -> dict:
    try:
        import pypdfium2 as pdfium
    except ImportError:
        # pypdfium2 belongs to the pdf skill and is NOT a declared dependency here —
        # putting a red badge on the xlsx skill for a library only the preview extras
        # need would be wrong. So the PDF is still produced and the blank-page check
        # degrades to a stated gap. But --png was asked for explicitly: quietly
        # producing no images would be the silent failure this file exists to avoid.
        if png_dir is not None:
            fail("--png needs pypdfium2 to rasterize the pages, and it is not "
                 "installed (pip install pypdfium2). The PDF itself does not need it")
        return {"pages": None, "blank_pages": None, "images": [],
                "note": "pypdfium2 is not installed, so the blank-page check did NOT "
                        "run — this PDF has not been checked for empty pages "
                        "(pip install pypdfium2)"}
    out: dict = {"images": []}
    doc = pdfium.PdfDocument(str(pdf))
    try:
        out["pages"] = len(doc)
        out["blank_pages"] = [i + 1 for i in range(len(doc))
                              if page_ink(doc[i]) < BLANK_INK]
        if src is not None:
            out["columns_off_first_page"] = split_columns(doc, src, sheet)
        if src is not None:
            out.update(hash_marks(doc, src, sheet))
        if png_dir is not None:
            png_dir.mkdir(parents=True, exist_ok=True)
            for i in range(len(doc)):
                img = png_dir / f"page-{i + 1:03d}.png"
                doc[i].render(scale=dpi / 72.0).to_pil().save(str(img))
                out["images"].append(str(img))
            # A shorter document rendered into the same directory leaves the tail of
            # the previous one behind, under names that read as part of THIS render.
            # Measured 2026-08-16 (059 §三十三): a 1-page preview landed in a folder
            # holding page-002.png from a 2-page one, and nothing said so. Only the
            # names this script itself produces are removed.
            stale = sorted(p.name for p in png_dir.glob("page-*.png")
                           if str(p) not in out["images"])
            for name in stale:
                (png_dir / name).unlink()
            if stale:
                out["stale_images_removed"] = stale
    finally:
        doc.close()
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in", dest="src", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    ap.add_argument("--sheet", help="render only this sheet")
    ap.add_argument("--png", type=Path, metavar="DIR", help="also write page images")
    ap.add_argument("--dpi", type=int, default=150)
    ap.add_argument("--timeout", type=int, default=180)
    ap.add_argument("--allow-blank", action="store_true",
                    help="accept a PDF whose pages carry no ink")
    ap.add_argument("--report", type=Path)
    args = ap.parse_args()

    def entry():
        if not args.src.is_file():
            fail(f"no such file: {args.src}")
        replaced = ensure_distinct(args.src, args.out)
        if not find_soffice():
            fail("LibreOffice not found. It is a required dependency of this skill — "
                 "rendering a spreadsheet means resolving column widths, page breaks "
                 "and number formats the way a spreadsheet application does, and "
                 "there is no pure-Python substitute. Install from libreoffice.org")
        if args.dpi < 36 or args.dpi > 600:
            fail(f"--dpi {args.dpi} is outside 36-600")

        blank_formulas = uncalculated(args.src)
        with tempfile.TemporaryDirectory(prefix="xlsx-pdf-") as td:
            work = Path(td)
            source = only_sheet(args.src, args.sheet, work) if args.sheet else args.src
            produced, err = convert(source, "pdf", work / "out", timeout=args.timeout)
            if produced is None:
                fail(err)
            args.out.parent.mkdir(parents=True, exist_ok=True)
            args.out.write_bytes(produced.read_bytes())

        info = inspect_pdf(args.out, args.png, args.dpi, args.src, args.sheet)
        if info.get("blank_pages") and not args.allow_blank:
            args.out.unlink(missing_ok=True)
            fail(f"page(s) {info['blank_pages']} render with no ink at all, so nothing "
                 f"was written. An empty preview looks like lost data; pass "
                 f"--allow-blank if a blank page is genuinely expected"
                 + (f". Note {blank_formulas} formula cell(s) have no cached value — "
                    f"run xlsx_recalc.py first" if blank_formulas else ""))

        report = {"in": args.src.name, "out": args.out.name,
                  "replaced_existing": replaced,
                  "sheet": args.sheet, "engine": "LibreOffice",
                  "hidden_sheets_not_rendered": hidden_sheets(args.src),
                  "uncalculated_formulas": blank_formulas, **info}
        if blank_formulas:
            # This sentence used to end "and therefore render EMPTY — the picture is
            # wrong in a way the picture cannot show", which is a claim about the
            # renderer that nothing here measures, and it is false on LibreOffice:
            # a formula cell with no cached result is COMPUTED at load (measured
            # 2026-08-16, 059 §三十 — two workbooks, one library-written with no
            # caches at all, both rendered real numbers). The claim was also exactly
            # inverted in practice: it fired on a correct picture, and stayed silent
            # on the one case where the page really was wrong (stale caches, which
            # LibreOffice renders verbatim because it does not honour fullCalcOnLoad).
            # So say only what is certain — what the FILE holds — and do not pretend
            # to know what came out on paper.
            report["warning"] = (
                f"{blank_formulas} formula cell(s) carry no cached result in the "
                f"file. Anything that reads values instead of rendering — this "
                f"skill's own reader included — sees them as empty; whether they "
                f"appear on the page depends on the renderer computing them, which "
                f"this script does not check. Run xlsx_recalc.py to put the numbers "
                f"into the file")
        if info.get("hash_marked_cells"):
            where = info.get("hash_marked_columns") or []
            report["hash_warning"] = (
                f"a number came out as ### on the page — the column is too narrow to "
                f"show it, so the value is in the file and NOT in this preview"
                + (f" (too narrow: {'、'.join(where)})" if where else "")
                + ". Run xlsx_write.py --autofit to widen by measured display width")
        if info.get("columns_off_first_page"):
            report["split_warning"] = (
                f"the sheet is wider than the paper: column(s) "
                f"{'、'.join(info['columns_off_first_page'])} were moved to a page of "
                f"their own, away from the row labels that give them meaning. Narrow "
                f"the columns, or render the sheet in landscape from a spreadsheet "
                f"application — a preview split this way looks complete on page 1")
        emit(report, args.report, "images")

    return run(entry)


if __name__ == "__main__":
    raise SystemExit(main())
