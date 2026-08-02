#!/usr/bin/env python3
"""X5 — find what is wrong with the formulas. X10 — check cross-sheet references.

Four classes of finding, in the order they cost you money:

  error       a cell already carries #REF! / #DIV/0! / #VALUE! / ... — reported
              with the reference that most likely caused it, not just the token
  missing     a formula points at a sheet the workbook does not contain. This is a
              #REF! that has not happened yet: openpyxl stores the string without
              complaint and the damage appears the first time Excel opens the file
  circular    a chain of cells that reads itself. Excel shows 0 and a status-bar
              warning most people never see
  uncalc      formulas with no cached value — normal for a file a library wrote,
              and worth saying out loud because it means every number a reader
              takes from this file right now is absent, not zero

    python3 xlsx_audit.py --in book.xlsx
    python3 xlsx_audit.py --in book.xlsx --out audit.json --fail-on error,missing

Exit 0 = audited, findings reported. `--fail-on` makes chosen classes exit 1, for
callers that want this as a gate. A clean audit is not proof of correct numbers —
it is proof the formulas point at things that exist.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from office.formula import (  # noqa: E402
    build_graph, error_tokens_in, find_cycles, missing_sheets, references,
)
from xlsxcommon import emit, fail, run  # noqa: E402

CLASSES = ("error", "missing", "circular", "uncalc")


def load(path: Path):
    import openpyxl
    if not path.is_file():
        fail(f"no such file: {path}")
    try:
        return (openpyxl.load_workbook(path, data_only=False),
                openpyxl.load_workbook(path, data_only=True))
    except Exception as e:  # noqa: BLE001 - openpyxl raises several unrelated types
        fail(f"cannot open {path.name} as a workbook: {type(e).__name__}: {e}")


def blame(formula: str, token: str, known: set[str]) -> str | None:
    """Which reference most likely produced this error token.

    A #REF! baked into the formula TEXT is its own explanation. Otherwise the best
    available answer is a reference to a sheet that is gone. Anything beyond that
    would need evaluation, and a guessed cause reads exactly like a known one — so
    when there is nothing solid to say, this says nothing.
    """
    if token == "#REF!" and "#REF!" in formula:
        return "the reference was deleted; #REF! is stored in the formula itself"
    gone = missing_sheets(formula, known)
    if gone:
        return f"points at missing sheet {', '.join(repr(g) for g in gone)}"
    return None


def audit(fwb, vwb) -> dict:
    known = set(fwb.sheetnames)
    findings: list[dict] = []
    formulas: dict[str, dict[str, str]] = {}
    counts = {"cells": 0, "formulas": 0}

    for ws in fwb.worksheets:
        vws = vwb[ws.title]
        per_sheet: dict[str, str] = {}
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                counts["cells"] += 1
                cached = vws[cell.coordinate].value
                is_formula = isinstance(v, str) and v.startswith("=")
                if is_formula:
                    counts["formulas"] += 1
                    per_sheet[cell.coordinate] = v
                where = f"{ws.title}!{cell.coordinate}"

                for text, view in ((v, "formula"), (cached, "cached value")):
                    if not isinstance(text, str):
                        continue
                    if view == "formula" and not is_formula:
                        continue
                    for tok in error_tokens_in(text):
                        findings.append({
                            "class": "error", "cell": where, "token": tok,
                            "view": view, "formula": v if is_formula else None,
                            "cause": blame(v, tok, known) if is_formula else None,
                        })
                if is_formula:
                    for gone in missing_sheets(v, known):
                        findings.append({
                            "class": "missing", "cell": where, "sheet": gone,
                            "formula": v,
                            "cause": f"no sheet named {gone!r} in this workbook "
                                     f"(have: {', '.join(sorted(known))})",
                        })
                    if cached is None:
                        findings.append({"class": "uncalc", "cell": where,
                                         "formula": v,
                                         "cause": "no cached result; whatever reads "
                                                  "this file sees an empty cell"})
        formulas[ws.title] = per_sheet

    graph, stats = build_graph(formulas)
    for cycle in find_cycles(graph):
        findings.append({"class": "circular", "cell": cycle[0],
                         "chain": " -> ".join(cycle),
                         "cause": f"{len(cycle) - 1} cell(s) form a loop; Excel "
                                  f"shows 0 and warns only in the status bar"})

    by_class = {c: sum(1 for f in findings if f["class"] == c) for c in CLASSES}
    return {"sheets": fwb.sheetnames, "counts": {**counts, **stats},
            "by_class": by_class, "findings": findings,
            # Truncation is a fact about the audit, not about the workbook, so it
            # rides in the report rather than being silently absorbed.
            "graph_truncated": stats["truncated"]}


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in", dest="src", required=True, type=Path)
    ap.add_argument("--out", type=Path, help="write the full report here")
    ap.add_argument("--fail-on", default="",
                    help=f"comma-separated classes that make this exit 1 "
                         f"({', '.join(CLASSES)})")
    args = ap.parse_args()

    state: dict = {}

    def entry():
        wanted = [c.strip() for c in args.fail_on.split(",") if c.strip()]
        unknown = [c for c in wanted if c not in CLASSES]
        if unknown:
            fail(f"--fail-on: unknown class(es) {', '.join(unknown)}; "
                 f"choose from {', '.join(CLASSES)}")
        fwb, vwb = load(args.src)
        report = {"file": args.src.name, **audit(fwb, vwb)}
        fwb.close()
        vwb.close()
        emit(report, args.out, "findings")
        state["hit"] = sum(report["by_class"][c] for c in wanted)

    rc = run(entry)
    # Exit 1 means "findings you asked to fail on"; exit 2 stays "could not audit".
    # Collapsing them would make a missing file and a dirty workbook the same event.
    return 1 if rc == 0 and state.get("hit") else rc


if __name__ == "__main__":
    raise SystemExit(main())
