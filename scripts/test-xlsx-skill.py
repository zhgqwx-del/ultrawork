#!/usr/bin/env python3
"""Behaviour tests for the xlsx skill's scripts (discussions/059 S3).

L1 proves each declared capability runs and produces an artifact; L2 proves the
artifact is a legal, non-lossy workbook. Neither can say whether the numbers are
RIGHT — that a width was counted in display units rather than `len()`, that the
edit reached the cell it named, that a surgical write really did leave the other
16 parts alone. Those claims are this file's job.

    python3 scripts/test-xlsx-skill.py
    python3 scripts/test-xlsx-skill.py --json

Every assertion runs twice: once against the real output of the real scripts (must
stay silent) and once against output carrying exactly the defect it hunts (must
fire). The flaws are not invented damage — each is the implementation somebody
reaches for first. `write-via-load-save` above all: `openpyxl.load_workbook(f)` →
`save()` is what every example on the internet does, and it is what silently drops
`xl/metadata.xml` and every customXml part in the file.

The run prints a flaw -> fired-checks matrix. "All the negative controls went red"
is not the claim; "the RIGHT check went red" is, and only the matrix shows the
difference.

Exit 0 = every assertion behaved, 1 = something did not.
"""
from __future__ import annotations

import argparse
import copy
import json
import shutil
import subprocess
import sys
import tempfile
import time
import zipfile
from pathlib import Path

# ── stdout must be UTF-8 on every platform, and on Windows it is not ──────────
# This gate prints ✅/❌ and Chinese. Windows encodes a CAPTURED stdout in the
# machine's ANSI code page and Python only defaults to UTF-8 from 3.15 (PEP 686);
# CI pins 3.11. Measured on CI: this script died with
# `UnicodeEncodeError: 'charmap' codec can't encode character '\u2705'` inside its
# own `print(json.dumps(...))`. The skills were fixed for this first and the GATES
# were missed — the same defect has two homes, and only one of them was product.
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        try:
            _stream.reconfigure(encoding="utf-8")
        except (ValueError, OSError):        # already detached / not reconfigurable
            pass


REPO = Path(__file__).resolve().parent.parent
SKILL = REPO / "skills" / "builtin" / "xlsx"
FIXTURES = SKILL / "fixtures"
BOOK = FIXTURES / "book.xlsx"
NARROW = FIXTURES / "narrow.xlsx"
PY = sys.executable

MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

# What book.xlsx is known to hold. Spelled out rather than read back from the file:
# an expectation derived from the artifact it checks agrees with itself no matter
# what the artifact says.
SHEETS = ["利润表", "汇总"]
CROSS_SHEET_CELL = ("汇总", "B2", "=利润表!B5")
# 营业收入合计（含其他业务） is 13 wide characters => 26 display units, + 2 padding.
LONG_LABEL = "营业收入合计（含其他业务）"
LONG_LABEL_WIDTH = 28
LONG_LABEL_LEN_WIDTH = 15        # what len()+2 would have produced — the defect

# The merge probe's three labels, sized so that each implementation lands on a
# different number: ignore-the-title => 18, count-the-title => 42, ignore-every-
# merge => 6. Same-length labels would make the first two indistinguishable.
MERGE_TITLE = "二零二六年第三季度经营分析报告与附注说明"   # 20 wide => 42 units
MERGE_VERTICAL = "累计毛利合计金额"                        # 8 wide  => 18 units
MERGE_EXPECT_WIDTH = 18
MERGE_TITLE_WIDTH = 42
MERGE_PLAIN_WIDTH = 6

# X12's memory calibration. Sizes chosen so the per-row cost is visibly different
# between them (0.252 -> 0.118 MB per 1k rows in streaming mode) while 50k still
# builds in under a second. The large one is generated, never committed: a 50k-row
# workbook under skills/builtin/ would dominate the sentinel hash.
MEM_SMALL, MEM_LARGE = 10_000, 50_000
# Measured 2026-08-02: streaming 5.91 MB vs eager 63.15 MB at 50k rows (10.7x).
# 3x leaves a wide margin while still being far outside measurement noise.
MEM_MIN_RATIO = 3.0
PNG_DPI = "120"

STDOUT_BUDGET = 6000             # bytes one call may print for a large workbook
SCALE_ROWS = 2000                # comfortably past xlsxcommon.STDOUT_ITEM_LIMIT


# Claims this host could not exercise. Reported separately and never folded into the
# pass count — the whole reason X3 carried a wrong expectation for a month is that a
# skip and a pass looked identical at a glance.
SKIPS: list[str] = []


def run_script(name: str, *args: str) -> subprocess.CompletedProcess:
    return subprocess.run([PY, str(SKILL / "scripts" / name), *map(str, args)],
                          capture_output=True, text=True, encoding="utf-8",
                          errors="replace", timeout=300)


def run_script_from(scripts, name: str, *args: str) -> subprocess.CompletedProcess:
    """Same call against an arbitrary copy of the scripts — for LIVE controls, which
    re-run the REAL entry point with the fix backed out instead of editing the numbers
    this file collected."""
    return subprocess.run([PY, str(Path(scripts) / name), *map(str, args)],
                          capture_output=True, text=True, encoding="utf-8",
                          errors="replace", timeout=300)


def patched_scripts(work: Path, edits: list[tuple[str, str]], name: str) -> Path:
    """A copy of the skill's scripts with each edit applied exactly once.

    Raises when an anchor does not match exactly once: a control arm that silently
    failed to apply is indistinguishable from one that applied and changed nothing.
    """
    dest = work / f"patched-{name}"
    if dest.exists():
        shutil.rmtree(dest)
    shutil.copytree(SKILL / "scripts", dest,
                    ignore=shutil.ignore_patterns("__pycache__"))
    for old, new in edits:
        hits = 0
        for py in list(dest.glob("*.py")) + list(dest.glob("office/*.py")):
            text = py.read_text(encoding="utf-8")
            if old in text:
                hits += text.count(old)
                py.write_text(text.replace(old, new), encoding="utf-8")
        if hits != 1:
            raise SystemExit(f"control {name!r}: anchor matched {hits} times, expected "
                             f"1 — the control did not replicate the defect")
    return dest


# The shipped-until-2026-08-05 shape: widths applied AFTER the first append, which
# write_only silently discards. The second edit removes the self-check, so the run
# succeeds while lying — which is precisely what shipped.
IMPORT_ORDER_ANCHOR = """    if widest:
        from openpyxl.utils import get_column_letter
        for i, w in widest.items():
            ws.column_dimensions[get_column_letter(i)].width = w
    for row in rows:
        ws.append([coerce_cell(v) for v in row])"""
IMPORT_ORDER_BROKEN = """    for row in rows:
        ws.append([coerce_cell(v) for v in row])
    if widest:
        from openpyxl.utils import get_column_letter
        for i, w in widest.items():
            ws.column_dimensions[get_column_letter(i)].width = w"""
SELFCHECK_ANCHOR = """    if in_file != len(widest):"""
SELFCHECK_OFF = """    if False:"""


# The CSV comes from 利润表, whose headers are CJK: counted in display units the
# widest is well past this, counted with len() it would be about half. Measured
# rather than guessed — the point of the floor is to separate the two.
CJK_IMPORT_MIN_WIDTH = 12


def collect_import_autofit(scripts, work: Path, tag: str, csv_src: Path) -> dict:
    """`--from <csv> --autofit` — the creation path, which had no coverage at all."""
    out_x = work / f"imp-{tag}.xlsx"
    r = run_script_from(scripts, "xlsx_convert.py", "--from", csv_src, "--out", out_x,
                        "--sheet", "导入", "--autofit")
    report = {}
    if r.returncode == 0 and r.stdout.strip():
        try:
            report = json.loads(r.stdout)
        except json.JSONDecodeError:
            report = {}
    return {"exit": r.returncode, "report": report,
            "cols_in_file": cols_in_xlsx(out_x) if out_x.is_file() else 0,
            "widths": widths_in_xlsx(out_x) if out_x.is_file() else {}}


def cols_in_xlsx(path: Path) -> int:
    import re
    import zipfile
    with zipfile.ZipFile(path) as z:
        name = next((n for n in z.namelist()
                     if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")), None)
        return len(re.findall(rb"<col\b", z.read(name))) if name else 0


def widths_in_xlsx(path: Path) -> dict:
    """Column letter -> width, read straight out of the artifact."""
    import openpyxl
    from contextlib import closing
    wb = openpyxl.load_workbook(path)
    with closing(wb):
        ws = wb.worksheets[0]
        return {k: d.width for k, d in ws.column_dimensions.items() if d.width}


def with_wide_ranges(dst: Path, rows: int, formulas: int) -> None:
    """A workbook whose reference SPAN is past the sweep's bound, carrying caches.

    Shape breaks the precise dependency walk, not size: every one of these formulas
    reads the whole of column A, so `formulas x rows` cells have to be expanded to
    build the graph. openpyxl never writes cached values and the whole point here is
    to have some, so they go into the sheet XML directly afterwards.
    """
    import openpyxl
    import re
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "宽"
    for r in range(1, rows + 1):
        ws.cell(row=r, column=1, value=r)
    for r in range(1, formulas + 1):
        ws.cell(row=r, column=2, value=f"=SUM(A1:A{rows})")
    wb.save(dst)
    wb.close()
    with zipfile.ZipFile(dst) as z:
        parts = {n: z.read(n) for n in z.namelist()}
    sheet = "xl/worksheets/sheet1.xml"
    # openpyxl emits `<f>...</f><v></v>` — an EMPTY cached value, which reads back as
    # None. Filling it is the difference between a fixture with caches and one that
    # only looks like it has them.
    parts[sheet] = re.sub(rb"</f>(<v>[^<]*</v>)?", rb"</f><v>1</v>", parts[sheet])
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as z:
        for name, blob in parts.items():
            z.writestr(name, blob)


# Shape, not size: 2,000 formulas each reading 6,000 rows is 12,000,000 referenced
# cells, well past the 500,000 the precise walk is bounded to.
WIDE_ROWS, WIDE_FORMULAS = 6_000, 2_000
# Measured on this host: bounded 0.26s, bound removed 6.9s, and the version that
# cleared cell by cell through `set_cached` 2.2s here and **54s** on a 10,000-row
# chain — the first implementation of the sweep never returned at all and was killed
# after nine minutes. 60s is far above every working number and far below every
# broken one; it pins "does not hang", not a performance target.
WIDE_SECONDS_CEILING = 60.0


def with_percent_column(dst: Path, col_b_width: float, wide: bool = False) -> None:
    """利润表 in miniature: a percent column too narrow for what it displays.

    `1.0877...` under `0.0%` shows as `108.8%` — six characters in a column six
    units wide, which Excel and LibreOffice both print as `###`. The value is a
    CACHED formula result, because that is the shape the real file has and the shape
    a width routine that only looks at text cells cannot see.
    """
    import openpyxl
    import re
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "利润表"
    ws["A1"], ws["B1"] = "科目", "同比"
    ws["A2"], ws["B2"] = "营业利润", "=1.0877192982456141"
    ws["B2"].number_format = "0.0%"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = col_b_width
    if wide:
        # Wider than the paper: LibreOffice moves 同比 to a page of its own. The extra
        # rows are not decoration — at two rows spread over two sheets of paper the
        # ink fraction falls under BLANK_INK and the blank-page guard refuses to write
        # the preview at all, so the fixture never reaches the check it is for.
        for r in range(3, 24):
            ws.cell(row=r, column=1, value=f"科目{r}")
            ws.cell(row=r, column=2, value=r / 7)
            ws.cell(row=r, column=2).number_format = "0.0%"
        for i in range(1, 3):
            ws.column_dimensions[get_column_letter(i)].width = 60
    wb.save(dst)
    wb.close()
    with zipfile.ZipFile(dst) as z:
        parts = {n: z.read(n) for n in z.namelist()}
    sheet = "xl/worksheets/sheet1.xml"
    parts[sheet] = re.sub(rb"</f>(<v>[^<]*</v>)?",
                          rb"</f><v>1.0877192982456141</v>", parts[sheet])
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as z:
        for name, blob in parts.items():
            z.writestr(name, blob)


def collect_percent_width(scripts, work: Path, tag: str) -> dict:
    """`--autofit` over a column whose NUMBER is what does not fit."""
    src = work / "pct-narrow.xlsx"
    if not src.exists():
        with_percent_column(src, col_b_width=6)
    first = work / f"pct-{tag}.xlsx"
    r1 = run_script_from(scripts, "xlsx_write.py", "--in", src, "--out", first,
                         "--autofit", "--report", work / f"pct-{tag}.json")
    # Running it again over its own output is the no-op case: nothing left to widen,
    # and the report has to be able to say that.
    again = work / f"pct-{tag}-again.xlsx"
    run_script_from(scripts, "xlsx_write.py", "--in", first, "--out", again,
                    "--autofit", "--report", work / f"pct-{tag}-again.json")
    pdf = work / f"pct-{tag}.pdf"
    run_script_from(scripts, "xlsx_pdf.py", "--in", first, "--out", pdf)
    return {"exit": r1.returncode,
            "report": json.loads((work / f"pct-{tag}.json").read_text(encoding="utf-8")),
            "noop": json.loads(
                (work / f"pct-{tag}-again.json").read_text(encoding="utf-8")),
            "width_before": widths_of(src, "利润表").get("B"),
            "width_after": widths_of(first, "利润表").get("B"),
            "rendered": pdf_text(pdf)}


def collect_split(scripts, work: Path, tag: str) -> dict:
    """A sheet wider than the paper — the columns on the right end up alone."""
    src = work / "pct-wide.xlsx"
    if not src.exists():
        with_percent_column(src, col_b_width=90, wide=True)
    report = work / f"split-{tag}.json"
    run_script_from(scripts, "xlsx_pdf.py", "--in", src,
                    "--out", work / f"split-{tag}.pdf", "--report", report)
    return json.loads(report.read_text(encoding="utf-8")) if report.exists() else {}


def collect_replaced(scripts, work: Path, tag: str) -> dict:
    """Each writing script, run twice at the same --out: fresh, then over itself."""
    src = BOOK
    runs: dict[str, list] = {}
    plans = (
        ("xlsx_write.py", lambda out: ("--in", src, "--out", out, "--sheet", "利润表",
                                       "--set", "B3=1240")),
        ("xlsx_recalc.py", lambda out: ("--in", src, "--out", out)),
        ("xlsx_pdf.py", lambda out: ("--in", work / "calc.xlsx", "--out", out)),
    )
    for name, argv in plans:
        out = work / f"rep-{tag}-{name}.{'pdf' if 'pdf' in name else 'xlsx'}"
        out.unlink(missing_ok=True)
        seen = []
        for _ in range(2):
            r = run_script_from(scripts, name, *argv(out))
            try:
                seen.append(json.loads(r.stdout).get("replaced_existing"))
            except json.JSONDecodeError:
                seen.append(f"no JSON on stdout (exit {r.returncode})")
        runs[name] = seen
    # Structural coverage: every script that guards its output must also report on
    # it. A new writer that forgets the field is the way this rots.
    missing = sorted(p.name for p in Path(scripts).glob("*.py")
                     if "ensure_distinct(" in p.read_text(encoding="utf-8")
                     and p.name != "xlsxcommon.py"
                     and "replaced_existing" not in p.read_text(encoding="utf-8"))
    return {"runs": runs, "scripts_without_the_field": missing}


def collect_hash_marks(scripts, work: Path, tag: str) -> dict:
    """Render a sheet whose number does not fit, and one that says `###` for real."""
    narrow = work / "pct-narrow.xlsx"
    if not narrow.exists():
        with_percent_column(narrow, col_b_width=6)
    literal = work / "pct-literal.xlsx"
    if not literal.exists():
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "利润表"
        ws["A1"], ws["B1"] = "科目", "备注"
        ws["A2"], ws["B2"] = "营业利润", "###"      # a cell that really holds it
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        wb.save(literal)
        wb.close()
    out = {}
    for label, src in (("narrow", narrow), ("literal", literal)):
        report = work / f"hash-{tag}-{label}.json"
        run_script_from(scripts, "xlsx_pdf.py", "--in", src,
                        "--out", work / f"hash-{tag}-{label}.pdf", "--report", report)
        out[label] = json.loads(report.read_text(encoding="utf-8")) \
            if report.exists() else {}
    return out


def collect_stale_images(scripts, work: Path, tag: str) -> dict:
    """Render a 1-page document into a directory holding a longer render's tail."""
    src = work / "pct-narrow.xlsx"
    if not src.exists():
        with_percent_column(src, col_b_width=6)
    png_dir = work / f"pngdir-{tag}"
    png_dir.mkdir(parents=True, exist_ok=True)
    for name in ("page-002.png", "page-009.png", "notes.png"):
        (png_dir / name).write_bytes(b"")
    report = work / f"stale-img-{tag}.json"
    run_script_from(scripts, "xlsx_pdf.py", "--in", src,
                    "--out", work / f"stale-img-{tag}.pdf", "--png", png_dir,
                    "--report", report)
    return {"report": json.loads(report.read_text(encoding="utf-8"))
            if report.exists() else {},
            "left": sorted(p.name for p in png_dir.iterdir())}


def pdf_text(pdf: Path) -> str | None:
    try:
        import pypdfium2 as pdfium
    except ImportError:
        return None
    if not pdf.is_file():
        return ""
    doc = pdfium.PdfDocument(str(pdf))
    try:
        return "\n".join(doc[i].get_textpage().get_text_range() for i in range(len(doc)))
    finally:
        doc.close()


def collect_audit_spotless(scripts, work: Path, tag: str, src: Path) -> dict:
    """An audit of a workbook with nothing wrong with it."""
    out = work / f"audit-spotless-{tag}.json"
    r = run_script_from(scripts, "xlsx_audit.py", "--in", src, "--out", out)
    return {"exit": r.returncode,
            "report": json.loads(out.read_text(encoding="utf-8"))
            if out.exists() else {}}


def collect_stale_bound(scripts, work: Path, tag: str) -> dict:
    """`--set` on a workbook whose dependency graph is too big to expand."""
    # NOT "wide.xlsx": the autofit section already owns that name in this work dir,
    # and reusing it silently handed this collector that file instead — 0 cached
    # cells, which the non-vacuity guard below caught on the first run.
    wide = work / "wide-ranges.xlsx"
    if not wide.exists():
        with_wide_ranges(wide, WIDE_ROWS, WIDE_FORMULAS)
    out_x = work / f"wide-ranges-{tag}.xlsx"
    report = work / f"wide-ranges-{tag}.json"
    started = time.monotonic()
    r = run_script_from(scripts, "xlsx_write.py", "--in", wide, "--out", out_x,
                        "--sheet", "宽", "--set", "A1=5", "--report", report)
    return {"exit": r.returncode, "seconds": time.monotonic() - started,
            "report": json.loads(report.read_text(encoding="utf-8"))
            if report.exists() else {},
            "cached_before": cached_cells(wide),
            "cached_after": cached_cells(out_x) if out_x.exists() else set()}


def cached_cells(path: Path) -> set[str]:
    """`{"Sheet!REF"}` for every formula cell that carries a cached result."""
    import openpyxl
    from contextlib import closing
    f = openpyxl.load_workbook(path, data_only=False)
    v = openpyxl.load_workbook(path, data_only=True)
    out: set[str] = set()
    with closing(f), closing(v):
        for name in f.sheetnames:
            for frow, vrow in zip(f[name].iter_rows(),
                                  v[name].iter_rows(values_only=True)):
                for cell, value in zip(frow, vrow):
                    if isinstance(cell.value, str) and cell.value.startswith("=") \
                            and value is not None:
                        out.add(f"{name}!{cell.coordinate}")
    return out


def parts_of(path: Path) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as z:
        return {n: z.read(n) for n in z.namelist() if not n.endswith("/")}


def sheet_child_order(path: Path, part: str = "xl/worksheets/sheet1.xml") -> list[str]:
    from lxml import etree
    root = etree.fromstring(parts_of(path)[part])
    return [str(c.tag).rsplit("}", 1)[-1] for c in root]


def widths_of(path: Path, sheet: str) -> dict[str, float]:
    import openpyxl
    wb = openpyxl.load_workbook(path)
    try:
        ws = wb[sheet]
        return {letter: dim.width for letter, dim in ws.column_dimensions.items()
                if dim.width is not None}
    finally:
        wb.close()


def cells_of(path: Path, sheet: str, refs: list[str]) -> dict[str, object]:
    import openpyxl
    wb = openpyxl.load_workbook(path)
    try:
        ws = wb[sheet]
        return {r: ws[r].value for r in refs}
    finally:
        wb.close()


def values_of(path: Path, sheet: str, refs: list[str]) -> dict[str, object]:
    """The CACHED values, not the formulas.

    `cells_of` opens the workbook with data_only=False and therefore hands back
    "=B3-B4" where a recalculated number is expected — which is what the first
    version of K5 asserted against, so the check failed on correct output.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb[sheet]
        return {r: ws[r].value for r in refs}
    finally:
        wb.close()


def style_ids(path: Path, sheet_part: str, refs: list[str]) -> dict[str, str | None]:
    from lxml import etree
    root = etree.fromstring(parts_of(path)[sheet_part])
    out: dict[str, str | None] = {r: None for r in refs}
    for c in root.iter(f"{MAIN}c"):
        if c.get("r") in out:
            out[c.get("r")] = c.get("s")
    return out


# ── inputs the fixtures deliberately do not carry ─────────────────────────────
def with_calc_chain(src: Path, dst: Path) -> None:
    """Add an `xl/calcChain.xml` to a package.

    openpyxl never writes one, so the fixture cannot carry it — but every workbook
    Excel has saved does, and carrying a stale one across a formula edit is what
    makes Excel report a good file as damaged. Built here so the drop is testable.
    """
    sys.path.insert(0, str(SKILL / "scripts"))
    from office.package import Package
    pkg = Package.open(src)
    pkg.write("xl/calcChain.xml",
              b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
              b'<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/'
              b'2006/main"><c r="B5" i="1"/><c r="D5" i="1"/></calcChain>')
    pkg.set_override("xl/calcChain.xml", "application/vnd.openxmlformats-"
                     "officedocument.spreadsheetml.calcChain+xml")
    pkg.add_relationship("xl/_rels/workbook.xml.rels",
                         "http://schemas.openxmlformats.org/officeDocument/2006/"
                         "relationships/calcChain", "calcChain.xml")
    pkg.save(dst)


def with_shared_formula(src: Path, dst: Path) -> None:
    """Turn D3 into the MASTER of a shared formula covering D3:D5.

    This is how Excel stores a formula filled down a column, and openpyxl does not
    produce it. Overwriting the master leaves D4 and D5 with `<f t="shared" si="0"/>`
    and no definition to share — a file that opens and shows nothing where the
    numbers were.
    """
    sys.path.insert(0, str(SKILL / "scripts"))
    from lxml import etree
    from office.package import Package
    pkg = Package.open(src)
    root = pkg.tree("xl/worksheets/sheet1.xml")
    for c in root.iter(f"{MAIN}c"):
        ref = c.get("r")
        if ref not in ("D3", "D4", "D5"):
            continue
        for child in list(c):
            c.remove(child)
        f = etree.SubElement(c, f"{MAIN}f")
        f.set("t", "shared")
        f.set("si", "0")
        if ref == "D3":
            f.set("ref", "D3:D5")
            f.text = "B3/C3-1"
    pkg.put_tree("xl/worksheets/sheet1.xml", root)
    pkg.save(dst)


def with_dangling_rel(src: Path, dst: Path) -> None:
    """A relationship pointing at a part that is not in the package.

    Damage the skill did not cause. It must be reported and must NOT block the edit —
    a fidelity-minded tool that refuses to touch an already-damaged file is useless
    on exactly the documents that need help.
    """
    sys.path.insert(0, str(SKILL / "scripts"))
    from office.package import Package
    pkg = Package.open(src)
    pkg.add_relationship("xl/_rels/workbook.xml.rels",
                         "http://schemas.openxmlformats.org/officeDocument/2006/"
                         "relationships/sheetMetadata", "metadata.xml")
    pkg.save(dst)


def big_workbook(path: Path, rows: int) -> None:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    ws.append(["科目", "金额", "备注"])
    for i in range(rows):
        ws.append([f"科目{i}", i * 7, f"备注{i}"])
    wb.save(path)
    wb.close()


# ── collect: run the real scripts once ────────────────────────────────────────
def collect(work: Path) -> dict:
    import openpyxl

    ctx: dict = {}

    # --- read -----------------------------------------------------------------
    r = run_script("xlsx_read.py", "--in", BOOK, "--sheet", "利润表",
                   "--range", "A1:D5", "--out", work / "read.json")
    read_json = json.loads((work / "read.json").read_text(encoding="utf-8"))
    rejects = []
    for label, args in (
            ("open-ended range", ("--range", "A1:D")),
            ("oversized range", ("--range", "A1:ZZ1048576")),
            ("unknown sheet", ("--sheet", "没有这个表", "--range", "A1:B2")),
            ("not a cell reference", ("--cells", "无效")),
    ):
        p = run_script("xlsx_read.py", "--in", BOOK, *args)
        rejects.append({"case": label, "exit": p.returncode,
                        "stderr": p.stderr.strip()})
    ctx["read"] = {"json": read_json, "rejects": rejects, "exit": r.returncode}

    # --- surgical write, and the control everyone writes instead ---------------
    edited = work / "edited.xlsx"
    w = run_script("xlsx_write.py", "--in", BOOK, "--out", edited, "--sheet", "利润表",
                   "--set", "B3=1310", "--set-formula", "D5=B5/C5-1",
                   "--append-row", "其他业务收入,88,74",
                   "--report", work / "write.json")
    write_report = json.loads((work / "write.json").read_text(encoding="utf-8"))

    naive = work / "naive.xlsx"
    wb = openpyxl.load_workbook(BOOK)
    ws = wb["利润表"]
    ws["B3"] = 1310
    ws["D5"] = "=B5/C5-1"
    ws.append(["其他业务收入", 88, 74])
    wb.save(naive)
    wb.close()

    before, after, naive_after = parts_of(BOOK), parts_of(edited), parts_of(naive)
    ctx["write"] = {
        "exit": w.returncode,
        "report": write_report,
        "parts_before": sorted(before),
        "parts_after": sorted(after),
        "identical": sorted(n for n in before if after.get(n) == before[n]),
        "naive_parts_after": sorted(naive_after),
        "naive_identical": sorted(n for n in before if naive_after.get(n) == before[n]),
        "cells": cells_of(edited, "利润表", ["B3", "D5", "B5", "A6", "B6", "C6"]),
        # D5 carries a percent number format in the fixture; an edit that resets it
        # is the "I changed one number and the column turned into General" failure.
        "styles": style_ids(edited, "xl/worksheets/sheet1.xml", ["D5", "D3"]),
        "styles_before": style_ids(BOOK, "xl/worksheets/sheet1.xml", ["D5", "D3"]),
        "sheet_order": sheet_child_order(edited),
    }

    # --- calcChain: dropped on a formula edit, kept on a value-only edit --------
    with_cc = work / "with-calcchain.xlsx"
    with_calc_chain(BOOK, with_cc)
    cc_formula = work / "cc-formula.xlsx"
    cc_value = work / "cc-value.xlsx"
    run_script("xlsx_write.py", "--in", with_cc, "--out", cc_formula,
               "--sheet", "利润表", "--set-formula", "D5=B5/C5-1")
    run_script("xlsx_write.py", "--in", with_cc, "--out", cc_value,
               "--sheet", "利润表", "--set", "B3=1310")
    ctx["calcchain"] = {
        "input_has": "xl/calcChain.xml" in parts_of(with_cc),
        "after_formula_edit": "xl/calcChain.xml" in parts_of(cc_formula),
        "after_value_edit": "xl/calcChain.xml" in parts_of(cc_value),
    }

    # --- shared formula: refused, not silently broken --------------------------
    shared = work / "shared.xlsx"
    with_shared_formula(BOOK, shared)
    sp = run_script("xlsx_write.py", "--in", shared, "--out", work / "shared-out.xlsx",
                    "--sheet", "利润表", "--set", "D3=0.5")
    sp_other = run_script("xlsx_write.py", "--in", shared,
                          "--out", work / "shared-other.xlsx",
                          "--sheet", "利润表", "--set", "B3=1310")
    ctx["shared"] = {
        "master_exit": sp.returncode, "master_stderr": sp.stderr.strip(),
        "master_wrote": (work / "shared-out.xlsx").exists(),
        # A cell that is NOT the shared master must still be editable, or the guard
        # is just "refuse to edit any workbook Excel has ever touched".
        "other_exit": sp_other.returncode,
    }

    # --- pre-existing damage: reported, not fatal ------------------------------
    damaged = work / "damaged.xlsx"
    with_dangling_rel(BOOK, damaged)
    dp = run_script("xlsx_write.py", "--in", damaged, "--out", work / "damaged-out.xlsx",
                    "--sheet", "利润表", "--set", "B3=1310",
                    "--report", work / "damaged.json")
    damaged_report = json.loads((work / "damaged.json").read_text(encoding="utf-8")) \
        if (work / "damaged.json").exists() else {}
    ctx["damage"] = {
        "exit": dp.returncode,
        "wrote": (work / "damaged-out.xlsx").exists(),
        "pre_existing": damaged_report.get("pre_existing_package_findings", []),
    }

    # --- autofit ---------------------------------------------------------------
    wide = work / "wide.xlsx"
    a = run_script("xlsx_write.py", "--in", NARROW, "--out", wide, "--autofit",
                   "--report", work / "fit.json")
    ctx["autofit"] = {
        "exit": a.returncode,
        "report": json.loads((work / "fit.json").read_text(encoding="utf-8")),
        "widths": widths_of(wide, "利润表"),
        "narrow_widths": widths_of(NARROW, "利润表"),
        "sheet_order": sheet_child_order(wide),
    }

    # A merged title must not drive its first column; a vertical merge must.
    #
    # The three labels have DELIBERATELY different lengths. The first version of this
    # probe used the same string for the horizontal and the vertical merge, so
    # counting the title and ignoring it produced the identical width and the
    # negative control could not fire — a control arm that cannot tell the two
    # implementations apart is not a control.
    merged = work / "merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "表"
    ws["A1"] = MERGE_TITLE                         # 20 wide chars => 42 units
    ws.merge_cells("A1:F1")                        # displayed across six columns
    ws["A3"] = "毛利"                              # 2 wide chars => 6 units
    ws["A5"] = MERGE_VERTICAL                      # 8 wide chars => 18 units
    ws.merge_cells("A5:A7")                        # vertical: no extra room
    wb.save(merged)
    wb.close()
    merged_out = work / "merged-wide.xlsx"
    run_script("xlsx_write.py", "--in", merged, "--out", merged_out, "--autofit")
    ctx["merge"] = {"width_a": widths_of(merged_out, "表").get("A")}

    # --- audit: clean input, then one carrying every class of defect -----------
    a = run_script("xlsx_audit.py", "--in", BOOK, "--out", work / "audit-clean.json")
    clean_audit = json.loads((work / "audit-clean.json").read_text(encoding="utf-8"))

    broken = work / "broken.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "利润表"
    ws["A1"] = 10
    ws["B1"] = "=预算表!A1"        # a sheet that does not exist: a #REF! in waiting
    ws["C1"] = "=#REF!*2"          # an error already baked into the formula text
    ws["D1"] = "=E1+1"             # two-cell cycle
    ws["E1"] = "=D1+1"
    ws["F1"] = "=F1"               # self-reference
    ws["G1"] = '="见 预算表!A1 的说明"'   # a reference inside a STRING — not a reference
    ws["H1"] = "=LOG10(A1)"        # a function whose name is shaped like a cell ref
    sm = wb.create_sheet("汇总")
    sm["A1"] = "=利润表!A1"        # a legitimate cross-sheet link: must NOT be flagged
    wb.save(broken)
    wb.close()
    b = run_script("xlsx_audit.py", "--in", broken, "--out", work / "audit-broken.json",
                   "--fail-on", "error,missing,circular")
    broken_audit = json.loads((work / "audit-broken.json").read_text(encoding="utf-8"))
    ctx["audit"] = {
        "clean_exit": a.returncode, "clean": clean_audit,
        "broken_exit": b.returncode, "broken": broken_audit,
        "cells": {f["cell"] for f in broken_audit["findings"]
                  if f["class"] != "uncalc"},
    }

    # --- formulas refused at write time, before they reach the file ------------
    refusals = {}
    for label, spec, sheet in (
            ("missing sheet", "B7==预算表!A1", "利润表"),
            ("error token baked in", "B7==#REF!+1", "利润表"),
    ):
        out = work / f"refused-{len(refusals)}.xlsx"
        p = run_script("xlsx_write.py", "--in", BOOK, "--out", out,
                       "--sheet", sheet, "--set-formula", spec)
        refusals[label] = {"exit": p.returncode, "stderr": p.stderr.strip(),
                           "wrote": out.exists()}
    linked = work / "linked.xlsx"
    ok = run_script("xlsx_write.py", "--in", BOOK, "--out", linked, "--sheet", "利润表",
                    "--set-formula", "C7==汇总!B2")
    ctx["refuse"] = {"cases": refusals, "legit_exit": ok.returncode,
                     "legit_formula": cells_of(linked, "利润表", ["C7"])["C7"]}

    # --- the rebuild path: openpyxl object model + graft -----------------------
    fmt = work / "formatted.xlsx"
    f = run_script("xlsx_format.py", "--in", BOOK, "--out", fmt, "--sheet", "利润表",
                   "--range", "B3:C5", "--number-format", "#,##0",
                   "--font-color", "0000FF", "--bold", "--fill", "FFF2CC",
                   "--border", "thin", "--report", work / "format.json")
    fmt_report = json.loads((work / "format.json").read_text(encoding="utf-8"))

    cond = work / "conditional.xlsx"
    run_script("xlsx_format.py", "--in", BOOK, "--out", cond, "--sheet", "利润表",
               "--rules", FIXTURES / "rules.json")
    panes = work / "panes.xlsx"
    run_script("xlsx_format.py", "--in", BOOK, "--out", panes, "--sheet", "利润表",
               "--freeze", "C6", "--filter", "A2:C5")
    charted = work / "charted.xlsx"
    c = run_script("xlsx_chart.py", "--in", BOOK, "--out", charted, "--sheet", "汇总",
                   "--type", "column", "--data", "利润表!B2:C4",
                   "--categories", "利润表!A3:A4", "--anchor", "D18",
                   "--title", "季度对比", "--report", work / "chart.json")
    chart_report = json.loads((work / "chart.json").read_text(encoding="utf-8"))

    import openpyxl as _o

    def style_probe(path: Path, sheet: str, refs: list[str]) -> dict:
        wb = _o.load_workbook(path)
        try:
            ws = wb[sheet]
            out = {}
            for r in refs:
                cell = ws[r]
                fill = cell.fill
                out[r] = {
                    "number_format": cell.number_format,
                    "bold": bool(cell.font.bold),
                    "color": cell.font.color.rgb if cell.font.color
                    and isinstance(cell.font.color.rgb, str) else None,
                    "size": cell.font.size, "name": cell.font.name,
                    "fill": fill.start_color.rgb if fill and fill.fill_type
                    and isinstance(fill.start_color.rgb, str) else None,
                    "border": cell.border.left.style,
                }
            return out
        finally:
            wb.close()

    def cf_probe(path: Path, sheet: str) -> dict:
        wb = _o.load_workbook(path)
        try:
            ws = wb[sheet]
            out: dict[str, list[str]] = {}
            for rng in ws.conditional_formatting:
                out.setdefault(str(rng.sqref), []).extend(r.type for r in rng.rules)
            return out
        finally:
            wb.close()

    def chart_probe(path: Path, sheet: str) -> dict:
        wb = _o.load_workbook(path)
        try:
            ws = wb[sheet]
            titles = []
            refs = []
            for ch in ws._charts:
                for s in ch.series:
                    titles.append(getattr(getattr(s.tx, "strRef", None), "f", None)
                                  if s.tx else None)
                    refs.append(s.val.numRef.f if s.val and s.val.numRef else None)
            return {"count": len(ws._charts), "series_titles": titles,
                    "series_refs": refs}
        finally:
            wb.close()

    base_style = style_probe(BOOK, "利润表", ["B3", "D3"])
    ctx["rebuild"] = {
        "format_exit": f.returncode, "chart_exit": c.returncode,
        "format_report": fmt_report, "chart_report": chart_report,
        "style": style_probe(fmt, "利润表", ["B3", "C5", "D3"]),
        "style_before": base_style,
        "custom_parts": {
            tag: sum(1 for n in parts_of(p) if n.startswith("customXml"))
            for tag, p in (("format", fmt), ("conditional", cond),
                           ("panes", panes), ("chart", charted))},
        "custom_parts_in": sum(1 for n in parts_of(BOOK) if n.startswith("customXml")),
        "cf": cf_probe(cond, "利润表"),
        "cf_before": cf_probe(BOOK, "利润表"),
        "panes": (lambda p: {"freeze": p[0], "filter": p[1]})(
            (lambda wb: (wb["利润表"].freeze_panes, wb["利润表"].auto_filter.ref))(
                _o.load_workbook(panes))),
        "panes_before": {"freeze": "A3", "filter": "A2:D5"},
        "chart": chart_probe(charted, "汇总"),
        "chart_before": chart_probe(BOOK, "汇总"),
    }

    # Fault injection: make the graft restore nothing and assert the rebuild REFUSES.
    # Same shape as the L2 gate's broken-soffice cases — a repair path nobody has
    # watched fail is not evidence that it repairs anything.
    probe = ("import sys, pathlib\n"
             f"sys.path.insert(0, {str(SKILL / 'scripts')!r})\n"
             "from office import rebuild as R\n"
             "R.graft_missing_parts = lambda base, prod: "
             "{'restored': [], 'skipped': [], 'restored_count': 0, 'skipped_count': 0}\n"
             f"out = pathlib.Path({str(work / 'nograft.xlsx')!r})\n"
             "try:\n"
             f"    R.rebuild(pathlib.Path({str(BOOK)!r}), out, lambda wb: None)\n"
             "    print('RAISED=no')\n"
             "except R.RebuildError as e:\n"
             "    print('RAISED=yes', 'customXml' in str(e))\n"
             "print('WROTE=', out.exists())\n")
    fault = subprocess.run([PY, "-c", probe], capture_output=True, text=True,
                           encoding="utf-8", timeout=300)
    ctx["graft_fault"] = {"stdout": fault.stdout.strip(), "exit": fault.returncode}

    # --- X4: the two engines, and the evaluator's measured coverage boundary ----
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "xlsx_calibration", REPO / "scripts" / "xlsx-evaluator-calibration.py")
    cal = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(cal)

    py_results = cal.python_results()
    lo_results, lo_err = cal.soffice_results()
    ctx["calibration"] = {
        "mismatches": [{"formula": f, "pinned": want, "python": py_results[f]}
                       for f, want in cal.CALIBRATION
                       if cal.norm(py_results[f]) != cal.norm(want)],
        "leaked": [{"formula": f, "why": why, "got": cal.python_refusals()[f]}
                   for f, why in cal.MUST_REFUSE
                   if cal.python_refusals()[f][0] != "REFUSED"],
        "unexercised": cal.unexercised(),
        "pinned_rows": len(cal.CALIBRATION),
        "must_refuse_rows": len(cal.MUST_REFUSE),
        "soffice_available": not lo_err,
        "soffice_reason": lo_err,
        "soffice_drift": [] if lo_err else
        [{"formula": f, "pinned": want, "soffice": lo_results[f]}
         for f, want in cal.CALIBRATION
         if cal.norm(lo_results[f]) != cal.norm(want)],
    }
    if lo_err:
        SKIPS.append(
            f"K4 (LibreOffice agrees with every pinned value) — {lo_err}. "
            f"Residual coverage: K1 still checks the python engine against the pins, "
            f"but NOTHING checked the pins themselves this run — that is exactly how "
            f"the L2 gate carried a wrong X3 expectation for a month.")

    recalc = work / "recalced.xlsx"
    rc = run_script("xlsx_recalc.py", "--in", BOOK, "--out", recalc,
                    "--report", work / "recalc.json")
    recalc_report = json.loads((work / "recalc.json").read_text(encoding="utf-8"))
    mixed = work / "mixed.xlsx"
    wb = openpyxl.load_workbook(BOOK)
    ws = wb["利润表"]
    ws["B8"] = "=SUM(B3:B4)"          # both engines can do this
    ws["B9"] = "=VLOOKUP(B3,A3:B4,2,0)"   # only LibreOffice can
    wb.save(mixed)
    wb.close()
    mx = run_script("xlsx_recalc.py", "--in", mixed, "--report", work / "mixed.json")
    mixed_report = json.loads((work / "mixed.json").read_text(encoding="utf-8"))
    py_only = run_script("xlsx_recalc.py", "--in", mixed, "--engine", "python",
                         "--report", work / "pyonly.json")
    py_only_report = json.loads((work / "pyonly.json").read_text(encoding="utf-8"))
    ctx["recalc"] = {
        "exit": rc.returncode, "report": recalc_report,
        "cached_after": values_of(recalc, "利润表", ["B5", "D5"]),
        "cached_before": values_of(BOOK, "利润表", ["B5", "D5"]),
        "formula_after": cells_of(recalc, "利润表", ["B5", "D5"]),
        "mixed": mixed_report, "mixed_exit": mx.returncode,
        "python_only": py_only_report, "python_only_exit": py_only.returncode,
    }

    # --- X11 conversions -------------------------------------------------------
    calc = work / "calc.xlsx"
    run_script("xlsx_recalc.py", "--in", BOOK, "--out", calc)

    # --- stale caches: an edit must not leave results that no longer match -------
    # `calc` is the only fixture here that HAS cached values — book.xlsx is
    # library-written and carries none, so this assertion run against it would pass
    # while looking at nothing.
    stale_out = work / "stale.xlsx"
    st = run_script("xlsx_write.py", "--in", calc, "--out", stale_out,
                    "--sheet", "利润表", "--set", "B3=1310",
                    "--report", work / "stale.json")
    ctx["stale"] = {
        "exit": st.returncode,
        "report": json.loads((work / "stale.json").read_text(encoding="utf-8")),
        # Read from the FILES by this gate. The report's own list is the claim under
        # test and cannot also be the evidence for it.
        "cached_before": cached_cells(calc),
        "cached_after": cached_cells(stale_out),
    }
    ctx["stale_bound"] = collect_stale_bound(SKILL / "scripts", work, "real")

    # A report with NOTHING in it — the shape most easily read as "no problems".
    # book.xlsx cannot play this part: it is library-written, so every formula is
    # `uncalc` and the audit always has findings. `calc.xlsx` has cached values.
    ctx["audit"] = {**ctx["audit"], "spotless": collect_audit_spotless(
        SKILL / "scripts", work, "real", calc)}
    ctx["percent_width"] = collect_percent_width(SKILL / "scripts", work, "real")
    ctx["split"] = collect_split(SKILL / "scripts", work, "real")
    ctx["hash"] = collect_hash_marks(SKILL / "scripts", work, "real")
    ctx["stale_images"] = collect_stale_images(SKILL / "scripts", work, "real")
    ctx["replaced"] = collect_replaced(SKILL / "scripts", work, "real")
    csv_out, back = work / "sheet.csv", work / "back.xlsx"
    run_script("xlsx_convert.py", "--in", calc, "--to", "csv", "--out", csv_out,
               "--sheet", "利润表")
    run_script("xlsx_convert.py", "--from", csv_out, "--out", back, "--sheet", "明细")
    jsonl = work / "rows.jsonl"
    run_script("xlsx_convert.py", "--in", calc, "--to", "jsonl", "--out", jsonl,
               "--sheet", "利润表", "--header-row", "2")
    jsonl_r1 = work / "rows-r1.jsonl"
    run_script("xlsx_convert.py", "--in", calc, "--to", "jsonl", "--out", jsonl_r1,
               "--sheet", "利润表")
    ctx["_csv_src"] = csv_out
    ctx["import_autofit"] = collect_import_autofit(SKILL / "scripts", work,
                                                    "real", csv_out)
    ctx["convert"] = {
        "csv_head": csv_out.read_bytes()[:3],
        "csv_text": csv_out.read_text(encoding="utf-8-sig"),
        "roundtrip": [list(r) for r in (lambda wb: [
            [c.value for c in row] for row in wb["明细"].iter_rows()])(
            openpyxl.load_workbook(back))],
        "jsonl": [json.loads(l) for l in jsonl.read_text(encoding="utf-8").splitlines()],
        "jsonl_row1": [json.loads(l) for l in
                       jsonl_r1.read_text(encoding="utf-8").splitlines()],
        "stats": json.loads(run_script("xlsx_convert.py", "--in", calc, "--stats",
                                       "--sheet", "利润表", "--header-row", "2").stdout),
    }

    # --- X12: bounded memory, measured on THIS script --------------------------
    MEM_PROBE = (
        "import sys, tracemalloc, runpy, openpyxl\n"
        "if sys.argv[1] == 'eager':\n"
        "    _o = openpyxl.load_workbook\n"
        "    openpyxl.load_workbook = lambda *a, **k: _o(*a, **{**k, 'read_only': False})\n"
        "script, book = sys.argv[2], sys.argv[3]\n"
        "sys.argv = [script, '--in', book, '--stats', '--sheet', '明细']\n"
        "tracemalloc.start()\n"
        "try:\n    runpy.run_path(script, run_name='__main__')\n"
        "except SystemExit:\n    pass\n"
        "print(tracemalloc.get_traced_memory()[1])\n")

    def peak_mb(mode: str, book: Path) -> float:
        r = subprocess.run([PY, "-c", MEM_PROBE, mode,
                            str(SKILL / "scripts" / "xlsx_convert.py"), str(book)],
                           capture_output=True, text=True, encoding="utf-8",
                           errors="replace", timeout=1800)
        # Diagnose rather than crash. On Windows this line raised
        # `AttributeError: 'NoneType' object has no attribute 'strip'` — a message
        # that says nothing about WHY, on a platform the author cannot run. The
        # probe's own exit code and stderr are the evidence; an AttributeError
        # throws them away. Same lesson as the truncated gate output earlier in
        # this branch: a diagnostic that hides the cause is not a diagnostic.
        out = (r.stdout or "").strip()
        last = out.splitlines()[-1] if out else ""
        if not last.isdigit():
            raise SystemExit(
                f"[error] the X12 memory probe ({mode}, {book.name}) produced no "
                f"number.\n  exit={r.returncode}\n  stdout={(r.stdout or '')[:400]!r}"
                f"\n  stderr={(r.stderr or '')[:800]!r}")
        return int(last) / 1024 / 1024

    mem: dict[int, dict[str, float]] = {}
    for rows in (MEM_SMALL, MEM_LARGE):
        big = work / f"mem{rows}.xlsx"
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("明细")
        ws.append(["科目", "金额", "备注"])
        for i in range(rows):
            ws.append([f"科目{i}", i * 7, f"备注{i}"])
        wb.save(big)
        wb.close()
        mem[rows] = {m: peak_mb(m, big) for m in ("stream", "eager")}
    ctx["memory"] = mem

    # --- X13 render ------------------------------------------------------------
    pdf, pages = work / "preview.pdf", work / "pages"
    p13 = run_script("xlsx_pdf.py", "--in", calc, "--out", pdf, "--png", pages,
                     "--dpi", PNG_DPI, "--report", work / "pdf.json")
    pdf_report = json.loads((work / "pdf.json").read_text(encoding="utf-8")) \
        if (work / "pdf.json").exists() else {}
    empty = work / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "空"
    wb.save(empty)
    wb.close()
    blank = run_script("xlsx_pdf.py", "--in", empty, "--out", work / "blank.pdf")
    raw = run_script("xlsx_pdf.py", "--in", BOOK, "--out", work / "raw.pdf",
                     "--report", work / "raw.json")
    raw_report = json.loads((work / "raw.json").read_text(encoding="utf-8")) \
        if (work / "raw.json").exists() else {}
    png_size = None
    imgs = sorted(pages.glob("*.png")) if pages.is_dir() else []
    if imgs:
        try:
            import fitz
            png_size = fitz.Pixmap(str(imgs[0])).width
        except Exception:  # noqa: BLE001
            png_size = None
    ctx["render"] = {
        "exit": p13.returncode, "report": pdf_report,
        "pdf_exists": pdf.is_file(), "images": len(imgs), "png_width": png_size,
        "blank_exit": blank.returncode, "blank_stderr": blank.stderr.strip(),
        "blank_wrote": (work / "blank.pdf").exists(),
        "uncalc_warning": raw_report.get("warning"),
        "uncalc_count": raw_report.get("uncalculated_formulas"),
    }

    # --- X14 finance convention ------------------------------------------------
    fin = work / "finance.xlsx"
    chk = run_script("xlsx_finance.py", "--in", calc, "--check")
    run_script("xlsx_finance.py", "--in", calc, "--out", fin, "--apply")
    rechk = run_script("xlsx_finance.py", "--in", fin, "--check",
                       "--fail-on-violation")
    ctx["finance"] = {
        "before": json.loads(chk.stdout),
        "after": json.loads(rechk.stdout),
        "recheck_exit": rechk.returncode,
        "font_before": style_probe(calc, "利润表", ["B3"])["B3"],
        "font_after": style_probe(fin, "利润表", ["B3"])["B3"],
        "custom_parts": sum(1 for n in parts_of(fin) if n.startswith("customXml")),
    }

    # --- stdout budget and the in-place contract -------------------------------
    big = work / "big.xlsx"
    big_workbook(big, SCALE_ROWS)
    scale = run_script("xlsx_read.py", "--in", big, "--sheet", "明细",
                       "--range", f"A1:C{SCALE_ROWS}", "--out", work / "big.json")
    big_json = json.loads((work / "big.json").read_text(encoding="utf-8"))
    in_place = {}
    for script, args in (
            ("xlsx_write.py", ("--in", BOOK, "--out", BOOK, "--set", "B3=1")),
            ("xlsx_read.py", ("--in", work / "nope.xlsx",)),
    ):
        p = run_script(script, *args)
        in_place[script] = {"exit": p.returncode, "stderr": p.stderr.strip()}
    ctx["scale"] = {
        "stdout_bytes": len(scale.stdout.encode("utf-8")),
        "file_cells": len(big_json.get("cells", [])),
        "note": json.loads(scale.stdout).get("cells_note", ""),
        # Stat'd here, by the gate, not read out of the report — the size in the note
        # is the thing under test and cannot also be the evidence for it.
        "file_bytes": (work / "big.json").stat().st_size,
        "in_place": in_place,
    }
    return ctx


# ── the assertions ────────────────────────────────────────────────────────────
CHECKS: dict[str, dict] = {}


def check(cid: str, title: str):
    def deco(fn):
        # A repeated id silently REPLACED the earlier check, taking its assertion out
        # of the run while the pass count went up. Caught 2026-08-16 by a control arm
        # that stopped firing for no visible reason: a second N9 had removed the
        # first. Deleting an assertion has to be louder than adding one.
        if cid in CHECKS:
            raise SystemExit(f"duplicate check id {cid!r}: {CHECKS[cid]['fn'].__name__}"
                             f" would be replaced by {fn.__name__}")
        CHECKS[cid] = {"id": cid, "title": title, "fn": fn}
        return fn
    return deco


@check("V0", "the fixtures actually give every assertion something to look at")
def v0_not_vacuous(ctx: dict) -> list[str]:
    """A silent check and a check with no subjects look identical from the outside.

    L1 shipped a green C4 while zero capabilities were declared; the same shape is
    available here the moment a fixture loses its CJK column or its second sheet.
    """
    out = []
    read = ctx["read"]["json"]
    formula_cells = [c for c in read["cells"] if c["formula"]]
    for label, got, need in (
            ("sheets in the inventory", len(read["sheets"]), 2),
            ("cells in the window", len(read["cells"]), 20),
            ("formula cells to report two views of", len(formula_cells), 3),
            ("rejected read requests", len(ctx["read"]["rejects"]), 4),
            ("parts in the input package", len(ctx["write"]["parts_before"]), 15),
            ("columns autofit had to widen", len(ctx["autofit"]["report"]["widths"]), 4),
            ("formulas in the clean audit", ctx["audit"]["clean"]["counts"]["formulas"], 5),
            # A property of the constructed INPUT, not of what the audit found —
            # counting findings here would make V0 fire on every detection defect
            # A1-A6 already own, and the matrix could no longer say which broke.
            ("formulas in the broken workbook",
             ctx["audit"]["broken"]["counts"]["formulas"], 8),
    ):
        if got < need:
            out.append(f"V0 only {got} {label} (need at least {need}) — the "
                       f"assertions that rest on them cannot fail")
    # S2's whole discriminating power rests on the fixture's font NOT being the
    # openpyxl default: a wholesale Font() assignment lands on Calibri 11, so if
    # B3 already were Calibri 11 the negative control would be a no-op that reads
    # as a pass. The first version of this fixture was exactly that.
    b3 = ctx["rebuild"]["style_before"]["B3"]
    if b3["name"] in (None, "Calibri") or b3["size"] in (None, 11.0):
        out.append(f"V0 the fixture's B3 font is {b3['name']!r} {b3['size']} — the "
                   f"openpyxl default, so S2's control cannot tell a preserved font "
                   f"from a reset one")
    return out


@check("R1", "read reports BOTH views of a formula cell: the value and the formula")
def r1_two_views(ctx: dict) -> list[str]:
    cells = {c["ref"]: c for c in ctx["read"]["json"]["cells"]}
    out = []
    target = cells.get("D3")
    if target is None:
        return ["R1 D3 is not in the window at all"]
    if not target["formula"]:
        out.append("R1 D3 holds a formula in the fixture and read reported none — a "
                   "reader that only returns values makes every formula invisible")
    if cells.get("B3", {}).get("value") != 1240:
        out.append(f"R1 B3 value is {cells.get('B3', {}).get('value')!r}, "
                   f"expected the stored 1240")
    return out


@check("R2", "an uncalculated formula is flagged, not reported as an empty cell")
def r2_uncalculated(ctx: dict) -> list[str]:
    cells = {c["ref"]: c for c in ctx["read"]["json"]["cells"]}
    # "there are no formula cells at all" is V0's finding, not this one — a check
    # that also reports its own vacuity cannot be told apart from the guard that
    # owns it, and the matrix then shows two checks for one defect.
    formulas = [c for c in cells.values() if c["formula"]]
    # openpyxl writes no cached values, so every formula in the fixture is
    # uncalculated. Reporting value=None without saying so is how an agent concludes
    # "the sheet is empty" from a file that is merely uncomputed.
    unflagged = [c["ref"] for c in formulas if c["value"] is None
                 and not c["uncalculated"]]
    if unflagged:
        return [f"R2 {', '.join(unflagged)} have no cached value and are not marked "
                f"`uncalculated` — indistinguishable from a genuinely empty cell"]
    return []


@check("R3", "an unreadable request is refused with a sentence, not clamped or crashed")
def r3_refusals(ctx: dict) -> list[str]:
    out = []
    for case in ctx["read"]["rejects"]:
        if case["exit"] != 2:
            out.append(f"R3 {case['case']}: exit {case['exit']}, expected 2 "
                       f"(a silently clamped range answers a question nobody asked)")
        elif not case["stderr"].startswith("error: ") or "Traceback" in case["stderr"]:
            out.append(f"R3 {case['case']}: stderr is not one actionable line: "
                       f"{case['stderr'][:80]!r}")
    return out


@check("R4", "the sheet inventory names every sheet and counts its formulas")
def r4_inventory(ctx: dict) -> list[str]:
    sheets = {s["name"]: s for s in ctx["read"]["json"]["sheets"]}
    out = []
    for name in SHEETS:
        if name not in sheets:
            out.append(f"R4 sheet {name!r} is missing from the inventory")
    summary = sheets.get("汇总")
    if summary and summary["formulas"] < 2:
        out.append(f"R4 汇总 reports {summary['formulas']} formulas; the fixture's "
                   f"cross-sheet references are what X10 will rest on")
    return out


@check("W1", "a surgical edit loses no part of the input")
def w1_no_part_lost(ctx: dict) -> list[str]:
    w = ctx["write"]
    lost = sorted(set(w["parts_before"]) - set(w["parts_after"]))
    out = []
    if lost:
        out.append(f"W1 the edit dropped {', '.join(lost)} — present in the input, "
                   f"gone from the output")
    if w["report"]["parts_lost"]:
        out.append(f"W1 the report itself admits losing {w['report']['parts_lost']}")
    # The control arm, measured rather than assumed: if the obvious implementation
    # loses nothing either, this whole capability is solving a problem that is not
    # there and the claim should be withdrawn, not asserted.
    naive_lost = set(w["parts_before"]) - set(w["naive_parts_after"])
    if not naive_lost:
        out.append("W1 the load→save control lost nothing, so this fixture cannot "
                   "demonstrate the difference the surgical path exists for")
    return out


@check("W2", "parts the edit did not touch come out byte-identical")
def w2_bytes_untouched(ctx: dict) -> list[str]:
    w = ctx["write"]
    total = len(w["parts_before"])
    kept, naive_kept = len(w["identical"]), len(w["naive_identical"])
    out = []
    # Only sheet1.xml should differ. Anything else means the writer rebuilt a part it
    # had no business rebuilding, and a rebuild is where content quietly changes.
    changed = sorted(set(w["parts_before"]) - set(w["identical"]))
    if changed != ["xl/worksheets/sheet1.xml"]:
        out.append(f"W2 the edit rewrote {changed}; only the edited worksheet should "
                   f"differ from the input")
    if kept <= naive_kept:
        out.append(f"W2 the surgical path kept {kept}/{total} parts byte-identical and "
                   f"load→save kept {naive_kept}/{total} — no measurable advantage")
    return out


@check("W3", "the edit reaches the named cell and keeps its formatting")
def w3_edit_lands(ctx: dict) -> list[str]:
    w = ctx["write"]
    out = []
    # Deliberately NOT the appended row: W4 owns that, and a check that also looks
    # at A6 fires on every append defect too, which makes the matrix unable to say
    # which of the two is broken.
    for ref, want in (("B3", 1310), ("D5", "=B5/C5-1")):
        got = w["cells"].get(ref)
        if got != want:
            out.append(f"W3 {ref} = {got!r} after the edit, expected {want!r}")
    if w["styles"]["D5"] != w["styles_before"]["D5"]:
        out.append(f"W3 D5's style index changed from {w['styles_before']['D5']!r} to "
                   f"{w['styles']['D5']!r} — its number format is an attribute of the "
                   f"cell, not of the value that was in it")
    return []if not out else out


@check("W4", "append-row lands after the last used row, overwriting nothing")
def w4_append(ctx: dict) -> list[str]:
    w = ctx["write"]
    # The fixture's last used row is 5 (毛利), so the appended row must be 6 and row
    # 5 must be untouched.
    if w["cells"].get("A6") != "其他业务收入":
        return [f"W4 the appended row is not at row 6 (A6 = {w['cells'].get('A6')!r})"]
    if w["cells"].get("B5") != "=B3-B4":
        return [f"W4 appending overwrote row 5: B5 = {w['cells'].get('B5')!r}"]
    return []


@check("W5", "overwriting a shared-formula master is refused, not silently done")
def w5_shared_formula(ctx: dict) -> list[str]:
    s = ctx["shared"]
    out = []
    if s["master_exit"] != 2:
        out.append(f"W5 overwriting the master of a shared formula exited "
                   f"{s['master_exit']}; the cells sharing it would be left with a "
                   f"reference to a definition that no longer exists")
    if s["master_wrote"]:
        out.append("W5 a file was written despite the refusal")
    if s["other_exit"] != 0:
        out.append(f"W5 editing an unrelated cell in the same workbook exited "
                   f"{s['other_exit']} — the guard is refusing whole files, not the "
                   f"one cell it is about")
    return out


@check("W6", "the calcChain cache is dropped on a formula edit and only then")
def w6_calcchain(ctx: dict) -> list[str]:
    c = ctx["calcchain"]
    out = []
    if not c["input_has"]:
        return ["W6 the constructed input carries no calcChain, so nothing is proven"]
    if c["after_formula_edit"]:
        out.append("W6 a stale calcChain survived a formula edit — this is what makes "
                   "Excel report a perfectly good file as damaged")
    if not c["after_value_edit"]:
        out.append("W6 the calcChain was dropped on a VALUE-only edit; the dependency "
                   "order did not change, so throwing it away costs a recalculation "
                   "for nothing")
    return out


@check("W7", "damage the edit did not cause is reported, not treated as fatal")
def w7_pre_existing(ctx: dict) -> list[str]:
    d = ctx["damage"]
    out = []
    if d["exit"] != 0 or not d["wrote"]:
        out.append(f"W7 a workbook that arrived with a dangling relationship was "
                   f"refused (exit {d['exit']}) — the files that most need editing "
                   f"are the damaged ones")
    if not d["pre_existing"]:
        out.append("W7 the pre-existing damage was not reported at all, so the user "
                   "has no way to know the file came in broken")
    return out


# Worked out by hand from book.xlsx and pinned here on purpose: deriving them from
# the fixture's formulas would re-implement the dependency walk that is under test.
# 利润表!B3 is read by D3 (=B3/C3-1) and B5 (=B3-B4); B5 is read by D5 and by
# 汇总!B2, which 汇总!B3 reads in turn — so the closure crosses a sheet boundary.
STALE_EXPECTED = {"利润表!D3", "利润表!B5", "利润表!D5", "汇总!B2", "汇总!B3"}
# Cells the edit cannot reach. If these lose their cached values the implementation
# is clearing every formula in the book — honest, but it throws away numbers that
# are still true.
STALE_UNTOUCHED = {"利润表!D4", "利润表!C5"}


@check("W8", "an edit clears the cached results it invalidated, and only those")
def w8_stale_caches(ctx: dict) -> list[str]:
    """A cached `<v>` over a changed input is not a stale number with a warning next
    to it — it is a WRONG number that reports itself as calculated.

    Measured 2026-08-16 on the L4 input (059 §三十): `--set B2=1350` left 毛利 488.2
    and 营业利润 47.6 in the file, `xlsx_read` marked both `uncalculated: false`, and
    the PDF preview put that table on the page — LibreOffice renders stale caches
    verbatim because it does not honour `fullCalcOnLoad`. Clearing them makes the
    same render produce the right numbers, because a formula cell with no cached
    result IS computed at load.

    Two-sided on purpose: clearing EVERY formula cache satisfies the first half and
    is the lazy implementation the second half exists to reject.
    """
    s = ctx["stale"]
    out = []
    before, after = s["cached_before"], s["cached_after"]
    absent = (STALE_EXPECTED | STALE_UNTOUCHED) - before
    if absent:
        return [f"W8 the fixture carries no cached value for {sorted(absent)}, so "
                f"this assertion is looking at nothing — a workbook with no caches "
                f"cannot show a stale one"]
    if s["exit"] != 0:
        return [f"W8 the edit exited {s['exit']}"]
    left_stale = sorted(STALE_EXPECTED & after)
    if left_stale:
        out.append(f"W8 {left_stale} were computed from the edited cell and still "
                   f"carry their old cached result — a reader sees those numbers as "
                   f"current, and LibreOffice renders them onto the page")
    over_cleared = sorted(STALE_UNTOUCHED - after)
    if over_cleared:
        out.append(f"W8 {over_cleared} do not depend on the edited cell and lost "
                   f"their cached values anyway — that discards numbers that are "
                   f"still true")
    reported = set(s["report"].get("caches_invalidated") or [])
    if reported != STALE_EXPECTED:
        out.append(f"W8 the report claims it invalidated {sorted(reported)} while "
                   f"the file says {sorted(before - after)} — a report that does not "
                   f"match the artifact is the failure this check is about")
    if not s["report"].get("caches_invalidated_note"):
        out.append("W8 caches were cleared and nothing in the report says so; the "
                   "caller finds out by reading an empty cell")
    return out


@check("W9", "the staleness sweep is bounded, and says so when it falls back")
def w9_stale_bound(ctx: dict) -> list[str]:
    """The precise walk expands every reference to single cells, which is quadratic
    on a shape real workbooks have.

    Measured 2026-08-16 (059 §三十): the first version of the sweep ran **over nine
    minutes** on a 10,000-row sheet of widening SUM ranges and had to be killed. A
    correctness fix that turns a sub-second command into a hang is not a fix. The
    fallback has to be safe as well as fast: clearing more than necessary throws away
    true numbers, but leaving one stale value behind is the original defect.
    """
    s = ctx["stale_bound"]
    out = []
    if len(s["cached_before"]) < WIDE_FORMULAS:
        return [f"W9 the fixture carries {len(s['cached_before'])} cached formula "
                f"cell(s), not {WIDE_FORMULAS} — nothing here can go stale, so the "
                f"assertion would pass on an implementation that does nothing"]
    if s["exit"] != 0:
        return [f"W9 the edit exited {s['exit']} on a workbook with a wide "
                f"dependency graph"]
    if s["seconds"] > WIDE_SECONDS_CEILING:
        out.append(f"W9 the edit took {s['seconds']:.1f}s (ceiling "
                   f"{WIDE_SECONDS_CEILING:.0f}s) — the dependency walk is being run "
                   f"on a graph it cannot afford to expand")
    if s["cached_after"]:
        out.append(f"W9 {len(s['cached_after'])} formula cell(s) kept a cached "
                   f"result the edit invalidated; falling back is allowed, leaving a "
                   f"wrong number behind is not")
    if not s["report"].get("caches_invalidated_coarse"):
        out.append("W9 the sweep gave up precision and the report does not say so — "
                   "the caller cannot tell 'only the dependents were cleared' from "
                   "'everything was cleared' without being told")
    return out


@check("C1", "column width counts a CJK character as two units, not one")
def c1_display_width(ctx: dict) -> list[str]:
    got = ctx["autofit"]["widths"].get("A")
    if got is None:
        return ["C1 column A got no explicit width at all"]
    if abs(got - LONG_LABEL_WIDTH) > 0.01:
        hint = (" — that is len()+2, which is exactly half the room the label needs"
                if abs(got - LONG_LABEL_LEN_WIDTH) < 0.01 else "")
        return [f"C1 column A width {got:g}, expected {LONG_LABEL_WIDTH} for "
                f"{LONG_LABEL!r}{hint}"]
    return []


@check("C2", "a formula's own text does not drive the width of its column")
def c2_formula_text(ctx: dict) -> list[str]:
    # 汇总!B holds =B2/利润表!B3 (14 characters). Counting it would demand ~16 units;
    # the widest DISPLAYED text in that column is 取值 at 6.
    widths = {(w["sheet"], w["column"]): w["width"]
              for w in ctx["autofit"]["report"]["widths"]}
    got = widths.get(("汇总", "B"))
    if got is None:
        return ["C2 汇总!B was not widened, so nothing is proven"]
    if got > 8:
        return [f"C2 汇总!B width {got:g} is wider than its widest displayed value; "
                f"the formula text is being measured and it is never shown"]
    return []


@check("C3", "a title merged across columns does not drive the first one; a vertical "
             "merge does")
def c3_merges(ctx: dict) -> list[str]:
    got = ctx["merge"]["width_a"]
    if got is None:
        return ["C3 column A got no width in the merge probe"]
    if abs(got - MERGE_EXPECT_WIDTH) < 0.01:
        return []
    if abs(got - MERGE_TITLE_WIDTH) < 0.01:
        return [f"C3 column A width {got:g} — that is the title merged across A:F, "
                f"which is DISPLAYED across six columns and must not size the first "
                f"one on its own"]
    if abs(got - MERGE_PLAIN_WIDTH) < 0.01:
        return [f"C3 column A width {got:g} — every merged cell was skipped, but a "
                f"merge inside a single column buys no horizontal room and still "
                f"has to fit ({MERGE_VERTICAL!r} needs {MERGE_EXPECT_WIDTH})"]
    return [f"C3 column A width {got:g}, expected {MERGE_EXPECT_WIDTH} "
            f"for the vertically merged {MERGE_VERTICAL!r}"]


@check("C4", "importing a CSV with --autofit puts the widths in the FILE, not only "
             "in the report")
def c4_import_autofit(ctx: dict) -> list[str]:
    """`--from x.csv --autofit` reported `widths_set: 5` and wrote no <cols> at all.

    openpyxl's write_only sheet streams out as rows arrive and emits <cols> at the
    START of it, so widths set after the first append are dropped without a word.
    Nothing here or anywhere else was looking: this is the creation path, and every
    width assertion above tests the EDIT path (`xlsx_write.py --autofit`), which
    writes the sheet XML surgically and was always fine. Two different code paths,
    one of them uncovered.
    """
    out = []
    got = ctx["import_autofit"]
    if got["exit"] != 0:
        out.append(f"C4 the import exited {got['exit']} — no widths to check")
        return out
    claimed = got["report"].get("widths_set")
    if not claimed:
        out.append(f"C4 --autofit reported widths_set={claimed!r} on a CSV that has "
                   f"text in every column")
    # The file is the fact; the report is a claim about it.
    if got["cols_in_file"] != claimed:
        out.append(f"C4 report says widths_set={claimed} but the sheet carries "
                   f"{got['cols_in_file']} <col> entr(ies) — a width that is not in "
                   f"the artifact is not a width")
    if got["report"].get("widths_in_file") != got["cols_in_file"]:
        out.append(f"C4 report's widths_in_file={got['report'].get('widths_in_file')!r} "
                   f"disagrees with the file's {got['cols_in_file']}")
    # And they must be CJK-aware, or this passes with widths that truncate.
    widest = max(got["widths"].values(), default=0)
    if widest < CJK_IMPORT_MIN_WIDTH:
        out.append(f"C4 widest imported column is {widest}, expected >= "
                   f"{CJK_IMPORT_MIN_WIDTH} for a header counted in wide characters "
                   f"(got {got['widths']})")
    return out


@check("C5", "a number too wide for its column is measured through its format")
def c5_number_width(ctx: dict) -> list[str]:
    """`###` is what a too-narrow NUMBER column shows, and the width routine used to
    skip every cell that was not text.

    Measured 2026-08-16 (059 §三十二): in 利润表 the only cell that actually failed
    to display was 营业利润's 同比 — `1.0877…` under `0.0%`, i.e. `108.8%`, six
    characters in a six-unit column — and it is precisely the kind of cell
    `--autofit` could not see. The判据 is taken off the PAGE, not out of the report:
    the report is the thing under test.
    """
    p = ctx["percent_width"]
    out = []
    if p["width_before"] != 6:
        return [f"C5 the fixture's percent column starts at {p['width_before']}, not "
                f"6 — it has to be too narrow or there is nothing to widen"]
    if p["rendered"] is None:
        SKIPS.append("C5 rendered check: pypdfium2 is not installed")
    elif "###" in p["rendered"]:
        out.append("C5 the rendered page still shows ### — the column was not made "
                   "wide enough for the number it displays")
    elif "108.8%" not in p["rendered"]:
        out.append(f"C5 neither ### nor 108.8% is on the page; the fixture did not "
                   f"render what this check is about")
    if (p["width_after"] or 0) <= 6:
        out.append(f"C5 the column is still {p['width_after']} units wide; the value "
                   f"it displays needs 8")
    widened = [w for w in p["report"].get("widths", [])
               if w.get("column") == "B" and w.get("reason", "").startswith("'108.8%'")]
    if not widened:
        out.append(f"C5 nothing in the report attributes the widening to the "
                   f"DISPLAYED text: {p['report'].get('widths')}")
    return out


@check("C6", "a run that widened nothing still says what it measured")
def c6_measured_when_nothing_changed(ctx: dict) -> list[str]:
    """An empty `changes` and a `--autofit` that did nothing look identical.

    Measured 2026-08-16 (059 §三十二): handed `changes: []` twice, a model went
    around the skill and set five widths by hand — larger than its own reading of
    this skill's formula said were needed — and split the table across two pages.
    """
    noop = ctx["percent_width"]["noop"]
    out = []
    if noop.get("widths"):
        return [f"C6 the second pass still widened {noop['widths']}; this check needs "
                f"the run that changed nothing"]
    measured = noop.get("widths_measured") or []
    if not measured:
        out.append("C6 a run that changed nothing reports nothing it measured — "
                   "indistinguishable from --autofit never having run")
    elif not any(m.get("verdict") == "already wide enough" and m.get("current")
                 for m in measured):
        out.append(f"C6 the measurements carry no current-vs-needed verdict: "
                   f"{measured[:2]}")
    if not noop.get("widths_note"):
        out.append("C6 nothing in the report says the columns were checked and found "
                   "wide enough")
    return out


@check("N10", "a table split across pages by its own width is reported")
def n10_split_columns(ctx: dict) -> list[str]:
    """Page 1 looks like a complete table. It is not: the columns that did not fit
    are on a page of their own, with no row labels beside them.
    """
    s = ctx["split"]
    out = []
    if s.get("pages") in (None, 0):
        SKIPS.append("N10 split check: the preview reported no page count")
        return []
    if s.get("pages") < 2:
        return [f"N10 the wide fixture rendered on {s.get('pages')} page(s); this "
                f"check needs a sheet that does not fit"]
    off = s.get("columns_off_first_page")
    if off is None:
        out.append("N10 the report does not say whether any column left page 1")
    elif "同比" not in off:
        out.append(f"N10 同比 is on page 2 of this preview and the report lists {off}")
    if not s.get("split_warning"):
        out.append("N10 columns were moved to their own page and no warning says so — "
                   "page 1 reads as the whole table")
    return out


@check("N11", "a number rendered as ### is reported, and not guessed at")
def n11_hash_marks(ctx: dict) -> list[str]:
    """`###` passes every other check on this preview: the page has ink, one page,
    no column moved off it — and the number the reader came for is not there.

    Measured 2026-08-16 (059 §三十三): the B9 deliverable rendered 营业利润's 同比,
    108.8%, as ### while every field in the report said the preview was fine.
    """
    h = ctx["hash"]
    out = []
    narrow, literal = h.get("narrow", {}), h.get("literal", {})
    if narrow.get("pages") is None:
        SKIPS.append("N11: the narrow preview reported no page count")
        return []
    if narrow.get("blank_pages") or narrow.get("columns_off_first_page"):
        return [f"N11 the narrow fixture also trips the blank/split checks "
                f"({narrow.get('blank_pages')}, "
                f"{narrow.get('columns_off_first_page')}) — then it does not show "
                f"that THIS check catches something they miss"]
    if narrow.get("hash_marked_cells") is not True:
        out.append(f"N11 a page showing ### was reported as "
                   f"{narrow.get('hash_marked_cells')!r}")
    if "利润表!B" not in (narrow.get("hash_marked_columns") or []):
        out.append(f"N11 the too-narrow column was not named: "
                   f"{narrow.get('hash_marked_columns')}")
    if not narrow.get("hash_warning"):
        out.append("N11 nothing warns that a value is in the file and not in the "
                   "picture")
    # The other half: a sheet that really holds "###" must not be called truncated.
    if literal.get("hash_marked_cells") is not None:
        out.append(f"N11 a sheet whose cell literally contains ### was judged "
                   f"{literal.get('hash_marked_cells')!r} instead of reported as "
                   f"undecidable")
    elif not literal.get("hash_marked_note"):
        out.append("N11 the undecidable case says nothing about why")
    return out


@check("N12", "a shorter render does not leave the previous one's pages behind")
def n12_stale_images(ctx: dict) -> list[str]:
    """Measured 2026-08-16 (059 §三十三): a 1-page preview landed in a folder that
    still held page-002.png from a 2-page render, and the report listed only the one
    image it wrote. The leftover reads as part of this document.
    """
    s = ctx["stale_images"]
    left, report = s["left"], s["report"]
    out = []
    if report.get("pages") != 1:
        return [f"N12 the fixture rendered {report.get('pages')} page(s); this check "
                f"needs a render shorter than what was already in the directory"]
    orphans = [n for n in left if n.startswith("page-") and n != "page-001.png"]
    if orphans:
        out.append(f"N12 {orphans} survived a render that produced one page")
    if "notes.png" not in left:
        out.append("N12 notes.png was deleted — only the page-NNN.png names this "
                   "script itself writes may be removed")
    removed = report.get("stale_images_removed") or []
    if sorted(removed) != ["page-002.png", "page-009.png"]:
        out.append(f"N12 the report says it removed {removed}; deleting a user's "
                   f"files without listing them is the silent part of this")
    return out


@check("E1","cols is written before sheetData, as the ECMA-376 sequence requires")
def e1_element_order(ctx: dict) -> list[str]:
    order = ctx["autofit"]["sheet_order"]
    if "cols" not in order:
        return ["E1 the autofit output has no cols element"]
    if "sheetData" not in order:
        return ["E1 the autofit output has no sheetData element"]
    if order.index("cols") > order.index("sheetData"):
        return [f"E1 cols appears after sheetData ({order}); CT_Worksheet is an "
                f"xsd:sequence and Excel offers to repair a file that gets it wrong"]
    return []


@check("O1", "stdout stays a summary on a large workbook")
def o1_stdout_budget(ctx: dict) -> list[str]:
    s = ctx["scale"]
    out = []
    if s["stdout_bytes"] > STDOUT_BUDGET:
        out.append(f"O1 reading {SCALE_ROWS} rows printed {s['stdout_bytes']} bytes to "
                   f"stdout (budget {STDOUT_BUDGET}) — that goes straight into the "
                   f"agent's context, and again across a Team delegation")
    if s["file_cells"] < SCALE_ROWS * 3:
        out.append(f"O1 the file written by --out holds only {s['file_cells']} cells; "
                   f"trimming stdout must not mean dropping the data")
    return out


@check("O3", "the pointer to the untrimmed file states that file's size")
def o3_pointer_size(ctx: dict) -> list[str]:
    """Trimming stdout MOVES the bytes; the note decides whether they come back.

    Measured in a live session on 2026-08-16: handed `"the full list is in <path>"`,
    a model answered `--out /tmp/cells.json && cat /tmp/cells.json` and pulled 5,752
    bytes through a trim that had held stdout to 259. The same move on a 2000x30
    sheet is 6,439,603 bytes. Whether following the pointer is safe is decided by one
    number — the size of the file — and that is the number the message omitted.
    """
    s = ctx["scale"]
    note, size = s["note"], s["file_bytes"]
    if not note:
        return ["O3 the trimmed report carries no pointer to the full data at all"]
    if size < 10_000:
        # Non-vacuity: on a tiny file the byte count could collide with a row number
        # or a cell count already in the note and pass for the wrong reason.
        return [f"O3 the --out file is only {size} bytes — too small for this "
                f"assertion to distinguish a stated size from a coincidence"]
    if str(size) not in note:
        return [f"O3 the note points at a {size}-byte file without saying how big it "
                f"is ({note[:110]!r}) — an agent decides whether to print that file "
                f"back into the conversation, and this is the message it decides on"]
    return []


@check("O4", "a run that replaced an existing file says so, and one that did not")
def o4_replaced_existing(ctx: dict) -> list[str]:
    """`--out == --in` is refused. A different path holding somebody else's bytes is
    not, and nothing in this skill said a word about it.

    The docx skill grew this field in §二十四 after a preview overwrote a fixture;
    xlsx had seven scripts that write files and not one of them mentioned it. Two
    directions on purpose — a field that is always True says nothing, which is the
    trap §二十四 named when it added the same thing next door.

    Three scripts are exercised by running them; the rest are covered structurally,
    because inventing valid arguments for every writer is how coverage lists stop
    matching the code.
    """
    r = ctx["replaced"]
    out = []
    for name, seen in sorted(r["runs"].items()):
        if seen != [False, True]:
            out.append(f"O4 {name} reported {seen} for [fresh path, same path again]; "
                       f"expected [False, True] — one value for both cases is not an "
                       f"answer")
    if r["scripts_without_the_field"]:
        out.append(f"O4 {r['scripts_without_the_field']} guard their output with "
                   f"ensure_distinct and never report replaced_existing")
    return out


@check("O2", "an output over its own input, or a missing file, gets one sentence")
def o2_contracts(ctx: dict) -> list[str]:
    out = []
    for script, r in ctx["scale"]["in_place"].items():
        if r["exit"] != 2:
            out.append(f"O2 {script} exited {r['exit']}, expected 2")
        elif "Traceback" in r["stderr"] or not r["stderr"].startswith("error: "):
            out.append(f"O2 {script} answered with {r['stderr'][:90]!r} instead of one "
                       f"actionable sentence")
    return out


@check("A1", "the audit finds an error token and names what caused it")
def a1_error(ctx: dict) -> list[str]:
    hits = [f for f in ctx["audit"]["broken"]["findings"]
            if f["class"] == "error" and f["cell"] == "利润表!C1"]
    if not hits:
        return ["A1 =#REF!*2 was not reported as an error"]
    if not hits[0].get("cause"):
        return ["A1 the error was reported with no cause — a token alone does not "
                "tell anyone which reference broke"]
    return []


@check("A2", "the audit finds a reference to a sheet that does not exist")
def a2_missing_sheet(ctx: dict) -> list[str]:
    """The #REF! that has not happened yet.

    Nothing in this stack refuses to STORE `=预算表!A1`; it becomes an error the
    first time Excel opens the file, long after anyone remembers writing it.
    """
    hits = [f for f in ctx["audit"]["broken"]["findings"]
            if f["class"] == "missing" and f["cell"] == "利润表!B1"]
    if not hits:
        return ["A2 a formula pointing at the non-existent sheet 预算表 was not "
                "reported"]
    if hits[0].get("sheet") != "预算表":
        return [f"A2 reported the wrong sheet name: {hits[0].get('sheet')!r}"]
    return []


@check("A3", "the audit finds circular chains, including a one-cell self-reference")
def a3_cycles(ctx: dict) -> list[str]:
    cycles = [f for f in ctx["audit"]["broken"]["findings"] if f["class"] == "circular"]
    starts = {f["cell"] for f in cycles}
    out = []
    if "利润表!F1" not in starts:
        out.append("A3 the self-reference =F1 was not reported")
    if not ({"利润表!D1", "利润表!E1"} & starts):
        out.append("A3 the two-cell cycle D1<->E1 was not reported")
    for f in cycles:
        if not f.get("chain") or "->" not in f["chain"]:
            out.append(f"A3 {f['cell']} was reported without the chain; a cycle with "
                       f"no path is not actionable")
    return out


@check("A4", "an uncalculated formula is reported, not counted as clean")
def a4_uncalc(ctx: dict) -> list[str]:
    clean = ctx["audit"]["clean"]
    n = clean["by_class"]["uncalc"]
    if n != clean["counts"]["formulas"]:
        return [f"A4 {n} of {clean['counts']['formulas']} formulas were reported as "
                f"uncalculated; a library-written workbook has NO cached values, so "
                f"every one of them should be"]
    return []


@check("A5", "the audit stays silent on a clean workbook, cross-sheet links included")
def a5_clean(ctx: dict) -> list[str]:
    clean = ctx["audit"]["clean"]
    noisy = {c: n for c, n in clean["by_class"].items()
             if c != "uncalc" and n}
    if noisy:
        return [f"A5 the clean fixture produced {noisy} — the 汇总 sheet's "
                f"cross-sheet formulas are legitimate and must not be findings"]
    if ctx["audit"]["clean_exit"] != 0:
        return [f"A5 auditing a clean workbook exited {ctx['audit']['clean_exit']}"]
    return []


@check("A6", "the reference parser ignores string literals and function names")
def a6_false_positives(ctx: dict) -> list[str]:
    """Both of these produced a false finding before they were handled.

    G1 holds `="见 预算表!A1 的说明"` — a sheet name inside a STRING, which is text,
    not a link. H1 holds `=LOG10(A1)`, where LOG10 matches the shape of a cell
    reference exactly; left alone it becomes a graph node and can invent a circular
    reference in a workbook that happens to use cell LOG10.
    """
    out = []
    flagged = ctx["audit"]["cells"]
    if "利润表!G1" in flagged:
        out.append("A6 a sheet name inside a string literal was treated as a "
                   "reference to a missing sheet")
    if "利润表!H1" in flagged:
        out.append("A6 LOG10( was parsed as a reference to cell LOG10")
    if "汇总!A1" in flagged:
        out.append("A6 a legitimate cross-sheet reference was flagged")
    return out


@check("A7", "the audit report says what it did NOT check, clean or not")
def a7_scope_in_the_report(ctx: dict) -> list[str]:
    """A clean audit proves every reference resolves. It proves nothing about the
    numbers, and `findings: []` does not say so.

    SKILL.md states this boundary twice, in bold, in the words the acceptance
    criterion asks for. Measured 2026-08-16 (059 §三十一): a model with the entire
    document in context still answered "没有任何问题 / 所有公式引用均合法" — while
    transcribing "45 个单元格，其中 17 个是公式" straight out of `counts`. So the
    statement belongs in the artifact being quoted, not only in the document.

    ⚠️ This assertion checks that the report CARRIES the scope. Nothing here — and
    nothing that can be written — checks that a model passes it on. See 059 §六·补八.
    """
    out = []
    spotless = ctx["audit"]["spotless"]["report"]
    broken = ctx["audit"]["broken"]
    if spotless.get("findings") or any(spotless.get("by_class", {"x": 1}).values()):
        return [f"A7 the spotless fixture reports "
                f"{spotless.get('by_class')} — this assertion is about the report "
                f"with NOTHING in it, and that is not this one"]
    if not spotless.get("counts", {}).get("formulas"):
        return ["A7 the spotless fixture has no formulas, so `formulas_evaluated: 0` "
                "sits next to nothing and says nothing"]
    for label, report in (("clean", spotless), ("with findings", broken)):
        scope = report.get("scope")
        if not scope:
            out.append(f"A7 the {label} report does not say what it checked; "
                       f"`findings` alone reads as 'no problems', and on the clean "
                       f"report that is exactly when it is over-read")
        elif "xlsx_recalc" not in scope:
            out.append(f"A7 the {label} report draws the boundary without naming "
                       f"what to run instead")
        if report.get("counts", {}).get("formulas_evaluated") != 0:
            out.append(f"A7 the {label} report has no `formulas_evaluated: 0` beside "
                       f"its formula count — a number is the part a reader copies")
    return out


@check("Q1", "a formula naming a missing sheet is refused, and nothing is written")
def q1_refuse_missing_sheet(ctx: dict) -> list[str]:
    c = ctx["refuse"]["cases"]["missing sheet"]
    out = []
    if c["exit"] != 2:
        out.append(f"Q1 writing =预算表!A1 exited {c['exit']}; storing it succeeds "
                   f"everywhere and only breaks when Excel opens the file")
    if c["wrote"]:
        out.append("Q1 a file was written despite the refusal")
    if "预算表" not in c["stderr"]:
        out.append("Q1 the refusal does not name the sheet that is missing")
    return out


@check("Q2", "a formula with an error token baked into its text is refused")
def q2_refuse_error_token(ctx: dict) -> list[str]:
    c = ctx["refuse"]["cases"]["error token baked in"]
    if c["exit"] != 2 or c["wrote"]:
        return [f"Q2 =#REF!+1 was accepted (exit {c['exit']}, wrote {c['wrote']}) — "
                f"that is an error value, not an expression"]
    return []


@check("Q3", "a legitimate cross-sheet formula is accepted")
def q3_accept_legit(ctx: dict) -> list[str]:
    """The control for Q1/Q2: a guard that refuses everything is not a guard."""
    r = ctx["refuse"]
    if r["legit_exit"] != 0:
        return [f"Q3 writing =汇总!B2 into a workbook that HAS 汇总 exited "
                f"{r['legit_exit']} — the check is rejecting valid links"]
    if r["legit_formula"] != "=汇总!B2":
        return [f"Q3 the accepted formula read back as {r['legit_formula']!r}"]
    return []


@check("S1", "formatting lands on the named range and nowhere else")
def s1_format_scope(ctx: dict) -> list[str]:
    s = ctx["rebuild"]["style"]
    out = []
    for ref in ("B3", "C5"):
        cell = s[ref]
        if cell["number_format"] != "#,##0":
            out.append(f"S1 {ref} number format is {cell['number_format']!r}")
        if not cell["bold"] or cell["color"] != "FF0000FF":
            out.append(f"S1 {ref} font is bold={cell['bold']} color={cell['color']}")
        if cell["fill"] != "FFFFF2CC":
            out.append(f"S1 {ref} fill is {cell['fill']!r}, expected FFFFF2CC")
        if cell["border"] != "thin":
            out.append(f"S1 {ref} border is {cell['border']!r}")
    # D3 is OUTSIDE B3:C5. A formatter that quietly applies to the whole sheet
    # passes every check above.
    if ctx["rebuild"]["style"]["D3"]["fill"] == "FFFFF2CC":
        out.append("S1 D3 is outside the requested range and was formatted anyway")
    return out


@check("S2", "setting one font attribute keeps the others the cell already had")
def s2_font_merge(ctx: dict) -> list[str]:
    """`Font(color=...)` replaces the whole font object.

    Assigning a fresh Font to change a colour silently resets size and typeface to
    the defaults — the "I made one number blue and the row changed size" failure.
    """
    before, after = ctx["rebuild"]["style_before"]["B3"], ctx["rebuild"]["style"]["B3"]
    out = []
    if after["size"] != before["size"]:
        out.append(f"S2 B3 font size changed from {before['size']} to {after['size']} "
                   f"while only the colour was asked for")
    if after["name"] != before["name"]:
        out.append(f"S2 B3 typeface changed from {before['name']!r} to {after['name']!r}")
    return out


@check("S3", "every conditional-format rule kind arrives, and the existing one survives")
def s3_conditional(ctx: dict) -> list[str]:
    cf, before = ctx["rebuild"]["cf"], ctx["rebuild"]["cf_before"]
    out = []
    want = {"D3:D5": "cellIs", "B3:B5": "colorScale", "C3:C5": "dataBar"}
    for rng, kind in want.items():
        kinds = cf.get(rng, [])
        if kind not in kinds:
            out.append(f"S3 no {kind} rule on {rng} (got {kinds or 'nothing'})")
    # The fixture already carried a rule on D3:D5. Adding one must not replace it.
    kept = len(cf.get("D3:D5", []))
    if kept < len(before.get("D3:D5", [])) + 1:
        out.append(f"S3 D3:D5 holds {kept} rule(s); the fixture's own rule was "
                   f"replaced rather than added to")
    return out


@check("S4", "freeze panes and auto filter are set to the requested values")
def s4_panes(ctx: dict) -> list[str]:
    p, before = ctx["rebuild"]["panes"], ctx["rebuild"]["panes_before"]
    out = []
    if p["freeze"] != "C6":
        out.append(f"S4 freeze_panes is {p['freeze']!r}, expected C6"
                   + (" — that is the fixture's own value, so nothing happened"
                      if p["freeze"] == before["freeze"] else ""))
    if p["filter"] != "A2:C5":
        out.append(f"S4 auto_filter is {p['filter']!r}, expected A2:C5"
                   + (" — the fixture's own value" if p["filter"] == before["filter"]
                      else ""))
    return out


@check("S5", "a chart is added, named from the header, pointing at the requested data")
def s5_chart(ctx: dict) -> list[str]:
    c, before = ctx["rebuild"]["chart"], ctx["rebuild"]["chart_before"]
    out = []
    if c["count"] != before["count"] + 1:
        out.append(f"S5 汇总 holds {c['count']} chart(s), was {before['count']} — "
                   f"the new one replaced the fixture's instead of joining it")
    if not c["series_titles"] or not all(c["series_titles"]):
        out.append("S5 a series has no title reference — without titles_from_data the "
                   "legend reads Series1/Series2 and the chart is unreadable")
    refs = " ".join(r for r in c["series_refs"] if r)
    if "利润表" not in refs:
        out.append(f"S5 no series points at the requested sheet: {c['series_refs']}")
    return out


@check("G1", "every rebuild-path script gets the dropped parts back")
def g1_graft(ctx: dict) -> list[str]:
    """The headline of this slice.

    openpyxl drops all three customXml parts on ANY save. Each of these four
    artifacts went through load→mutate→save, so each is a chance to lose them.
    """
    r = ctx["rebuild"]
    want = r["custom_parts_in"]
    out = []
    if want < 3:
        return [f"G1 the input carries only {want} customXml part(s); the graft has "
                f"nothing to prove"]
    for tag, got in r["custom_parts"].items():
        if got != want:
            out.append(f"G1 {tag}: {got}/{want} customXml parts survived the rebuild")
    if not r["format_report"]["grafted"]:
        out.append("G1 the report claims nothing was grafted, yet openpyxl always "
                   "drops these parts — the repair is not being recorded")
    if r["format_report"]["still_missing"]:
        out.append(f"G1 still missing after the graft: "
                   f"{r['format_report']['still_missing']}")
    return out


@check("G2", "a rebuild that cannot restore what it lost writes nothing")
def g2_graft_fault(ctx: dict) -> list[str]:
    """Fault injection: the graft is disabled and the rebuild must refuse.

    Without this, G1 passing proves the parts survive — not that anything would
    notice if they stopped. A repair path nobody has watched fail is not evidence.
    """
    got = ctx["graft_fault"]["stdout"]
    out = []
    if "RAISED=yes" not in got:
        out.append(f"G2 with the graft disabled the rebuild did not refuse: {got!r}")
    elif "True" not in got.split("RAISED=yes")[1].split("\n")[0]:
        out.append("G2 it refused but did not name the part it lost")
    if "WROTE= False" not in got:
        out.append(f"G2 a file was written despite the refusal: {got!r}")
    return out


@check("K1", "the python engine matches every pinned value")
def k1_pins(ctx: dict) -> list[str]:
    m = ctx["calibration"]["mismatches"]
    return [f"K1 {x['formula']}: pinned {x['pinned']!r}, python {x['python']!r}"
            for x in m[:6]]


@check("K2", "everything outside the boundary is refused, never quietly computed")
def k2_refusals(ctx: dict) -> list[str]:
    """The one behaviour the evaluator must never have.

    A cross-check has value only if it can be wrong loudly. A function that is not
    implemented but returns something plausible is worse than no evaluator at all.
    """
    return [f"K2 {x['formula']} must be refused ({x['why']}) — got {x['got']!r}"
            for x in ctx["calibration"]["leaked"][:6]]


@check("K3", "every function the evaluator claims to support is actually exercised")
def k3_coverage(ctx: dict) -> list[str]:
    missing = ctx["calibration"]["unexercised"]
    if missing:
        return [f"K3 SUPPORTED names never run by the corpus: {', '.join(missing)} — "
                f"a claim nothing measures is a wish list (L1's C5, one level down)"]
    return []


@check("K4", "LibreOffice agrees with every pinned value")
def k4_pin_truth(ctx: dict) -> list[str]:
    """Checks the PINS, not the engine.

    K1 compares the python engine against these numbers; only this compares the
    numbers against the authority they came from. Skipped where LibreOffice is
    absent — and named in the skip list, because that is the exact hole X3 sat in.
    """
    c = ctx["calibration"]
    if not c["soffice_available"]:
        return []
    return [f"K4 {x['formula']}: pinned {x['pinned']!r}, LibreOffice {x['soffice']!r} "
            f"— the PIN is wrong, not the engine" for x in c["soffice_drift"][:6]]


@check("K5", "recalculation writes cached values a value-reader can actually see")
def k5_cached(ctx: dict) -> list[str]:
    r = ctx["recalc"]
    out = []
    if any(v is not None for v in r["cached_before"].values()):
        out.append("K5 the input already had cached values, so writing them proves "
                   "nothing")
    for ref, want in (("B5", 471),):
        got = r["cached_after"].get(ref)
        if not isinstance(got, (int, float)) or abs(float(got) - want) > 1e-6:
            out.append(f"K5 {ref} reads back as {got!r} after recalculation, "
                       f"expected {want}")
    # The formula must survive: a "recalculation" that replaces =B3-B4 with 471 has
    # destroyed the model to produce a number.
    if r["formula_after"].get("B5") != "=B3-B4":
        out.append(f"K5 the formula in B5 became {r['formula_after'].get('B5')!r} — "
                   f"caching a result must not overwrite the formula")
    return out


@check("K6", "a formula only one engine can do is reported, not silently dropped")
def k6_partial(ctx: dict) -> list[str]:
    """VLOOKUP: LibreOffice computes it, the python engine refuses it.

    The report must say so. Without this, "0 disagreements" would be reachable by
    an engine that simply declines everything.
    """
    rep = ctx["recalc"]["mixed"]
    out = []
    unsupported = [f for f in rep["findings"] if f["class"] == "unsupported"]
    if not any("VLOOKUP" in (f.get("formula") or "") for f in unsupported):
        out.append("K6 the VLOOKUP the python engine cannot do was not reported as "
                   "unsupported")
    for f in unsupported:
        if not f.get("formula"):
            out.append(f"K6 {f['cell']} is reported unsupported without its formula "
                       f"text — the caller is told nothing it can act on")
    if rep["cross_checked"] >= rep["formulas"]:
        out.append(f"K6 the report claims {rep['cross_checked']} of {rep['formulas']} "
                   f"formulas were cross-checked, but one engine could not do them all")
    return out


@check("K7", "a single-engine run never presents itself as cross-checked")
def k7_single_engine(ctx: dict) -> list[str]:
    solo, both = ctx["recalc"]["python_only"], ctx["recalc"]["mixed"]
    out = []
    if solo["cross_checked_by_two_engines"]:
        out.append("K7 --engine python reported itself as cross-checked by two engines")
    if solo["cross_checked"]:
        out.append(f"K7 --engine python reports {solo['cross_checked']} cross-checked "
                   f"cells; nothing checked it")
    if not both["cross_checked_by_two_engines"] and ctx["calibration"]["soffice_available"]:
        out.append("K7 the two-engine run does not report itself as cross-checked")
    # The boolean above was the whole story until 2026-08-16, and it was not enough:
    # a model ran --engine python, took the numbers, and told the user they were
    # recalculated without a word about how many engines had seen them (059 §三十五).
    # A skippable flag next to a sentence loses; so the fact gets a sentence.
    note = solo.get("single_engine_note") or ""
    if not note:
        out.append("K7 a single-engine run says so only in a boolean — the one field "
                   "a reader skips is a boolean among numbers")
    elif "python" not in note:
        out.append(f"K7 the single-engine note does not name which engine: {note[:80]!r}")
    if both.get("single_engine_note") and ctx["calibration"]["soffice_available"]:
        out.append("K7 a cross-checked run carries the single-engine caveat too — a "
                   "warning on every run is a warning on none")
    return out


@check("N1", "a CSV round trip keeps the values, and the file carries a BOM")
def n1_csv(ctx: dict) -> list[str]:
    c = ctx["convert"]
    out = []
    if c["csv_head"] != b"\xef\xbb\xbf":
        out.append(f"N1 the CSV has no UTF-8 BOM (starts {c['csv_head']!r}); Excel "
                   f"reads a BOM-less UTF-8 CSV as the local codepage and Chinese "
                   f"arrives as mojibake")
    rows = c["roundtrip"]
    flat = [v for row in rows for v in row]
    for want in ("营业收入", 1240, "毛利"):
        if want not in flat:
            out.append(f"N1 {want!r} did not survive xlsx -> csv -> xlsx")
    if any(isinstance(v, str) and v.strip().lstrip("-").isdigit() for v in flat):
        out.append("N1 a numeric column came back as text after the round trip")
    return out


@check("N2", "the header row names the JSON keys, and rows above it are dropped")
def n2_headers(ctx: dict) -> list[str]:
    c = ctx["convert"]
    out = []
    if not c["jsonl"]:
        return ["N2 the JSONL export is empty"]
    keys = set(c["jsonl"][0])
    if not {"科目", "本季度"} <= keys:
        out.append(f"N2 --header-row 2 did not use row 2 as the keys: {sorted(keys)}")
    # The fixture carries a merged title in row 1. Taking it on faith turns that
    # title into a column name, which is what the default run should show.
    if c["jsonl_row1"] and "科目" in set(c["jsonl_row1"][0]):
        out.append("N2 the default run also produced the row-2 keys, so --header-row "
                   "cannot be shown to do anything")
    return out


@check("N3", "reading a large sheet does not scale memory with row count")
def n3_memory(ctx: dict) -> list[str]:
    """X12's actual claim, measured on THIS script rather than on openpyxl.

    Measuring openpyxl's two modes would prove openpyxl behaves — not that the
    script uses read_only. The eager arm is the same script with read_only forced
    off: the plausible wrong implementation.
    """
    m = ctx["memory"]
    out = []
    small, large = m[MEM_SMALL], m[MEM_LARGE]
    ratio = large["eager"] / large["stream"] if large["stream"] else 0
    if ratio < MEM_MIN_RATIO:
        out.append(f"N3 at {MEM_LARGE:,} rows the script peaked at "
                   f"{large['stream']:.2f} MB against {large['eager']:.2f} MB with "
                   f"read_only forced off — only {ratio:.1f}x, expected at least "
                   f"{MEM_MIN_RATIO}x")
    per_small = small["stream"] / (MEM_SMALL / 1000)
    per_large = large["stream"] / (MEM_LARGE / 1000)
    if per_large >= per_small:
        out.append(f"N3 per-row cost did not fall with size ({per_small:.3f} -> "
                   f"{per_large:.3f} MB per 1k rows) — that is linear growth, which "
                   f"is exactly what streaming is supposed to avoid")
    # The control arm has to be linear, or the comparison above proves nothing.
    eager_small = small["eager"] / (MEM_SMALL / 1000)
    eager_large = large["eager"] / (MEM_LARGE / 1000)
    if eager_large < eager_small * 0.5:
        out.append(f"N3 the eager control is not linear either ({eager_small:.3f} -> "
                   f"{eager_large:.3f} MB per 1k rows); the two arms cannot be told "
                   f"apart and the measurement means nothing")
    return out


@check("N4", "streaming aggregates are numerically right")
def n4_stats(ctx: dict) -> list[str]:
    s = ctx["convert"]["stats"]["stats"][0]
    cols = {c["header"]: c for c in s["columns"]}
    out = []
    if s["rows"] != 3:
        out.append(f"N4 counted {s['rows']} data rows, expected 3")
    q = cols.get("本季度")
    if q is None:
        return out + [f"N4 no 本季度 column in the stats: {sorted(cols)}"]
    # 1240 + 769 + 471 = 2480
    for key, want in (("count", 3), ("sum", 2480.0), ("min", 471), ("max", 1240)):
        if abs(float(q[key]) - float(want)) > 1e-6:
            out.append(f"N4 本季度 {key} = {q[key]!r}, expected {want!r}")
    if abs(q["mean"] - 2480 / 3) > 1e-6:
        out.append(f"N4 本季度 mean = {q['mean']!r}")
    return out


@check("N5", "the render produces a PDF and the page images that were asked for")
def n5_render(ctx: dict) -> list[str]:
    r = ctx["render"]
    out = []
    if r["exit"] != 0 or not r["pdf_exists"]:
        out.append(f"N5 rendering exited {r['exit']} / pdf present {r['pdf_exists']}")
    if not r["report"].get("pages"):
        out.append("N5 the report does not say how many pages were produced")
    if r["images"] != r["report"].get("pages"):
        out.append(f"N5 {r['images']} PNG(s) for {r['report'].get('pages')} page(s)")
    if r["report"].get("blank_pages"):
        out.append(f"N5 page(s) {r['report']['blank_pages']} came out blank")
    # 120 dpi on a portrait A4 is ~991px wide. A renderer ignoring --dpi lands near
    # the 72-dpi default (~595px).
    if r["png_width"] is not None and r["png_width"] < 800:
        out.append(f"N5 the first page image is {r['png_width']}px wide; --dpi "
                   f"{PNG_DPI} was ignored")
    return out


@check("N6", "a render with no ink is refused, and nothing is written")
def n6_blank(ctx: dict) -> list[str]:
    r = ctx["render"]
    out = []
    if r["blank_exit"] != 2:
        out.append(f"N6 rendering an empty sheet exited {r['blank_exit']}; a blank "
                   f"preview looks exactly like lost data")
    if r["blank_wrote"]:
        out.append("N6 a PDF was written despite the refusal")
    return out


@check("N7", "an uncalculated workbook is warned about, without claiming what the "
             "renderer did")
def n7_uncalculated(ctx: dict) -> list[str]:
    """The warning may state what the FILE holds. It may not state what came out on
    paper, because nothing in that script looks at the page.

    It used to say the cells "therefore render EMPTY — the picture is wrong in a way
    the picture cannot show". Measured 2026-08-16 (059 §三十) on two workbooks —
    one with 7 cleared caches, one library-written with none at all — LibreOffice
    computed every one of them and the numbers were in the PDF text layer. The claim
    was not merely unmeasured, it was **inverted**: it fired on a correct picture,
    and said nothing in the one case where the page really was wrong (stale caches,
    which LibreOffice renders verbatim).
    """
    r = ctx["render"]
    if not r["uncalc_count"]:
        return ["N7 the uncalculated fixture reported 0 formula cells without a "
                "cached value, so nothing is proven"]
    warning = r["uncalc_warning"]
    if not warning:
        return ["N7 no warning for a workbook whose formula cells hold no result — "
                "every consumer that reads values sees those cells as empty"]
    out = []
    for claim in ("render EMPTY", "renders EMPTY", "render empty"):
        if claim in warning:
            out.append(f"N7 the warning asserts {claim!r}, which this script never "
                       f"measured and which LibreOffice contradicts — it computes a "
                       f"formula cell that has no cached result")
            break
    if "xlsx_recalc" not in warning:
        out.append("N7 the warning does not name the one command that fixes it")
    return out


@check("N8", "the finance convention is audited by role, applied, and then clean")
def n8_finance(ctx: dict) -> list[str]:
    f = ctx["finance"]
    out = []
    before, after = f["before"], f["after"]
    if not before["cells_in_scope"]:
        return ["N8 no cell fell under the convention, so nothing was checked"]
    if not before["violations"]:
        out.append("N8 the ordinary fixture reported zero violations; it does not "
                   "follow the convention, so this proves the check has no teeth")
    if not before["by_role"].get("input") or not before["by_role"].get("link"):
        out.append(f"N8 violations were not attributed by role: {before['by_role']}")
    if after["violations"]:
        out.append(f"N8 --apply left {after['violations']} violation(s)")
    if f["recheck_exit"] != 0:
        out.append(f"N8 the re-check exited {f['recheck_exit']} after --apply")
    if f["custom_parts"] != 3:
        out.append(f"N8 the recoloured workbook kept {f['custom_parts']}/3 customXml "
                   f"parts — the finance path goes through the graft too")
    return out


@check("N9", "recolouring changes the colour and nothing else about the font")
def n9_font_preserved(ctx: dict) -> list[str]:
    b, a = ctx["finance"]["font_before"], ctx["finance"]["font_after"]
    out = []
    if a["color"] == b["color"]:
        return ["N9 B3's colour did not change, so nothing is proven"]
    for key in ("size", "name"):
        if a[key] != b[key]:
            out.append(f"N9 B3's font {key} changed from {b[key]!r} to {a[key]!r} "
                       f"while only the colour was in scope")
    return out


# ── negative controls ─────────────────────────────────────────────────────────
def flaw_write_via_load_save(ctx, work):
    """The implementation everyone reaches for first, measured on the same edit."""
    w = ctx["write"]
    w["parts_after"] = list(w["naive_parts_after"])
    w["identical"] = list(w["naive_identical"])
    w["report"] = {**w["report"],
                   "parts_lost": sorted(set(w["parts_before"]) - set(w["naive_parts_after"]))}
    return ctx


def flaw_write_rebuilds_every_part(ctx, work):
    ctx["write"]["identical"] = []
    return ctx


def flaw_edit_misses_the_cell(ctx, work):
    ctx["write"]["cells"]["B3"] = 1240
    return ctx


def flaw_edit_resets_style(ctx, work):
    ctx["write"]["styles"] = {**ctx["write"]["styles"], "D5": None}
    return ctx


def flaw_append_overwrites_last_row(ctx, work):
    ctx["write"]["cells"] = {**ctx["write"]["cells"], "A6": None, "B5": "其他业务收入"}
    return ctx


def flaw_shared_formula_overwritten(ctx, work):
    ctx["shared"] = {**ctx["shared"], "master_exit": 0, "master_wrote": True}
    return ctx


def flaw_shared_guard_refuses_whole_file(ctx, work):
    ctx["shared"] = {**ctx["shared"], "other_exit": 2}
    return ctx


def flaw_keeps_stale_calcchain(ctx, work):
    ctx["calcchain"] = {**ctx["calcchain"], "after_formula_edit": True}
    return ctx


def flaw_drops_calcchain_always(ctx, work):
    ctx["calcchain"] = {**ctx["calcchain"], "after_value_edit": False}
    return ctx


def flaw_refuses_damaged_input(ctx, work):
    ctx["damage"] = {**ctx["damage"], "exit": 2, "wrote": False}
    return ctx


def flaw_hides_pre_existing_damage(ctx, work):
    ctx["damage"] = {**ctx["damage"], "pre_existing": []}
    return ctx


def flaw_width_counts_len(ctx, work):
    """`len(s) + 2` — the width every implementation that has not met CJK computes."""
    ctx["autofit"] = {**ctx["autofit"],
                      "widths": {**ctx["autofit"]["widths"],
                                 "A": float(LONG_LABEL_LEN_WIDTH)}}
    return ctx


def flaw_width_counts_formula_text(ctx, work):
    report = copy.deepcopy(ctx["autofit"]["report"])
    for w in report["widths"]:
        if (w["sheet"], w["column"]) == ("汇总", "B"):
            w["width"] = 16
    ctx["autofit"] = {**ctx["autofit"], "report": report}
    return ctx


def flaw_width_counts_merged_title(ctx, work):
    ctx["merge"] = {"width_a": float(MERGE_TITLE_WIDTH)}
    return ctx


def flaw_width_ignores_every_merge(ctx, work):
    """"Skip merged cells" implemented without asking which direction they span."""
    ctx["merge"] = {"width_a": float(MERGE_PLAIN_WIDTH)}
    return ctx


def flaw_cols_appended_after_sheetdata(ctx, work):
    order = [t for t in ctx["autofit"]["sheet_order"] if t != "cols"]
    ctx["autofit"] = {**ctx["autofit"], "sheet_order": order + ["cols"]}
    return ctx


def flaw_read_values_only(ctx, work):
    read = copy.deepcopy(ctx["read"]["json"])
    for c in read["cells"]:
        c["formula"] = None
    ctx["read"] = {**ctx["read"], "json": read}
    return ctx


def flaw_read_uncalculated_as_empty(ctx, work):
    read = copy.deepcopy(ctx["read"]["json"])
    for c in read["cells"]:
        c["uncalculated"] = False
    ctx["read"] = {**ctx["read"], "json": read}
    return ctx


def flaw_read_clamps_bad_range(ctx, work):
    rejects = copy.deepcopy(ctx["read"]["rejects"])
    rejects[1] = {**rejects[1], "exit": 0, "stderr": ""}
    ctx["read"] = {**ctx["read"], "rejects": rejects}
    return ctx


def flaw_read_raises_traceback(ctx, work):
    rejects = copy.deepcopy(ctx["read"]["rejects"])
    rejects[0] = {**rejects[0], "exit": 1,
                  "stderr": "Traceback (most recent call last):\n  ...\n"
                            "ValueError: Value must be of type <class 'int'>"}
    ctx["read"] = {**ctx["read"], "rejects": rejects}
    return ctx


def flaw_inventory_forgets_a_sheet(ctx, work):
    read = copy.deepcopy(ctx["read"]["json"])
    read["sheets"] = [s for s in read["sheets"] if s["name"] != "汇总"]
    ctx["read"] = {**ctx["read"], "json": read}
    return ctx


# The shape that shipped until 2026-08-16: no staleness sweep at all, so a value
# edit left every dependent formula holding the result it had before the edit.
STALE_SWEEP_ANCHOR = """        stale = invalidate_stale_caches(args.src, wb, edited) if edited else \\
            {"cells": [], "truncated": [], "coarse": None}"""
STALE_SWEEP_OFF = """        stale = {"cells": [], "truncated": [], "coarse": None}"""
# The lazy alternative: clear every cached formula in the book instead of the ones
# the edit can actually reach. Honest, and it throws away numbers still true.
STALE_CLOSURE_ANCHOR = """sorted(n for n in seen if n in cached)"""
STALE_CLOSURE_BLANKET = """sorted(cached)"""


def collect_stale(scripts, work: Path, tag: str) -> dict:
    """Re-run the real writer from `scripts` and re-measure the FILES it produced."""
    out_x = work / f"stale-{tag}.xlsx"
    report = work / f"stale-{tag}.json"
    r = run_script_from(scripts, "xlsx_write.py", "--in", work / "calc.xlsx",
                        "--out", out_x, "--sheet", "利润表", "--set", "B3=1310",
                        "--report", report)
    return {"exit": r.returncode,
            "report": json.loads(report.read_text(encoding="utf-8"))
            if report.exists() else {},
            "cached_before": cached_cells(work / "calc.xlsx"),
            "cached_after": cached_cells(out_x) if out_x.exists() else set()}


def flaw_live_stale_caches_left_behind(ctx, work):
    """Exactly what shipped: 毛利 and 营业利润 keep the values they had before the
    edit, and every reader is told they are current."""
    scripts = patched_scripts(work, [(STALE_SWEEP_ANCHOR, STALE_SWEEP_OFF)], "stalekeep")
    ctx["stale"] = collect_stale(scripts, work, "stalekeep")
    return ctx


def flaw_live_stale_sweep_clears_everything(ctx, work):
    """No dependency walk — blank every cached formula in the workbook."""
    scripts = patched_scripts(work, [(STALE_CLOSURE_ANCHOR, STALE_CLOSURE_BLANKET)],
                              "staleall")
    ctx["stale"] = collect_stale(scripts, work, "staleall")
    return ctx


# The shape that shipped until 2026-08-16: a single-engine run said so in a boolean
# and nowhere else.
SINGLE_ENGINE_ANCHOR = """        if len(engines) < 2:"""
SINGLE_ENGINE_OFF = """        if False:"""
# ...and the other way it rots: put the caveat on every run, including the ones that
# really were cross-checked.
SINGLE_ENGINE_ALWAYS = """        if True:"""


def flaw_live_single_engine_note_missing(ctx, work):
    scripts = patched_scripts(work, [(SINGLE_ENGINE_ANCHOR, SINGLE_ENGINE_OFF)],
                              "nosolonote")
    ctx["recalc"] = {**ctx["recalc"], **_reengine(scripts, work, "nosolonote")}
    return ctx


def flaw_live_single_engine_note_on_every_run(ctx, work):
    scripts = patched_scripts(work, [(SINGLE_ENGINE_ANCHOR, SINGLE_ENGINE_ALWAYS)],
                              "alwaysnote")
    ctx["recalc"] = {**ctx["recalc"], **_reengine(scripts, work, "alwaysnote")}
    return ctx


def _reengine(scripts, work: Path, tag: str) -> dict:
    """Re-run both engine modes against the mixed fixture and re-read the reports."""
    out = {}
    for key, extra in (("python_only", ("--engine", "python")), ("mixed", ())):
        report = work / f"eng-{tag}-{key}.json"
        run_script_from(scripts, "xlsx_recalc.py", "--in", work / "mixed.xlsx",
                        "--report", report, *extra)
        out[key] = json.loads(report.read_text(encoding="utf-8")) \
            if report.exists() else {}
    return out


# The shape that shipped until 2026-08-16: the guard refused --out == --in and said
# nothing about replacing anybody else's file.
REPLACED_ANCHOR = """    return out.exists()"""
REPLACED_OFF = """    return False"""
REPLACED_FIELD_ANCHOR = """                  "replaced_existing": replaced,
                  "changes": changes, "widths": fit["changes"],"""
REPLACED_FIELD_OFF = """                  "changes": changes, "widths": fit["changes"],"""


def flaw_live_replacement_never_reported(ctx, work):
    """The field exists and is always False — present, and saying nothing."""
    scripts = patched_scripts(work, [(REPLACED_ANCHOR, REPLACED_OFF)], "norep")
    ctx["replaced"] = collect_replaced(scripts, work, "norep")
    return ctx


def flaw_live_one_writer_drops_the_field(ctx, work):
    scripts = patched_scripts(work, [(REPLACED_FIELD_ANCHOR, REPLACED_FIELD_OFF)],
                              "dropfield")
    ctx["replaced"] = collect_replaced(scripts, work, "dropfield")
    return ctx


# Nothing looked at whether a number survived the column it was printed in.
HASH_ANCHOR = """            out.update(hash_marks(doc, src, sheet))"""
HASH_OFF = """            pass"""
# ...and the plausible half-measure: call every ### a truncation, including the ones
# a sheet really contains.
HASH_LITERAL_ANCHOR = """    if literal:"""
HASH_LITERAL_OFF = """    if False:"""
# ...and a shorter render left the previous one's pages in place.
STALE_IMG_ANCHOR = """            for name in stale:
                (png_dir / name).unlink()"""
STALE_IMG_OFF = """            stale = []"""


def flaw_live_no_hash_check(ctx, work):
    scripts = patched_scripts(work, [(HASH_ANCHOR, HASH_OFF)], "nohash")
    ctx["hash"] = collect_hash_marks(scripts, work, "nohash")
    return ctx


def flaw_live_hash_ignores_literal_content(ctx, work):
    scripts = patched_scripts(work, [(HASH_LITERAL_ANCHOR, HASH_LITERAL_OFF)], "hashlit")
    ctx["hash"] = collect_hash_marks(scripts, work, "hashlit")
    return ctx


def flaw_live_stale_images_kept(ctx, work):
    scripts = patched_scripts(work, [(STALE_IMG_ANCHOR, STALE_IMG_OFF)], "staleimg")
    ctx["stale_images"] = collect_stale_images(scripts, work, "staleimg")
    return ctx


# The shape that shipped until 2026-08-16: only text drove the width, so the one
# cell in 利润表 that actually failed to display — a percentage — was invisible to it.
NUMBER_WIDTH_ANCHOR = """                        shown = displayed_text(cached, cell.number_format)"""
NUMBER_WIDTH_OFF = """                        shown = None"""
# ...and the report listed only what it CHANGED, so "measured, all fine" and "the
# flag did nothing" were the same two bytes.
MEASURED_ANCHOR = """                measured.append(note)"""
MEASURED_OFF = """                pass"""
# ...and nothing looked at whether the table survived the paper it was printed on.
SPLIT_ANCHOR = """            out["columns_off_first_page"] = split_columns(doc, src, sheet)"""
SPLIT_OFF = """            pass"""


def flaw_live_width_ignores_numbers(ctx, work):
    scripts = patched_scripts(work, [(NUMBER_WIDTH_ANCHOR, NUMBER_WIDTH_OFF)], "numblind")
    ctx["percent_width"] = collect_percent_width(scripts, work, "numblind")
    return ctx


def flaw_live_width_reports_only_changes(ctx, work):
    scripts = patched_scripts(work, [(MEASURED_ANCHOR, MEASURED_OFF)], "nomeasure")
    ctx["percent_width"] = collect_percent_width(scripts, work, "nomeasure")
    return ctx


def flaw_live_split_not_detected(ctx, work):
    scripts = patched_scripts(work, [(SPLIT_ANCHOR, SPLIT_OFF)], "nosplit")
    ctx["split"] = collect_split(scripts, work, "nosplit")
    return ctx


# The shape that shipped until 2026-08-16: the report drew no boundary at all, and
# `findings: []` was the whole answer.
AUDIT_SCOPE_ANCHOR = """            "scope": "references, not values: every formula was resolved to the \""""
AUDIT_SCOPE_OFF = """            "unused_scope": "references, not values: every formula was resolved to the \""""
AUDIT_COUNT_ANCHOR = """"counts": {**counts, **stats, "formulas_evaluated": 0},"""
AUDIT_COUNT_OFF = """"counts": {**counts, **stats},"""
# The plausible half-measure: say it only when there is something to say. That
# removes the sentence from the one report most likely to be read as "all clear".
AUDIT_SCOPE_ONLY_WHEN_DIRTY = """            "scope": None if not findings else "references, not values: every formula was resolved to the \""""


def flaw_live_audit_states_no_scope(ctx, work):
    scripts = patched_scripts(work, [(AUDIT_SCOPE_ANCHOR, AUDIT_SCOPE_OFF),
                                     (AUDIT_COUNT_ANCHOR, AUDIT_COUNT_OFF)], "noscope")
    ctx["audit"] = {**ctx["audit"],
                    "spotless": collect_audit_spotless(scripts, work, "noscope",
                                                       work / "calc.xlsx"),
                    "broken": _reaudit(scripts, work, "noscope")}
    return ctx


def flaw_live_audit_scope_only_when_dirty(ctx, work):
    scripts = patched_scripts(work, [(AUDIT_SCOPE_ANCHOR,
                                      AUDIT_SCOPE_ONLY_WHEN_DIRTY)], "dirtyonly")
    ctx["audit"] = {**ctx["audit"],
                    "spotless": collect_audit_spotless(scripts, work, "dirtyonly",
                                                       work / "calc.xlsx"),
                    "broken": _reaudit(scripts, work, "dirtyonly")}
    return ctx


def _reaudit(scripts, work: Path, tag: str) -> dict:
    out = work / f"audit-broken-{tag}.json"
    run_script_from(scripts, "xlsx_audit.py", "--in", work / "broken.xlsx",
                    "--out", out, "--fail-on", "error,missing,circular")
    return json.loads(out.read_text(encoding="utf-8")) if out.exists() else {}


# Removing the bound puts the precise walk back on a graph it cannot afford: the
# 12M-reference fixture takes 6.9s here instead of 0.26s, and the report stops
# saying it fell back at all.
SPAN_BOUND_ANCHOR = """MAX_DEPENDENCY_SPAN = 500_000"""
SPAN_BOUND_OFF = """MAX_DEPENDENCY_SPAN = 10_000_000_000"""
# Falling back to coarse and then not clearing anything — fast, reports the
# fallback, and leaves every stale number exactly where it was.
COARSE_KEEPS_ANCHOR = """        stale, coarse = sorted(cached), ("""
COARSE_KEEPS_BROKEN = """        stale, coarse = [], ("""


def flaw_live_dependency_walk_unbounded(ctx, work):
    scripts = patched_scripts(work, [(SPAN_BOUND_ANCHOR, SPAN_BOUND_OFF)], "nobound")
    ctx["stale_bound"] = collect_stale_bound(scripts, work, "nobound")
    return ctx


def flaw_live_coarse_fallback_clears_nothing(ctx, work):
    scripts = patched_scripts(work, [(COARSE_KEEPS_ANCHOR, COARSE_KEEPS_BROKEN)],
                              "coarsenoop")
    ctx["stale_bound"] = collect_stale_bound(scripts, work, "coarsenoop")
    return ctx


# The sentence that shipped until 2026-08-16 — a claim about the rendered page that
# the script never looked at, and that LibreOffice contradicts.
UNCALC_WARNING_ANCHOR = """            report["warning"] = (
                f"{blank_formulas} formula cell(s) carry no cached result in the "
                f"file. Anything that reads values instead of rendering — this "
                f"skill's own reader included — sees them as empty; whether they "
                f"appear on the page depends on the renderer computing them, which "
                f"this script does not check. Run xlsx_recalc.py to put the numbers "
                f"into the file")"""
UNCALC_WARNING_SHIPPED = """            report["warning"] = (
                f"{blank_formulas} formula cell(s) have no cached value and therefore "
                f"render EMPTY — the picture is wrong in a way the picture cannot "
                f"show. Run xlsx_recalc.py first")"""


def flaw_live_uncalc_warning_claims_the_page(ctx, work):
    scripts = patched_scripts(work, [(UNCALC_WARNING_ANCHOR, UNCALC_WARNING_SHIPPED)],
                              "uncalcclaim")
    report = work / "raw-claim.json"
    run_script_from(scripts, "xlsx_pdf.py", "--in", BOOK,
                    "--out", work / "raw-claim.pdf", "--report", report)
    raw = json.loads(report.read_text(encoding="utf-8")) if report.exists() else {}
    ctx["render"] = {**ctx["render"], "uncalc_warning": raw.get("warning"),
                     "uncalc_count": raw.get("uncalculated_formulas")}
    return ctx


# The wording that shipped until 2026-08-16: a bare path, no size. Backing the fix
# out means re-running the real reader with it, not editing the string this file
# collected — the defect is what the entry point PRINTS.
POINTER_ANCHOR = """    return (f"; the full list is in {out} ({size} bytes) — read the entries you need "
            f"out of it, or re-run with a narrower --range/--cells. Printing a file "
            f"that size back is the context blowout this trim exists to prevent")"""
POINTER_SHIPPED = """    return f"; the full list is in {out}\""""
POINTER_WRONG_SIZE = """    return (f"; the full list is in {out} ({size // 10} bytes) — read the entries "
            f"you need out of it")"""


def collect_scale_note(scripts, work: Path, tag: str) -> dict:
    """Re-run the real reader from `scripts` and re-measure what it printed."""
    out_json = work / f"big-{tag}.json"
    r = run_script_from(scripts, "xlsx_read.py", "--in", work / "big.xlsx",
                        "--sheet", "明细", "--range", f"A1:C{SCALE_ROWS}",
                        "--out", out_json)
    return {"stdout_bytes": len(r.stdout.encode("utf-8")),
            "file_cells": len(json.loads(out_json.read_text(encoding="utf-8"))
                              .get("cells", [])),
            "note": json.loads(r.stdout).get("cells_note", ""),
            "file_bytes": out_json.stat().st_size}


def flaw_live_pointer_without_a_size(ctx, work):
    """Exactly what shipped: `the full list is in <path>` and nothing about its size."""
    scripts = patched_scripts(work, [(POINTER_ANCHOR, POINTER_SHIPPED)], "ptrbare")
    ctx["scale"] = {**ctx["scale"], **collect_scale_note(scripts, work, "ptrbare")}
    return ctx


def flaw_live_pointer_states_a_wrong_size(ctx, work):
    """A size that was not measured. Stating one is worse than stating none: the
    caller now has a number to plan against and it is off by an order of magnitude."""
    scripts = patched_scripts(work, [(POINTER_ANCHOR, POINTER_WRONG_SIZE)], "ptrwrong")
    ctx["scale"] = {**ctx["scale"], **collect_scale_note(scripts, work, "ptrwrong")}
    return ctx


def flaw_stdout_dumps_every_cell(ctx, work):
    ctx["scale"] = {**ctx["scale"], "stdout_bytes": 480_000}
    return ctx


def flaw_trimming_loses_data(ctx, work):
    """Trimming stdout by dropping the rows instead of routing them to --out."""
    ctx["scale"] = {**ctx["scale"], "file_cells": 20}
    return ctx


def flaw_in_place_allowed(ctx, work):
    scale = copy.deepcopy(ctx["scale"])
    scale["in_place"]["xlsx_write.py"] = {"exit": 0, "stderr": ""}
    ctx["scale"] = scale
    return ctx


def flaw_fixture_stops_exercising_autofit(ctx, work):
    """A fixture whose columns are nearly all wide enough already.

    V0's own control has to be something no other check owns, or it cannot be told
    apart from them. Dropping a sheet — the obvious choice — is what
    `inventory-forgets-a-sheet` already does, so it would prove nothing about V0.
    The 汇总!B entry C2 reads is kept, so only V0 notices.
    """
    report = copy.deepcopy(ctx["autofit"]["report"])
    report["widths"] = [w for w in report["widths"]
                        if (w["sheet"], w["column"]) == ("汇总", "B")]
    ctx["autofit"] = {**ctx["autofit"], "report": report}
    return ctx


def _drop_findings(ctx, keep):
    """Rebuild the broken-workbook audit keeping only `keep(f)` findings."""
    a = copy.deepcopy(ctx["audit"])
    a["broken"]["findings"] = [f for f in a["broken"]["findings"] if keep(f)]
    a["broken"]["by_class"] = {
        c: sum(1 for f in a["broken"]["findings"] if f["class"] == c)
        for c in a["broken"]["by_class"]}
    a["cells"] = {f["cell"] for f in a["broken"]["findings"] if f["class"] != "uncalc"}
    ctx["audit"] = a
    return ctx


def flaw_audit_scans_formulas_only(ctx, work):
    """Reporting the token but never saying which reference produced it."""
    a = copy.deepcopy(ctx["audit"])
    for f in a["broken"]["findings"]:
        if f["class"] == "error":
            f["cause"] = None
    ctx["audit"] = a
    return ctx


def flaw_audit_misses_missing_sheet(ctx, work):
    """The plausible implementation: only scan for tokens already present."""
    return _drop_findings(ctx, lambda f: f["class"] != "missing")


def flaw_audit_no_cycle_detection(ctx, work):
    return _drop_findings(ctx, lambda f: f["class"] != "circular")


def flaw_audit_cycle_without_chain(ctx, work):
    a = copy.deepcopy(ctx["audit"])
    for f in a["broken"]["findings"]:
        if f["class"] == "circular":
            f["chain"] = ""
    ctx["audit"] = a
    return ctx


def flaw_audit_misses_self_reference(ctx, work):
    """=F1 is the cycle a graph walker that never revisits its start will skip."""
    return _drop_findings(ctx, lambda f: f["cell"] != "利润表!F1")


def flaw_audit_treats_uncalc_as_clean(ctx, work):
    a = copy.deepcopy(ctx["audit"])
    a["clean"]["by_class"]["uncalc"] = 0
    ctx["audit"] = a
    return ctx


def flaw_audit_flags_clean_cross_sheet(ctx, work):
    a = copy.deepcopy(ctx["audit"])
    a["clean"]["by_class"]["missing"] = 2
    ctx["audit"] = a
    return ctx


def flaw_parser_reads_string_literals(ctx, work):
    a = copy.deepcopy(ctx["audit"])
    a["cells"] = set(a["cells"]) | {"利润表!G1"}
    ctx["audit"] = a
    return ctx


def flaw_parser_reads_function_names(ctx, work):
    a = copy.deepcopy(ctx["audit"])
    a["cells"] = set(a["cells"]) | {"利润表!H1"}
    ctx["audit"] = a
    return ctx


def flaw_write_stores_missing_sheet_ref(ctx, work):
    r = copy.deepcopy(ctx["refuse"])
    r["cases"]["missing sheet"] = {"exit": 0, "stderr": "", "wrote": True}
    ctx["refuse"] = r
    return ctx


def flaw_write_stores_error_token(ctx, work):
    r = copy.deepcopy(ctx["refuse"])
    r["cases"]["error token baked in"] = {"exit": 0, "stderr": "", "wrote": True}
    ctx["refuse"] = r
    return ctx


def flaw_write_refuses_every_cross_sheet(ctx, work):
    """A guard that refuses all cross-sheet links passes Q1/Q2 and is useless."""
    r = copy.deepcopy(ctx["refuse"])
    r["legit_exit"] = 2
    ctx["refuse"] = r
    return ctx


def _style(ctx, ref, **kw):
    r = copy.deepcopy(ctx["rebuild"])
    r["style"][ref].update(kw)
    ctx["rebuild"] = r
    return ctx


def flaw_format_applies_to_whole_sheet(ctx, work):
    return _style(ctx, "D3", fill="FFFFF2CC")


def flaw_format_misses_the_range(ctx, work):
    return _style(ctx, "C5", number_format="General", bold=False, color=None,
                  fill=None, border=None)


def flaw_font_replaced_wholesale(ctx, work):
    """`cell.font = Font(color=...)` — the one-liner that resets size and face."""
    return _style(ctx, "B3", size=11.0, name="Calibri")


def flaw_conditional_replaces_existing(ctx, work):
    r = copy.deepcopy(ctx["rebuild"])
    r["cf"]["D3:D5"] = ["cellIs"]
    ctx["rebuild"] = r
    return ctx


def flaw_conditional_drops_a_kind(ctx, work):
    r = copy.deepcopy(ctx["rebuild"])
    r["cf"].pop("C3:C5", None)
    ctx["rebuild"] = r
    return ctx


def flaw_panes_left_at_the_fixture_value(ctx, work):
    r = copy.deepcopy(ctx["rebuild"])
    r["panes"] = dict(r["panes_before"])
    ctx["rebuild"] = r
    return ctx


def flaw_chart_replaces_the_existing_one(ctx, work):
    r = copy.deepcopy(ctx["rebuild"])
    r["chart"]["count"] = r["chart_before"]["count"]
    ctx["rebuild"] = r
    return ctx


def flaw_chart_series_unnamed(ctx, work):
    """titles_from_data left off: the legend reads Series1, Series2."""
    r = copy.deepcopy(ctx["rebuild"])
    r["chart"]["series_titles"] = [None] * len(r["chart"]["series_titles"])
    ctx["rebuild"] = r
    return ctx


def flaw_rebuild_without_graft(ctx, work):
    """What `load → mutate → save` does on its own: the customXml parts are gone."""
    r = copy.deepcopy(ctx["rebuild"])
    r["custom_parts"] = {k: 0 for k in r["custom_parts"]}
    r["format_report"] = {**r["format_report"], "grafted": []}
    ctx["rebuild"] = r
    return ctx


def flaw_graft_reports_a_repair_it_did_not_make(ctx, work):
    r = copy.deepcopy(ctx["rebuild"])
    r["format_report"] = {**r["format_report"],
                          "still_missing": ["customXml/item1.xml"]}
    ctx["rebuild"] = r
    return ctx


def flaw_rebuild_writes_anyway(ctx, work):
    ctx["graft_fault"] = {"stdout": "RAISED=no\nWROTE= True", "exit": 0}
    return ctx


def _cal(ctx, **kw):
    c = copy.deepcopy(ctx["calibration"])
    c.update(kw)
    ctx["calibration"] = c
    return ctx


def flaw_evaluator_gets_a_number_wrong(ctx, work):
    """The first draft's answer for -2^2: right in most languages, wrong in Excel."""
    return _cal(ctx, mismatches=[{"formula": "=-2^2", "pinned": 4, "python": -4.0}])


def flaw_evaluator_guesses_instead_of_refusing(ctx, work):
    return _cal(ctx, leaked=[{"formula": "=VLOOKUP(H1,H1:I3,2,0)",
                              "why": "not implemented", "got": ("VALUE", 20.0)}])


def flaw_supported_is_a_wish_list(ctx, work):
    return _cal(ctx, unexercised=["MEDIAN", "XLOOKUP"])


def flaw_pinned_value_is_wrong(ctx, work):
    """X3's defect, one level down: the expectation itself is wrong."""
    return _cal(ctx, soffice_available=True,
                soffice_drift=[{"formula": "=MOD(-7,3)", "pinned": -1, "soffice": 2}])


def flaw_recalc_leaves_no_cached_values(ctx, work):
    r = copy.deepcopy(ctx["recalc"])
    r["cached_after"] = {"B5": None, "D5": None}
    ctx["recalc"] = r
    return ctx


def flaw_recalc_overwrites_the_formula(ctx, work):
    r = copy.deepcopy(ctx["recalc"])
    r["formula_after"]["B5"] = 471
    ctx["recalc"] = r
    return ctx


def flaw_unsupported_reported_without_the_formula(ctx, work):
    r = copy.deepcopy(ctx["recalc"])
    for f in r["mixed"]["findings"]:
        if f["class"] == "unsupported":
            f["formula"] = None
    ctx["recalc"] = r
    return ctx


def flaw_unsupported_silently_dropped(ctx, work):
    r = copy.deepcopy(ctx["recalc"])
    r["mixed"]["findings"] = [f for f in r["mixed"]["findings"]
                              if f["class"] != "unsupported"]
    ctx["recalc"] = r
    return ctx


def flaw_single_engine_claims_cross_check(ctx, work):
    r = copy.deepcopy(ctx["recalc"])
    r["python_only"]["cross_checked_by_two_engines"] = True
    ctx["recalc"] = r
    return ctx


def flaw_csv_without_bom(ctx, work):
    c = copy.deepcopy(ctx["convert"])
    c["csv_head"] = b"\xe7\xa7\x91\xe7"
    ctx["convert"] = c
    return ctx


def flaw_csv_roundtrip_stringifies_numbers(ctx, work):
    c = copy.deepcopy(ctx["convert"])
    c["roundtrip"] = [[str(v) if isinstance(v, (int, float)) else v for v in row]
                      for row in c["roundtrip"]]
    ctx["convert"] = c
    return ctx


def flaw_header_row_ignored(ctx, work):
    """--header-row accepted and then not used: row 1 becomes the keys anyway."""
    c = copy.deepcopy(ctx["convert"])
    c["jsonl"] = copy.deepcopy(c["jsonl_row1"])
    ctx["convert"] = c
    return ctx


def flaw_reader_forgets_read_only(ctx, work):
    """The one-character omission this whole capability is about."""
    m = copy.deepcopy(ctx["memory"])
    for rows in m:
        m[rows]["stream"] = m[rows]["eager"]
    ctx["memory"] = m
    return ctx


def flaw_memory_grows_linearly(ctx, work):
    m = copy.deepcopy(ctx["memory"])
    per = m[MEM_SMALL]["stream"] / (MEM_SMALL / 1000)
    m[MEM_LARGE]["stream"] = per * (MEM_LARGE / 1000) * 1.05
    ctx["memory"] = m
    return ctx


def flaw_stats_miscounts(ctx, work):
    c = copy.deepcopy(ctx["convert"])
    for col in c["stats"]["stats"][0]["columns"]:
        if col["header"] == "本季度":
            col["sum"] = 9999.0
    ctx["convert"] = c
    return ctx


def flaw_render_ignores_dpi(ctx, work):
    r = copy.deepcopy(ctx["render"])
    r["png_width"] = 595
    ctx["render"] = r
    return ctx


def flaw_render_drops_images(ctx, work):
    r = copy.deepcopy(ctx["render"])
    r["images"] = 1
    ctx["render"] = r
    return ctx


def flaw_blank_render_accepted(ctx, work):
    r = copy.deepcopy(ctx["render"])
    r.update(blank_exit=0, blank_wrote=True)
    ctx["render"] = r
    return ctx


def flaw_uncalculated_not_warned(ctx, work):
    r = copy.deepcopy(ctx["render"])
    r["uncalc_warning"] = None
    ctx["render"] = r
    return ctx


def flaw_finance_check_has_no_teeth(ctx, work):
    f = copy.deepcopy(ctx["finance"])
    f["before"]["violations"] = 0
    f["before"]["by_role"] = {"input": 0, "formula": 0, "link": 0}
    ctx["finance"] = f
    return ctx


def flaw_finance_apply_leaves_violations(ctx, work):
    f = copy.deepcopy(ctx["finance"])
    f["after"]["violations"] = 4
    ctx["finance"] = f
    return ctx


def flaw_finance_resets_the_font(ctx, work):
    f = copy.deepcopy(ctx["finance"])
    f["font_after"] = {**f["font_after"], "size": 11.0, "name": "Calibri"}
    ctx["finance"] = f
    return ctx


def flaw_live_import_widths_after_append(ctx, work):
    """Widths applied after the first append — write_only drops them. The self-check
    added with the fix catches it, so this arm proves that check is load-bearing."""
    scripts = patched_scripts(work, [(IMPORT_ORDER_ANCHOR, IMPORT_ORDER_BROKEN)],
                              "impafter")
    ctx["import_autofit"] = collect_import_autofit(scripts, work, "impafter",
                                                   ctx["_csv_src"])
    return ctx


def flaw_live_import_reports_widths_it_did_not_write(ctx, work):
    """The same order bug WITH the self-check removed — exactly what shipped: exit 0,
    `widths_set: 5`, and not one <col> in the file."""
    scripts = patched_scripts(work, [(IMPORT_ORDER_ANCHOR, IMPORT_ORDER_BROKEN),
                                     (SELFCHECK_ANCHOR, SELFCHECK_OFF)], "impsilent")
    ctx["import_autofit"] = collect_import_autofit(scripts, work, "impsilent",
                                                   ctx["_csv_src"])
    return ctx


FLAWS = [
    ("csv-written-without-a-bom", flaw_csv_without_bom, {"N1"}, ""),
    ("LIVE: the pointer names the file but not its size (as it shipped until "
     "2026-08-16)", flaw_live_pointer_without_a_size, {"O3"}, ""),
    ("LIVE: the pointer states a size nobody measured",
     flaw_live_pointer_states_a_wrong_size, {"O3"}, ""),
    ("LIVE: an edit leaves the cached results it invalidated (as it shipped until "
     "2026-08-16)", flaw_live_stale_caches_left_behind, {"W8"}, ""),
    ("LIVE: the staleness sweep blanks every formula instead of the dependents",
     flaw_live_stale_sweep_clears_everything, {"W8"}, ""),
    ("LIVE: the uncalculated warning asserts what the page shows",
     flaw_live_uncalc_warning_claims_the_page, {"N7"}, ""),
    ("LIVE: a single-engine run says so only in a boolean (as it shipped until "
     "2026-08-16)", flaw_live_single_engine_note_missing, {"K7"}, ""),
    ("LIVE: the single-engine caveat is attached to every run",
     flaw_live_single_engine_note_on_every_run, {"K7"}, ""),
    ("LIVE: replacing an existing file is never reported (as it shipped until "
     "2026-08-16)", flaw_live_replacement_never_reported, {"O4"}, ""),
    ("LIVE: one writer drops the replaced_existing field",
     flaw_live_one_writer_drops_the_field, {"O4"}, ""),
    ("LIVE: nothing notices a number rendered as ###", flaw_live_no_hash_check,
     {"N11"}, ""),
    ("LIVE: every ### is called a truncation, including real content",
     flaw_live_hash_ignores_literal_content, {"N11"}, ""),
    ("LIVE: a shorter render leaves the previous one's pages in place",
     flaw_live_stale_images_kept, {"N12"}, ""),
    ("LIVE: column width ignores numbers, so ### stays (as it shipped until "
     "2026-08-16)", flaw_live_width_ignores_numbers, {"C5"}, ""),
    ("LIVE: the width report lists only what it changed",
     flaw_live_width_reports_only_changes, {"C6"}, ""),
    ("LIVE: nothing notices the table was split across pages",
     flaw_live_split_not_detected, {"N10"}, ""),
    ("LIVE: the audit report draws no boundary at all (as it shipped until "
     "2026-08-16)", flaw_live_audit_states_no_scope, {"A7"}, ""),
    ("LIVE: the audit states its scope only when it found something",
     flaw_live_audit_scope_only_when_dirty, {"A7"}, ""),
    ("LIVE: the dependency walk runs unbounded on a graph it cannot expand",
     flaw_live_dependency_walk_unbounded, {"W9"}, ""),
    ("LIVE: the coarse fallback reports itself and then clears nothing",
     flaw_live_coarse_fallback_clears_nothing, {"W9"}, ""),
    ("LIVE: import applies widths after the first append (write_only drops them)",
     flaw_live_import_widths_after_append, {"C4"}, ""),
    ("LIVE: import reports widths it did not write (as it shipped until 2026-08-05)",
     flaw_live_import_reports_widths_it_did_not_write, {"C4"}, ""),
    ("csv-roundtrip-turns-numbers-into-text", flaw_csv_roundtrip_stringifies_numbers,
     {"N1"}, ""),
    ("header-row-accepted-then-ignored", flaw_header_row_ignored, {"N2"}, ""),
    ("reader-forgets-read-only", flaw_reader_forgets_read_only, {"N3"}, ""),
    ("memory-grows-linearly-with-rows", flaw_memory_grows_linearly, {"N3"}, ""),
    ("streaming-stats-miscount", flaw_stats_miscounts, {"N4"}, ""),
    ("render-ignores-dpi", flaw_render_ignores_dpi, {"N5"}, ""),
    ("render-drops-a-page-image", flaw_render_drops_images, {"N5"}, ""),
    ("blank-render-handed-back-as-a-preview", flaw_blank_render_accepted, {"N6"}, ""),
    ("uncalculated-workbook-rendered-without-a-warning",
     flaw_uncalculated_not_warned, {"N7"}, ""),
    ("finance-check-reports-nothing-on-an-ordinary-sheet",
     flaw_finance_check_has_no_teeth, {"N8"}, ""),
    ("finance-apply-leaves-violations", flaw_finance_apply_leaves_violations,
     {"N8"}, ""),
    ("finance-apply-resets-the-font", flaw_finance_resets_the_font, {"N9"}, ""),

    ("evaluator-gets-a-number-wrong", flaw_evaluator_gets_a_number_wrong, {"K1"}, ""),
    ("evaluator-guesses-instead-of-refusing", flaw_evaluator_guesses_instead_of_refusing,
     {"K2"}, ""),
    ("supported-list-is-a-wish-list", flaw_supported_is_a_wish_list, {"K3"}, ""),
    ("the-pinned-value-itself-is-wrong", flaw_pinned_value_is_wrong, {"K4"}, ""),
    ("recalc-leaves-no-cached-values", flaw_recalc_leaves_no_cached_values, {"K5"}, ""),
    ("recalc-overwrites-the-formula-with-its-result",
     flaw_recalc_overwrites_the_formula, {"K5"}, ""),
    ("unsupported-reported-without-its-formula",
     flaw_unsupported_reported_without_the_formula, {"K6"}, ""),
    ("unsupported-formulas-silently-dropped", flaw_unsupported_silently_dropped,
     {"K6"}, ""),
    ("single-engine-run-claims-it-was-cross-checked",
     flaw_single_engine_claims_cross_check, {"K7"}, ""),

    ("format-applies-to-the-whole-sheet", flaw_format_applies_to_whole_sheet,
     {"S1"}, ""),
    ("format-never-reaches-the-range", flaw_format_misses_the_range, {"S1"}, ""),
    ("font-assigned-wholesale-resets-size-and-face", flaw_font_replaced_wholesale,
     {"S2"}, ""),
    ("conditional-rule-replaces-the-existing-one", flaw_conditional_replaces_existing,
     {"S3"}, ""),
    ("conditional-drops-a-rule-kind", flaw_conditional_drops_a_kind, {"S3"}, ""),
    ("panes-left-at-the-fixture-value", flaw_panes_left_at_the_fixture_value,
     {"S4"}, ""),
    ("chart-replaces-the-existing-one", flaw_chart_replaces_the_existing_one,
     {"S5"}, ""),
    ("chart-series-come-out-unnamed", flaw_chart_series_unnamed, {"S5"}, ""),
    ("rebuild-without-the-graft", flaw_rebuild_without_graft, {"G1"}, ""),
    ("graft-claims-a-repair-it-did-not-make", flaw_graft_reports_a_repair_it_did_not_make,
     {"G1"}, ""),
    ("rebuild-writes-a-lossy-file-anyway", flaw_rebuild_writes_anyway, {"G2"}, ""),

    ("audit-names-the-token-but-not-the-cause", flaw_audit_scans_formulas_only,
     {"A1"}, ""),
    ("audit-only-scans-for-existing-tokens", flaw_audit_misses_missing_sheet,
     {"A2"}, ""),
    ("audit-does-not-look-for-cycles", flaw_audit_no_cycle_detection, {"A3"}, ""),
    ("audit-reports-a-cycle-without-the-chain", flaw_audit_cycle_without_chain,
     {"A3"}, ""),
    ("audit-misses-a-one-cell-self-reference", flaw_audit_misses_self_reference,
     {"A3"}, ""),
    ("audit-counts-uncalculated-formulas-as-clean", flaw_audit_treats_uncalc_as_clean,
     {"A4"}, ""),
    ("audit-flags-a-legitimate-cross-sheet-link", flaw_audit_flags_clean_cross_sheet,
     {"A5"}, ""),
    ("parser-reads-references-inside-strings", flaw_parser_reads_string_literals,
     {"A6"}, ""),
    ("parser-reads-LOG10-as-a-cell", flaw_parser_reads_function_names, {"A6"}, ""),
    ("write-stores-a-reference-to-a-missing-sheet", flaw_write_stores_missing_sheet_ref,
     {"Q1"}, ""),
    ("write-stores-an-error-token-as-a-formula", flaw_write_stores_error_token,
     {"Q2"}, ""),
    ("write-refuses-every-cross-sheet-link", flaw_write_refuses_every_cross_sheet,
     {"Q3"}, ""),

    ("write-via-openpyxl-load-save", flaw_write_via_load_save, {"W1", "W2"},
     "W2 also fires, and it must: load→save both loses parts and rebuilds the ones "
     "it keeps. W1 owns the loss, W2 owns the rebuild, and this one flaw is where "
     "both come from"),
    ("write-rebuilds-every-part", flaw_write_rebuilds_every_part, {"W2"}, ""),
    ("edit-does-not-reach-the-cell", flaw_edit_misses_the_cell, {"W3"}, ""),
    ("edit-resets-the-cell-format", flaw_edit_resets_style, {"W3"}, ""),
    ("append-overwrites-the-last-row", flaw_append_overwrites_last_row, {"W4"}, ""),
    ("shared-formula-master-overwritten", flaw_shared_formula_overwritten, {"W5"}, ""),
    ("shared-formula-guard-refuses-the-whole-file",
     flaw_shared_guard_refuses_whole_file, {"W5"}, ""),
    ("keeps-a-stale-calcchain", flaw_keeps_stale_calcchain, {"W6"}, ""),
    ("drops-the-calcchain-on-every-edit", flaw_drops_calcchain_always, {"W6"}, ""),
    ("refuses-an-already-damaged-input", flaw_refuses_damaged_input, {"W7"}, ""),
    ("hides-pre-existing-damage", flaw_hides_pre_existing_damage, {"W7"}, ""),

    ("width-counts-characters-not-display-units", flaw_width_counts_len, {"C1"}, ""),
    ("width-measures-the-formula-text", flaw_width_counts_formula_text, {"C2"}, ""),
    ("width-counts-a-title-merged-across-columns", flaw_width_counts_merged_title,
     {"C3"}, ""),
    ("width-ignores-merges-in-both-directions", flaw_width_ignores_every_merge,
     {"C3"}, ""),
    ("cols-appended-after-sheetdata", flaw_cols_appended_after_sheetdata, {"E1"}, ""),

    ("read-returns-values-only", flaw_read_values_only, {"R1", "V0"},
     "V0 also fires, honestly: a reader that reports no formulas at all leaves the "
     "window with zero formula cells, and R1/R2 then have no subjects. V0 is the "
     "check that says so"),
    ("read-reports-uncalculated-as-empty", flaw_read_uncalculated_as_empty, {"R2"}, ""),
    ("read-clamps-an-oversized-range", flaw_read_clamps_bad_range, {"R3"}, ""),
    ("read-answers-with-a-traceback", flaw_read_raises_traceback, {"R3"}, ""),
    ("inventory-forgets-a-sheet", flaw_inventory_forgets_a_sheet, {"R4", "V0"},
     "V0 also fires: with only one sheet left the fixture no longer carries the "
     "cross-sheet reference the later capabilities rest on. Both findings are true "
     "and they say different things"),

    ("stdout-dumps-every-cell", flaw_stdout_dumps_every_cell, {"O1"}, ""),
    ("trimming-drops-the-data-instead-of-filing-it", flaw_trimming_loses_data,
     {"O1"}, ""),
    ("writer-overwrites-its-own-input", flaw_in_place_allowed, {"O2"}, ""),

    ("CONTROL: fixture stops exercising autofit",
     flaw_fixture_stops_exercising_autofit, {"V0"}, ""),
]


def fired(ctx: dict) -> dict[str, list[str]]:
    out = {}
    for cid, c in CHECKS.items():
        findings = c["fn"](ctx)
        if findings:
            out[cid] = findings
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()

    for f in (BOOK, NARROW):
        if not f.is_file():
            print(f"[error] fixture missing: {f} (run fixtures/make_fixtures.py)",
                  file=sys.stderr)
            return 1

    results = []
    SKIPS.clear()
    with tempfile.TemporaryDirectory(prefix="xlsx-skill-test-") as td:
        work = Path(td)
        base = collect(work)

        clean = fired(base)
        results.append({"case": "real output of the real scripts", "expect": "silence",
                        "ok": not clean,
                        "detail": [f for v in clean.values() for f in v]})

        matrix = []
        for name, mutate, expected, cascade in FLAWS:
            ctx = mutate(copy.deepcopy(base), work)
            got = fired(ctx)
            unexpected = set(got) - expected
            missing = expected - set(got)
            matrix.append({"flaw": name, "expected": sorted(expected),
                           "fired": sorted(got), "cascade_note": cascade})
            detail = []
            if missing:
                detail.append(f"expected {sorted(missing)} to fire and it did not")
            if unexpected and not cascade:
                detail.append(f"unexpected checks fired: {sorted(unexpected)} — either "
                              f"a real cascade that needs documenting, or a check "
                              f"measuring the wrong thing")
            results.append({"case": f"flaw: {name}",
                            "expect": f"fires {sorted(expected)}",
                            "ok": not detail, "detail": detail,
                            "fired": sorted(got)})

    failed = [r for r in results if not r["ok"]]
    if args.json:
        print(json.dumps({"results": results, "matrix": matrix,
                          "skipped": SKIPS, "failed": len(failed)},
                         ensure_ascii=False, indent=2))
        return 1 if failed else 0

    for r in results:
        print(f"{'PASS' if r['ok'] else 'FAIL'}  [{r['expect']}] {r['case']}")
        for d in r["detail"][:6]:
            print(f"      · {d}")
    print("\n[flaw -> fired] every row must light the check that owns the defect, "
          "and nothing else:")
    for m in matrix:
        extra = sorted(set(m["fired"]) - set(m["expected"]))
        note = f"   (also {', '.join(extra)}: {m['cascade_note']})" if extra else ""
        print(f"  {m['flaw']:<46} -> {', '.join(m['fired']) or '(nothing)'}{note}")
    if SKIPS:
        print("\n[skipped] claims this host could not exercise:")
        for note in SKIPS:
            print(f"  - {note}")
    print(f"\n[xlsx-skill] {len(results) - len(failed)} passed, {len(failed)} failed, "
          f"{len(CHECKS)} assertions"
          + (f", {len(SKIPS)} skipped" if SKIPS else ""))
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
