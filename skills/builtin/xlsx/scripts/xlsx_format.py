#!/usr/bin/env python3
"""X6 number formats, fonts, fills, borders. X7 conditional formatting. X9 panes.

These go through the openpyxl object model and then the graft (see
`office/rebuild.py`), because creating a conditional-format rule or a border means
writing styles.xml, the sheet and their rels graph consistently — hand-editing that
is how a workbook ends up in Excel's repair dialog.

    # X6 — a range, formatted
    python3 xlsx_format.py --in book.xlsx --out out.xlsx --sheet 利润表 \
            --range B3:C5 --number-format "#,##0" --font-color FF0000FF --bold \
            --fill FFFFF2CC --border thin

    # X9 — freeze the header rows and put a filter on the header row
    python3 xlsx_format.py --in book.xlsx --out out.xlsx --sheet 利润表 \
            --freeze A3 --filter A2:D5

    # X7 — conditional formatting from a spec file
    python3 xlsx_format.py --in book.xlsx --out out.xlsx --sheet 利润表 \
            --rules rules.json

`rules.json` is a list, because a rule has more moving parts than a flag can carry:

    [{"range": "D3:D5", "type": "cellIs", "operator": "lessThan",
      "formula": ["0"], "font_color": "FF9C0006", "fill": "FFFFC7CE"},
     {"range": "B3:B5", "type": "colorScale",
      "colors": ["FFF8696B", "FFFCFCFF", "FF63BE7B"]},
     {"range": "C3:C5", "type": "dataBar", "color": "FF638EC6"}]

Number formats are written verbatim. `"#,##0"` and `"0.0%"` are Excel's own
notation, not a locale-aware abstraction — inventing one here would mean guessing
what the user's Excel will do with it.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from office.rebuild import RebuildError, rebuild  # noqa: E402
from xlsxcommon import XlsxError, emit, ensure_distinct, fail, run  # noqa: E402

BORDER_STYLES = ("thin", "medium", "thick", "double", "dotted", "dashed", "hair", "none")
RULE_TYPES = ("cellIs", "expression", "colorScale", "dataBar")


def parse_range(spec: str) -> str:
    from openpyxl.utils.cell import range_boundaries
    try:
        c0, r0, c1, r1 = range_boundaries(spec)
    except Exception as e:  # noqa: BLE001
        fail(f"{spec!r} is not a cell or range (expected e.g. B3 or B3:C5): {e}")
    if None in (c0, r0, c1, r1):
        fail(f"{spec!r} is an open-ended range; give both corners, e.g. B3:C5")
    return spec


def rgb(value: str, what: str) -> str:
    """Excel wants AARRGGBB. A bare RRGGBB is the common mistake and is silent —
    the colour simply comes out wrong — so it is completed rather than rejected."""
    v = value.strip().lstrip("#").upper()
    if len(v) == 6:
        v = "FF" + v
    if len(v) != 8 or any(c not in "0123456789ABCDEF" for c in v):
        fail(f"{what} {value!r} is not a hex colour (RRGGBB or AARRGGBB)")
    return v


def load_rules(path: Path) -> list[dict]:
    if not path.is_file():
        fail(f"no such rules file: {path}")
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        fail(f"{path.name} is not valid JSON: {e}")
    if not isinstance(data, list) or not data:
        fail(f"{path.name} must be a non-empty list of rules")
    for i, rule in enumerate(data, 1):
        if not isinstance(rule, dict):
            fail(f"{path.name} rule {i} is not an object")
        if not rule.get("range"):
            fail(f"{path.name} rule {i} has no `range`")
        kind = rule.get("type")
        if kind not in RULE_TYPES:
            fail(f"{path.name} rule {i}: type {kind!r} is not one of "
                 f"{', '.join(RULE_TYPES)}")
        if kind in ("cellIs", "expression") and not rule.get("formula"):
            fail(f"{path.name} rule {i}: a {kind} rule needs `formula`")
        if kind == "colorScale" and len(rule.get("colors") or []) not in (2, 3):
            fail(f"{path.name} rule {i}: a colorScale needs 2 or 3 `colors`")
    return data


def build_rule(rule: dict):
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, Rule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.styles.differential import DifferentialStyle

    kind = rule["type"]
    if kind == "colorScale":
        colors = [rgb(c, "colorScale colour") for c in rule["colors"]]
        if len(colors) == 2:
            return ColorScaleRule(start_type="min", start_color=colors[0],
                                  end_type="max", end_color=colors[1])
        return ColorScaleRule(start_type="min", start_color=colors[0],
                              mid_type="percentile", mid_value=50, mid_color=colors[1],
                              end_type="max", end_color=colors[2])
    if kind == "dataBar":
        return DataBarRule(start_type="min", end_type="max",
                           color=rgb(rule.get("color", "FF638EC6"), "dataBar colour"))
    font = Font(color=rgb(rule["font_color"], "rule font colour")) \
        if rule.get("font_color") else None
    fill = PatternFill(start_color=rgb(rule["fill"], "rule fill"),
                       end_color=rgb(rule["fill"], "rule fill"), fill_type="solid") \
        if rule.get("fill") else None
    if font is None and fill is None:
        fail(f"rule on {rule['range']} changes nothing: give `font_color` or `fill`")
    dxf = DifferentialStyle(font=font, fill=fill)
    return Rule(type=kind, operator=rule.get("operator"), formula=rule["formula"],
                dxf=dxf, stopIfTrue=bool(rule.get("stop_if_true")))


def apply_all(args, rules: list[dict]) -> tuple:
    """Returns (mutate, changes) — `changes` fills in as the mutation runs."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    changes: list[dict] = []

    def mutate(wb):
        if args.sheet and args.sheet not in wb.sheetnames:
            fail(f"no sheet named {args.sheet!r} (have: {', '.join(wb.sheetnames)})")
        ws = wb[args.sheet] if args.sheet else wb.worksheets[0]

        if args.area:
            side = None if args.border in (None, "none") else Side(
                style=args.border,
                color=rgb(args.border_color, "border colour") if args.border_color
                else None)
            border = Border(left=side, right=side, top=side, bottom=side) \
                if args.border else None
            font_args = {}
            if args.font_color:
                font_args["color"] = rgb(args.font_color, "--font-color")
            if args.bold:
                font_args["bold"] = True
            if args.italic:
                font_args["italic"] = True
            if args.font_size:
                font_args["size"] = args.font_size
            if args.font_name:
                font_args["name"] = args.font_name
            fill = PatternFill(start_color=rgb(args.fill, "--fill"),
                               end_color=rgb(args.fill, "--fill"),
                               fill_type="solid") if args.fill else None
            touched = 0
            for row in ws[args.area]:
                for cell in row:
                    if args.number_format:
                        cell.number_format = args.number_format
                    if font_args:
                        # Copy the existing font so setting only a colour does not
                        # silently reset the size and face the cell already had.
                        base = cell.font
                        cell.font = Font(name=font_args.get("name", base.name),
                                         size=font_args.get("size", base.size),
                                         bold=font_args.get("bold", base.bold),
                                         italic=font_args.get("italic", base.italic),
                                         color=font_args.get("color", base.color))
                    if fill is not None:
                        cell.fill = fill
                    if border is not None:
                        cell.border = border
                    if args.wrap:
                        cell.alignment = Alignment(wrap_text=True,
                                                   horizontal=cell.alignment.horizontal,
                                                   vertical=cell.alignment.vertical)
                    touched += 1
            changes.append({"op": "format", "sheet": ws.title, "range": args.area,
                            "cells": touched,
                            "number_format": args.number_format,
                            "font": font_args or None,
                            "fill": args.fill, "border": args.border})

        for rule in rules:
            ws.conditional_formatting.add(parse_range(rule["range"]), build_rule(rule))
            changes.append({"op": "conditional", "sheet": ws.title,
                            "range": rule["range"], "type": rule["type"]})

        if args.freeze:
            ws.freeze_panes = args.freeze
            changes.append({"op": "freeze", "sheet": ws.title, "at": args.freeze})
        if args.filter:
            ws.auto_filter.ref = parse_range(args.filter)
            changes.append({"op": "filter", "sheet": ws.title, "ref": args.filter})
        return {"changes": changes}

    return mutate, changes


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in", dest="src", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    ap.add_argument("--sheet", help="default = the first sheet")
    ap.add_argument("--range", dest="area", metavar="B3:C5")
    ap.add_argument("--number-format", metavar='"#,##0"')
    ap.add_argument("--font-color", metavar="AARRGGBB")
    ap.add_argument("--font-size", type=float)
    ap.add_argument("--font-name")
    ap.add_argument("--bold", action="store_true")
    ap.add_argument("--italic", action="store_true")
    ap.add_argument("--wrap", action="store_true")
    ap.add_argument("--fill", metavar="AARRGGBB")
    ap.add_argument("--border", choices=BORDER_STYLES)
    ap.add_argument("--border-color", metavar="AARRGGBB")
    ap.add_argument("--rules", type=Path, help="conditional formatting spec (JSON)")
    ap.add_argument("--freeze", metavar="A3", help="cell that stays top-left")
    ap.add_argument("--filter", metavar="A2:D5")
    ap.add_argument("--report", type=Path)
    args = ap.parse_args()

    def entry():
        ensure_distinct(args.src, args.out)
        cell_opts = (args.number_format or args.font_color or args.bold or args.italic
                     or args.font_size or args.font_name or args.fill or args.border
                     or args.wrap)
        if args.area and not cell_opts:
            fail("--range was given with nothing to apply; add --number-format / "
                 "--font-color / --fill / --border / --bold / --wrap")
        if cell_opts and not args.area:
            fail("cell formatting needs --range to say which cells")
        rules = load_rules(args.rules) if args.rules else []
        if not (args.area or rules or args.freeze or args.filter):
            fail("nothing to do: pass --range with a format, --rules, --freeze "
                 "or --filter")
        if args.area:
            parse_range(args.area)
        mutate, _ = apply_all(args, rules)
        try:
            report = rebuild(args.src, args.out, mutate)
        except RebuildError as e:
            fail(str(e))
        emit({"in": args.src.name, "out": args.out.name, **report},
             args.report, "changes", "grafted")

    return run(entry)


if __name__ == "__main__":
    raise SystemExit(main())
