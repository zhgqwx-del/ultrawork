#!/usr/bin/env python3
"""Edit a .xlsx: set cells by A1 ref, append rows. ultrawork doc-edit skill."""
from __future__ import annotations
import argparse, sys

def _require():
    try:
        import openpyxl  # noqa: F401
        return openpyxl
    except ImportError:
        print("Missing dependency: openpyxl (pip install openpyxl)", file=sys.stderr)
        raise SystemExit(1)

def _coerce(value: str):
    for cast in (int, float):
        try:
            return cast(value)
        except ValueError:
            continue
    return value

def main(argv):
    ap = argparse.ArgumentParser(description="Edit a .xlsx in place (or --out)")
    ap.add_argument("file")
    ap.add_argument("--set", nargs=2, metavar=("SHEET!A1", "VALUE"), action="append", default=[],
                    help="set a cell, e.g. --set Sheet1!B2 hello")
    ap.add_argument("--append-row", nargs="+", metavar=("SHEET", "VAL"), action="append", default=[],
                    help="append a row: --append-row Sheet1 a b c")
    ap.add_argument("--out", help="write to this path instead of in place")
    args = ap.parse_args(argv)
    openpyxl = _require()
    try:
        wb = openpyxl.load_workbook(args.file)
    except Exception as exc:  # noqa: BLE001
        print(f"Error opening {args.file}: {exc}", file=sys.stderr)
        return 1

    def sheet(name):
        if name not in wb.sheetnames:
            print(f"Sheet not found: {name} (have: {', '.join(wb.sheetnames)})", file=sys.stderr)
            raise SystemExit(1)
        return wb[name]

    sets = 0
    for ref, value in args.set:
        if "!" not in ref:
            print(f"Bad cell ref (need Sheet!A1): {ref}", file=sys.stderr)
            return 1
        sname, cell = ref.split("!", 1)
        sheet(sname)[cell] = _coerce(value)
        sets += 1
    rows = 0
    for spec in args.append_row:
        if len(spec) < 2:
            print("--append-row needs SHEET then at least one value", file=sys.stderr)
            return 1
        sheet(spec[0]).append([_coerce(v) for v in spec[1:]])
        rows += 1
    out = args.out or args.file
    wb.save(out)
    print(f"Saved {out} (cells set: {sets}, rows appended: {rows})")
    return 0

if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
