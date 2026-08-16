#!/usr/bin/env python3
"""L3 —— 真实语料回归门禁（discussions/059 §5 L3，S5）。

**这一层存在的理由是实证不是仪式。** ADR-070 P2（LaTeX）手编用例 100% 通过，两个真缺陷
全靠 279 个真实公式抓出来；S6 那一刀三个真缺陷里两个是「别人产的文档」和「别人的机器」
逼出来的，没有一个是手编用例发现的。L0/L1/L2 三层都只看它们被设计去看的那一面，
而**手编夹具的形状是我想象出来的**。

跑法（059 §5 定的）：全语料过一遍 read → edit → validate 环，统计崩溃率 / 损坏率。

语料：`scripts/l3-corpus-manifest.json` 记清单，字节走缓存（不入 git）。
      先跑 `python3 scripts/fetch-l3-corpus.py`。

────────────────────────────────────────────────────────────────────────────
判据怎么量（三个数分开记，不合并成一个「通过率」）
────────────────────────────────────────────────────────────────────────────

**崩溃** = stderr 出现**裸 traceback**，或进程被信号打死 / 超时。**与退出码约定无关。**
  ⚠️ 第一版我写的是「退出码 ∉ {0,2}」，那是错的，而且错得很隐蔽：新技能的契约是
  `docxcommon.run()` 写下的「exit 2 = 可操作的错误，exit 1 留给真崩溃」，**旧
  doc-edit 没有这个约定** —— 它打印一行干净的 `Error opening …` 然后 return 1。
  按退出码判，旧臂会因为**用了另一套约定**而被记成崩溃，比出来的差距是我造的。
  改成只看「有没有一墙 Python 糊到 agent 脸上」，两条臂才是同一把尺。
  退出码另记一笔 **契约违规**（rc ∉ {0,2} 且无 traceback）—— 只对新技能有意义，
  因为只有它声明过那个契约。

**拒绝** = 非 0 退出但没有裸 traceback。加密件 / 损坏件 / 不支持格式的正确行为。

**损坏** = 输入过了合法性检查、整个环跑完（每步 rc=0），但**输出**没过。
  「没过」的定义直接**复用 L2**（import `office-skills-selftest.py`），
  不是第二套标准 —— 否则 L2 绿而 L3 红时没人说得清哪个对。
  合法性 = D1/D4/D5（docx）· X1/X2（xlsx）；保真度 = F1/F2 · F1/F3 · P5。

**输入本来就坏** 单列，不计进损坏率（垃圾进不算我们的账），但**逐条具名打印** ——
  沉默与通过长得一样。

拒绝也单列。它不是缺陷，但「读不了多少份真实文档」是能力缺口，
  藏在「没崩溃」里就成了好消息。

基线：docx/xlsx 从 git 取 S6 删掉之前的旧 `doc-edit` 脚本（`{BASELINE_REF}`），
      **绝不落回工作树**，跑同一份语料同一个环，两臂的完整命令行逐条打印。
      **pdf 没有前身**（S2 之前的 pdf 技能是「上游 Apache 版 + 零脚本」）⇒
      它没有基线可比，判据只能是绝对值。这一点不编。

用法：
    python3 scripts/test-office-l3-corpus.py                 # 新技能，全语料
    python3 scripts/test-office-l3-corpus.py --baseline      # 加跑旧 doc-edit 对照臂
    python3 scripts/test-office-l3-corpus.py --characterize  # 只打印语料刻画表
    python3 scripts/test-office-l3-corpus.py --selftest      # 门禁自己的正负控制
    python3 scripts/test-office-l3-corpus.py --require-corpus  # 语料缺失 = 红（CI 用）
"""
from __future__ import annotations

import argparse
import collections
import importlib.util
import json
import re
import shutil
import subprocess
import sys
import tempfile
import time
import zipfile
from pathlib import Path

# S6 血泪：Windows 在 stdout **被捕获时**按 ANSI 代码页编码 ⇒ 打印第一个中文字就
# UnicodeEncodeError 退出 1，而 agent 总是捕获 stdout。抄 docxcommon.py 顶部。
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
    except Exception:
        pass

REPO_ROOT = Path(__file__).resolve().parent.parent
SKILLS = REPO_ROOT / "skills" / "builtin"
L2_PATH = REPO_ROOT / "scripts" / "office-skills-selftest.py"
FETCHER = REPO_ROOT / "scripts" / "fetch-l3-corpus.py"

# S6 收官 commit；它的父提交是 doc-edit 还带 docx_*/xlsx_* 四个脚本的最后一版。
BASELINE_REF = "0a5b0987^"
BASELINE_DIR = "skills/builtin/doc-edit/scripts"

PROBE = "L3 探针 probe"          # 刻意含中文：编码问题在纯 ASCII 探针下不现形
STEP_TIMEOUT = 120               # 秒；挂住比崩掉更糟，所以超时按崩溃记
KINDS = ("docx", "xlsx", "pdf")

# 059 §5 L3 写下的目标量。少于这个数就不是「全语料过一遍」。
CORPUS_MIN = {"docx": 20, "xlsx": 15, "pdf": 15}

TRACEBACK_MARK = "Traceback (most recent call last):"

# 具名认领的残余崩溃。**不是豁免名单，是「必须说出理由」名单**：在这里的每次运行都
# 打印出来，不在这里的任何崩溃在没有基线时一律判红。名单变长会在 diff 里显形。
KNOWN_CRASHES: dict[str, str] = {
    "5317294594523136.pdf":
        "pypdf 上游缺陷：该文件的 /Root 指向一个 NumberObject，pypdf 在 "
        "_reader.py:229 抛 AttributeError（不是它自己的 PdfReadError），"
        "所以落不进「文件坏了 = 一句话」那条分支。把 catch-all 放宽到 "
        "AttributeError 会把我们自己未来的 bug 一起藏起来，代价不划算。"
        "该文件来自 ClusterFuzz 生成的畸形样本，不是用户会拿到的文档。",
}

# L2 里哪些检查用来判「输出坏了」。分两类是因为它们回答的问题不同：
#   合法性 —— 产物自身是否还是一个合法的 OOXML/PDF（也拿去判输入健康）
#   保真度 —— 相对输入丢没丢东西（只对输出有意义）
LEGALITY = {"docx": {"D1", "D4", "D5"}, "xlsx": {"X1", "X2"}, "pdf": set()}
FIDELITY = {"docx": {"F1", "F2"}, "xlsx": {"F1", "F3"}, "pdf": {"P5"}}


# ══ 结果类型 ═══════════════════════════════════════════════════════════════════

class Outcome:
    OK = "ok"
    CORRUPT = "corrupt"        # 环跑完了，输出坏了 —— 这是最严重的
    CRASH = "crash"            # 裸 traceback / 被信号打死 / 超时
    REFUSED = "refused"        # 非 0 退出但无 traceback —— 结构化报错，正确行为
    INPUT_BAD = "input_bad"    # 输入本身就不合法，不计我们的账
    NOT_RUN = "not_run"        # 该臂没有对应入口（如 pdf 的基线臂）
    GATE_LIMIT = "gate_limit"  # **门禁自己**判不了这一份（L2 子进程超时/崩了）。
                               # 不计进任何率的分子 —— 那是我的局限，不是技能的缺陷；
                               # 但从可用分母里扣掉并具名，否则它冒充「通过」。



# POSIX 上被信号打死 = 负退出码；**Windows 上没有负数** —— 硬崩溃返回的是
# NTSTATUS 异常码这样的大正数（0xC0000005 访问违例 = 3221225477，
# 0xC000013A Ctrl-C = 3221225786）。只写 `rc < 0` 的话，Windows 上的段错误会被
# 判成「拒绝」，也就是**把崩溃记成正确行为** —— 错的方向里最糟的那个。
# 这不是从 CI 学来的，是本刀在 Windows 换行缺陷之后自查出来的：
# 「不留从未执行过的平台分支」。
NTSTATUS_FAILURE = 0xC0000000


def _died_hard(rc: int) -> bool:
    return rc < 0 or rc >= NTSTATUS_FAILURE


class Step:
    def __init__(self, name: str, cmd: list[str], rc: int, out: str, err: str, secs: float):
        self.name, self.cmd, self.rc = name, cmd, rc
        self.out, self.err, self.secs = out, err, secs
        self.mem_bounded = False       # 这一步有没有被 ulimit 包起来（仅 Linux）
        # 这一步**当时**的时间预算。读全局的话，预算被临时改过再改回时报告会印错
        # 数字（控制臂把预算压到 60s，why() 却照旧印 120s）—— 报告印错数字，
        # 与被测对象印错数字是同一类问题。
        self.budget = STEP_TIMEOUT

    @property
    def crashed(self) -> bool:
        """裸 traceback，或被信号打死 / 超时（rc < 0）。**与退出码约定无关。**

        按 rc ∉ {0,2} 判会让旧 doc-edit 因为「用了另一套退出码约定」被记成崩溃 ——
        它对打不开的文件是 `print(...); return 1`，一行干净的话，不是一墙 Python。
        那样比出来的差距是我造的，不是两个实现的差距。
        """
        if self.hit_mem_cap:
            return False          # 见 hit_mem_cap：这是门禁的局限，单独归类
        if self.hit_ring_mem_bound:
            return True           # 与超时同类
        return TRACEBACK_MARK in self.err or _died_hard(self.rc)

    @property
    def hit_mem_cap(self) -> bool:
        """撞的是**我给的**内存上限 —— 那是门禁的局限，不是被测对象的缺陷。

        只对 L2 worker 成立：上限是它自己给自己设的。环里的外部脚本**没有**上限，
        它们真的抛 MemoryError 时那是一次真实的资源失败，不该被我洗成「门禁局限」。
        """
        if self.name != "l2":
            return False
        return self.rc == MEMCAP_RC or any(m in self.err for m in MEMCAP_MARKS)

    @property
    def refused(self) -> bool:
        return self.rc != 0 and not self.crashed and not self.hit_mem_cap

    # 每个入口自己声明过的退出码。docx_validate 用 **rc=1 表示「有 schema 违规」**，
    # 明细走 stdout —— 那是它 docstring 里写下的约定，不是契约违规。
    # ⚠️ 第一版没有这张表，于是 3 份真实 Word 文档被贴上「契约违规」，
    # 而判红的是我的标签，不是被测对象。
    ALLOWED_RC = {"validate": (0, 1, 2)}

    @property
    def breaks_contract(self) -> bool:
        """新技能自己声明过 exit 2；旧实现没声明过，所以这一笔只对新臂有意义。"""
        allowed = self.ALLOWED_RC.get(self.name, (0, 2))
        return self.rc not in allowed and not self.crashed

    @property
    def hit_ring_mem_bound(self) -> bool:
        """撞的是**环步骤**的内存上界 —— 与超时同类，算被测实现的属性。"""
        return self.mem_bounded and self.rc != 0 and any(
            m in (self.err or "") for m in RING_MEM_MARKS)

    def why(self) -> str:
        """一行说清楚它为什么没过。

        ⚠️ 第一版打的是 stderr 的**第一行** —— 而裸 traceback 的第一行永远是那句
        没有信息量的 `Traceback (most recent call last):`。CI 上四份 Windows 独有的
        崩溃因此完全无法诊断，我不得不本地反推机制。有 traceback 时要打**最后一行**
        （异常类型与消息就在那儿）。
        """
        if self.rc == -9:
            return f"{self.name}: 超时 >{self.budget}s"
        if self.hit_ring_mem_bound:
            return (f"{self.name}: 撞 {RING_MEM_LIMIT_KB // 1024 ** 2}GB 内存上界"
                    f"（与超时同类：这个实现处理不了这份文件）")
        lines = [x.strip() for x in (self.err or "").splitlines() if x.strip()]
        tb = TRACEBACK_MARK in (self.err or "")
        pick = (lines[-1] if tb and lines else (lines[0] if lines else ""))
        return f"{self.name}(rc={self.rc}){' [裸 traceback]' if tb else ''}: {pick[:160]}"


# L2 校验子进程的地址空间上限。⚠️ **只管得住 L2 worker（我们自己的代码）**：
# 环里的外部脚本（技能脚本、旧 doc-edit）只有 120s 的时间边界，**没有内存边界**。
# 每个子进程的地址空间上限。**「机器扛不住」不能被记成「技能崩了」** ——
# 语料里那份最大行号的 workbook 让 L2 的 openpyxl 校验实测吃到 4.7 GB，
# CI 上 4 路并行时把 ubuntu runner 整个打死（run 30911346681：exit 143 +
# "The runner has received a shutdown signal"）。给了上限之后，撞上限的那一份
# 落成 `gate_limit`（门禁自己的局限，具名打印、不进任何率的分子），
# 而不是让整轮消失。POSIX 才有 setrlimit；Windows 上这条是 no-op，
# **所以这个上限在 Windows 上没有生效，别当成三平台都有的保证。**
STEP_MEM_LIMIT = int(__import__("os").environ.get(
    "ULTRAWORK_L3_MEMCAP_MB", str(3 * 1024))) * 1024 ** 2
MEMCAP_MARK = "__L3_MEMCAP__"          # worker 自己打的、确定的标记
# L2 worker 的结果标记：stdout 上除了结果还可能有别人打的字（实测 pymupdf 1.28 的
# `import fitz` 弃用警告），所以结果自带一个只有它会打的前缀，读的一侧只取它之后的。
L2_JSON_MARK = "__L3_L2_JSON__"
MEMCAP_RC = 3                          # 以及一个不与 0/1/2 撞车的退出码
# 分配失败在不同层会以不同措辞出现：Python 抛 MemoryError，C 扩展可能是
# glibc / libstdc++ 的话。全都算「撞上限」，因为它们说的是同一件事。
MEMCAP_MARKS = (MEMCAP_MARK, "MemoryError", "Cannot allocate memory",
                "Out of memory", "std::bad_alloc")


def self_cap_memory() -> str:
    """让**子进程自己**给自己设地址空间上限，并把结果说出来。

    ⚠️ 不用 `preexec_fn`：① macOS 上 `RLIMIT_AS` 设不了（实测 preexec_fn 直接抛
    `SubprocessError`）；② `preexec_fn` 在**多线程父进程**里本来就不安全，而这个
    门禁正是线程池并行的。所以上限由 worker 在自己进程里设，设不上就如实说。
    """
    try:
        import resource
        resource.setrlimit(resource.RLIMIT_AS, (STEP_MEM_LIMIT, STEP_MEM_LIMIT))
        return f"cap={STEP_MEM_LIMIT // 1024 ** 2}MB"
    except Exception as e:                           # noqa: BLE001
        return f"nocap({type(e).__name__})"


# 环里的**外部脚本**（技能脚本、旧 doc-edit）的内存上界。
# ⚠️ 只在 Linux 生效：macOS 的 RLIMIT_AS 设不了，Windows 没有 ulimit。
# 走 `sh -c 'ulimit -v …; exec "$@"'` 而**不是** preexec_fn —— 后者在多线程父进程里
# 不安全，而这个门禁正是线程池并行的。
#
# 为什么必须有：L3 语料里 `issue_174.xlsx`（142 KB，但 max_row=1048576）进入**基线臂**
# 时，旧 `xlsx_read.py` 会 `openpyxl.load_workbook()` 全量加载它 —— 本机同类加载实测
# 4.7 GB。CI 上 ubuntu runner 因此被打死两次（exit 143 + shutdown signal），
# **两次都停在进度条的同一个字符上**，而那一刻唯一无界的进程就是它。
#
# 撞上界与**超时同等对待，都记崩溃**：对被测实现而言「处理不了这份文件」是它的属性
# （新 xlsx_read 在同一份文件上 3.2 秒、几百 MB 就干完了）。这与 L2 worker 撞我自己
# 设的上限不同 —— 那一个才是「门禁的局限」。
RING_MEM_LIMIT_KB = 3 * 1024 * 1024
RING_MEM_MARKS = ("MemoryError", "Cannot allocate memory", "std::bad_alloc",
                  "Out of memory")


def _bounded(cmd: list[str]) -> tuple[list[str], bool]:
    if not sys.platform.startswith("linux"):
        return cmd, False
    return (["sh", "-c", f'ulimit -v {RING_MEM_LIMIT_KB}; exec "$@"', "sh", *cmd], True)


def run_step(name: str, cmd: list[str], cwd: Path | None = None,
             stdin: str | None = None, bound_memory: bool = False) -> Step:
    t0 = time.time()
    bounded = False
    if bound_memory:
        cmd, bounded = _bounded(cmd)
    try:
        r = subprocess.run(cmd, cwd=str(cwd) if cwd else None, capture_output=True,
                           text=True, encoding="utf-8", errors="replace",
                           input=stdin, timeout=STEP_TIMEOUT)
        rc, out, err = r.returncode, r.stdout, r.stderr
    except subprocess.TimeoutExpired as e:
        rc, out, err = -9, (e.stdout or b"").decode("utf-8", "replace") if isinstance(
            e.stdout, bytes) else (e.stdout or ""), f"timeout after {STEP_TIMEOUT}s"
    st = Step(name, cmd, rc, out, err, time.time() - t0)
    st.mem_bounded = bounded
    return st


# ══ 环的定义 ═══════════════════════════════════════════════════════════════════
# 每个环都是**语料无关**的：不依赖某份文档里恰好有什么字。docx 的编辑用
# --append-paragraph 而不是 --replace，就是因为 --replace 要先知道文里有什么。

def first_sheet(src: Path) -> str | None:
    """两条臂必须写同一张表，否则比的是「谁挑了张更好写的表」。

    旧 `xlsx_edit.py --append-row` 的第一个参数**就是表名**（`--append-row Sheet1 a b`），
    新 `xlsx_write.py --append-row` 收的是逗号串、表由 `--sheet` 定。不显式对齐，
    旧臂会因为「第一个值被当成表名、剩下没值了」整片报错 —— 那是我的测量错，
    不是它的缺陷。（S6 同一形状：`$args` 没加引号，整批报 unrecognized arguments。）
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(src, read_only=True)
        try:
            return wb.sheetnames[0] if wb.sheetnames else None
        finally:
            wb.close()
    except Exception:                                # noqa: BLE001
        return None


def ring_new(kind: str, src: Path, td: Path,
             sheet: str | None = None) -> list[tuple[str, list[str]]]:
    py = sys.executable
    s = SKILLS / kind / "scripts"
    if kind == "docx":
        return [
            ("read", [py, str(s / "docx_read.py"), "--in", str(src),
                      "--outline", "--tables", "--text", "--out", str(td / "read.json")]),
            ("edit", [py, str(s / "docx_edit.py"), "--in", str(src),
                      "--out", str(td / "edited.docx"), "--append-paragraph", PROBE,
                      "--report", str(td / "edit.json")]),
            ("validate", [py, str(s / "docx_validate.py"), "--in", str(td / "edited.docx"),
                          "--report", str(td / "validate.json")]),
        ]
    if kind == "xlsx":
        return [
            ("read", [py, str(s / "xlsx_read.py"), "--in", str(src),
                      "--out", str(td / "read.json")]),
            ("edit", [py, str(s / "xlsx_write.py"), "--in", str(src),
                      "--out", str(td / "edited.xlsx"), "--append-row", f"{PROBE},1",
                      *(["--sheet", sheet] if sheet else []),
                      "--report", str(td / "edit.json")]),
            ("validate", [py, str(s / "xlsx_audit.py"), "--in", str(td / "edited.xlsx"),
                          "--out", str(td / "audit.json")]),
        ]
    return [
        ("read", [py, str(s / "pdf_info.py"), "--in", str(src), "--out", str(td / "info.json")]),
        ("read2", [py, str(s / "pdf_extract.py"), "--in", str(src), "--pages", "1",
                   "--out", str(td / "text.json")]),
        ("edit", [py, str(s / "pdf_pages.py"), "--op", "rotate", "--in", str(src),
                  "--out", str(td / "edited.pdf"), "--degrees", "90", "--pages", "1",
                  "--report", str(td / "edit.json")]),
    ]


def ring_baseline(kind: str, src: Path, td: Path, base: Path,
                  sheet: str | None = None) -> list[tuple[str, list[str]]]:
    """旧 doc-edit 的最近等价入口。**它没有 validate 这一步** —— 那本身就是结果之一。"""
    py = sys.executable
    if kind == "docx":
        return [
            ("read", [py, str(base / "docx_read.py"), str(src), "--json"]),
            ("edit", [py, str(base / "docx_edit.py"), str(src),
                      "--append-paragraph", PROBE, "--out", str(td / "edited.docx")]),
        ]
    if kind == "xlsx":
        return [
            ("read", [py, str(base / "xlsx_read.py"), str(src), "--json"]),
            # 旧实现的 --append-row 第一个 token 是表名，见 first_sheet() 的说明。
            ("edit", [py, str(base / "xlsx_edit.py"), str(src),
                      "--append-row", sheet or "Sheet1", PROBE, "1",
                      "--out", str(td / "edited.xlsx")]),
        ]
    return []   # pdf 没有前身


EDITED = {"docx": "edited.docx", "xlsx": "edited.xlsx", "pdf": "edited.pdf"}


# ══ L2 复用 ════════════════════════════════════════════════════════════════════

_L2 = None


def l2():
    global _L2
    if _L2 is None:
        spec = importlib.util.spec_from_file_location("l2", L2_PATH)
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)          # type: ignore[union-attr]
        _L2 = m
    return _L2


GATE_TIMEOUT_MARK = "__L3_GATE_LIMIT__"


def l2_findings(path: Path, only: set[str], baseline: Path | None = None,
                touched: list[int] | None = None) -> list[str]:
    """在**子进程**里跑 L2 检查，带超时。

    ⚠️ 第一版是 in-process 的，被真实语料当场打脸：pdfplumber 的语料里有一份
    1,048,576 行的 workbook，L2 的 X1 用 `openpyxl.load_workbook`（非 read_only）
    读它，实测 **4.7 GB RSS 且不收敛** —— 一份病态文件挂死整轮，CI 上直接 OOM。
    而这正是 L3 存在的意义：手编夹具里不会有这种东西。

    子进程还顺带守住第二件事：lxml / pypdfium 这类 C 扩展在畸形输入上是会 **段错误**
    的，in-process 的话整个门禁跟着死，而且死状看起来像「跑完了」。
    """
    if not only:
        return []
    spec = {"path": str(path), "only": sorted(only),
            "baseline": str(baseline) if baseline else None, "touched": touched or []}
    st = run_step("l2", [sys.executable, str(Path(__file__).resolve()), "--l2-worker"],
                  stdin=json.dumps(spec))
    parse_err = ""
    if st.rc == 0:
        # 只取标记之后的那一段。整条 stdout 当 JSON 读是错的 —— 见 L2_JSON_MARK。
        head, mark, payload = st.out.rpartition(L2_JSON_MARK)
        if not mark:
            parse_err = (f"stdout 里没有 {L2_JSON_MARK}（worker 没走到写结果那一步？）"
                         f" 前 160 字：{st.out.strip()[:160]!r}")
        else:
            try:
                return json.loads(payload)["findings"]
            except Exception as e:                   # noqa: BLE001
                parse_err = f"标记后的内容不是 JSON（{type(e).__name__}）：{payload[:160]!r}"
        del head
    # ⚠️ 三种情况此前塌陷成同一句话，而它引用的是 stderr 的最后一行 ——
    # 那一行永远是 self_cap_memory() 的 cap=/nocap()，与失败原因毫无关系。
    # CI 2026-08-16 三平台同时红在这里：pymupdf 1.28 起 `import fitz` 往 **stdout**
    # 打一行弃用警告，把 JSON 协议冲掉了，而报出来的却是 `rc=0: nocap(ValueError)`。
    why = ("超时 >%ds" % STEP_TIMEOUT if st.rc == -9 else
           f"L2 子进程 rc=0 但结果读不出来：{parse_err}" if parse_err else
           f"L2 子进程 rc={st.rc}: {(st.err or '').strip().splitlines()[-1][:120] if st.err else ''}")
    return [f"{GATE_TIMEOUT_MARK} {'/'.join(sorted(only))} 判不了：{why}"]


def l2_worker() -> int:
    print(self_cap_memory(), file=sys.stderr)   # 让「有没有上限」在结果里看得见
    try:
        return _l2_worker_body()
    except MemoryError:
        # 撞上限最可能发生在 import lxml/openpyxl/fitz 那一步，所以兜的是整个函数体。
        # 打一个**确定的**标记而不是指望上游措辞 —— 上游措辞是会变的。
        print(MEMCAP_MARK, file=sys.stderr)
        return MEMCAP_RC


def _l2_worker_body() -> int:
    """结果走 stdout，而 stdout 不是我们独占的。

    ⚠️ 2026-08-16 CI 三平台同时红：**pymupdf 1.28 起 `import fitz` 会往 stdout 打**
    「warning: The `fitz` API is deprecated…」，而 L2 的 P5 正是 `import fitz`。
    那一行混进结果里 ⇒ `json.loads` 失败 ⇒ 每一份健康 pdf 都被记成「判不了」，
    **而退出码是 0**，所以从外面看像是门禁自己的局限，不像上游改了行为。
    本机（pymupdf 1.27.2.3）一个字都不打，所以本机全绿 —— 又一次「只在别的机器上现形」。

    两道都要，因为它们防的不是同一件事：
    ① 把 stdout 让出来 —— 检查函数与它们 import 的任何库打的字一律去 stderr；
    ② 结果前面打一个标记，读的一侧只取标记之后 —— 万一有 C 扩展直接往 fd 1 写，
       ① 拦不住，而 ② 仍然读得对。
    """
    spec = json.loads(sys.stdin.read())
    expect: dict = {}
    if spec["baseline"]:
        expect["baseline"] = spec["baseline"]
    if spec["touched"]:
        expect["touched_pages"] = spec["touched"]
    real_stdout = sys.stdout
    sys.stdout = sys.stderr
    try:
        findings, _skipped, _inert = l2().run_checks(
            Path(spec["path"]), expect, only=set(spec["only"]),
            allow_missing=l2().ALL_TIERS)
    finally:
        sys.stdout = real_stdout
    real_stdout.write(L2_JSON_MARK + json.dumps({"findings": findings}))
    return 0


REQUIRED_PART = {"docx": "word/document.xml", "xlsx": "xl/workbook.xml"}


def zip_names(z) -> dict[str, str]:
    """条目名 → 原始名，**斜杠归一化后**。

    ⚠️ `zipfile.namelist()` 不是平台无关的：CPython 的 `ZipInfo.__init__` 里有
    `if os.sep != "/" and os.sep in filename: filename = filename.replace(os.sep, "/")`，
    所以一个（违反 ZIP 规范地）把条目存成 `xl\workbook.xml` 的档案，
    **在 Windows 上读回 `xl/workbook.xml`，在 POSIX 上读回 `xl\workbook.xml`**。
    L3 语料里 `issue_530.xlsx` 正是这样：同一份字节，输入门在两台机器上给出相反判定
    （Windows 放行、macOS 判「没有 xl/workbook.xml」），于是两边的分母都不一样。
    测量必须与平台无关，所以这里自己归一化，不依赖 zipfile 的平台行为。
    """
    return {n.replace("\\", "/"): n for n in z.namelist()}


def input_is_bad(kind: str, src: Path) -> list[str]:
    """输入门控 —— **只做结构性判断**，不用任何被测技能站着的那个库。

    ⚠️ 第一版拿 L2 的 X1（`openpyxl.load_workbook`）当输入门，那是循环论证：
    `xlsx_read.py` 自己就站在 openpyxl 上，凡是 openpyxl 读不了的都被划进
    「输入本来就坏」，被测技能的崩溃率自动变好看 —— 分母是它自己划的。
    现在输入门只问三件与实现无关的事：能不能当 zip 打开、必需 part 在不在、
    XML 是不是良构。openpyxl 读不了但结构完好的文件**留在分母里**，
    技能在它上面拒绝还是崩溃，如实计。

    「输入在某条 L2 检查上本来就不合法」另走一条路（`input_legality`）：
    不扣分母，只**屏蔽那一条检查**对输出的判定 —— 否则输入自带的违规会被
    记成我们弄坏的。
    """
    if kind == "pdf":
        # PDF 没有 L2 合法性层，得自己挑一个中立读者。
        # ⚠️ **不能用 pypdf** —— `pdf` 技能自己就站在 pypdf 上（`pdfcommon.open_reader`），
        # 拿它判「输入好不好」就是让被测对象决定自己的分母：凡是它读不了的都被划成
        # 「输入本来就坏」，它的崩溃率自动变好看。改用 pdfminer（完全独立的实现，
        # 且已经因为 pdfplumber 在依赖里）。
        try:
            import contextlib
            import io
            from pdfminer.high_level import extract_pages
            with contextlib.redirect_stderr(io.StringIO()):
                next(iter(extract_pages(str(src), maxpages=1)))
            return []
        except StopIteration:
            return ["pdfminer 解析出 0 页"]
        except Exception as e:                       # noqa: BLE001
            return [f"pdfminer 打不开：{type(e).__name__}: {str(e)[:120]}"]
    try:
        with zipfile.ZipFile(src) as z:
            bad = z.testzip()
            if bad:
                return [f"zip 条目损坏：{bad}"]
            names = set(zip_names(z))
            need = REQUIRED_PART[kind]
            if need not in names:
                return [f"OOXML 包里没有 {need}（可能根本不是这个格式）"]
            from lxml import etree
            raw = zip_names(z)
            for n in (need, "[Content_Types].xml"):
                if n not in names:
                    return [f"OOXML 包里没有 {n}"]
                try:
                    etree.fromstring(z.read(raw[n]))
                except etree.XMLSyntaxError as e:
                    return [f"{n} 不是良构 XML：{str(e)[:100]}"]
    except Exception as e:                           # noqa: BLE001
        return [f"不是可读的 zip：{type(e).__name__}: {e}"]
    return []


def input_legality(kind: str, src: Path) -> set[str]:
    """输入**自带**的 L2 违规检查号。这些检查对该文件的输出判定要屏蔽掉。

    不扣分母 —— 一份 D4 元素序本来就不合规的真实 Word 文档，仍然应该被读、被改、
    被计入崩溃率；只是它的输出再报 D4 时，那不是我们弄坏的。
    """
    f = l2_findings(src, LEGALITY[kind])
    if any(GATE_TIMEOUT_MARK in x for x in f):
        out = set(LEGALITY[kind])         # 判不了 ⇒ 全屏蔽，并在结果里具名
    else:
        out = {x.split()[0] for x in f if x.split()}
    if kind == "docx":
        # 技能自己的 schema 校验也要做输入门控，否则**输入本来就不合 ECMA-376**
        # 的真实文档会被记成技能拒绝了它。L3 实测：45 份公开 Word 夹具里有 3 份
        # 输入就不合规（`<w:tbl>` 缺必需的 `<w:tblPr>`），XSD 判得对，不是假阳性。
        v = run_step("validate", [sys.executable,
                                  str(SKILLS / "docx/scripts/docx_validate.py"),
                                  "--in", str(src)])
        if v.rc != 0:
            out.add("VALIDATE")
    return out


# ══ 单文件跑一遍 ═══════════════════════════════════════════════════════════════

class FileResult:
    def __init__(self, src: Path, kind: str, arm: str):
        self.src, self.kind, self.arm = src, kind, arm
        self.outcome = Outcome.OK
        self.reasons: list[str] = []
        self.steps: list[Step] = []
        self.secs = 0.0
        self.contract: list[str] = []   # rc ∉ {0,2} 且无 traceback（只对新臂有意义）


def run_one(kind: str, src: Path, arm: str, base: Path | None,
            input_verdict: list[str] | None = None,
            masked: set[str] | None = None) -> FileResult:
    res = FileResult(src, kind, arm)
    if input_verdict:
        res.outcome = Outcome.INPUT_BAD
        res.reasons = input_verdict
        return res
    masked = masked or set()
    with tempfile.TemporaryDirectory(prefix="l3-") as tds:
        td = Path(tds)
        sheet = first_sheet(src) if kind == "xlsx" else None
        ring = (ring_new(kind, src, td, sheet) if arm == "new"
                else ring_baseline(kind, src, td, base, sheet))   # type: ignore[arg-type]
        if not ring:
            res.outcome = Outcome.NOT_RUN
            res.reasons = ["该臂没有对应入口"]
            return res
        for name, cmd in ring:
            st = run_step(name, cmd, bound_memory=True)
            res.steps.append(st)
            res.secs += st.secs
            if st.hit_mem_cap:
                res.outcome = Outcome.GATE_LIMIT
                res.reasons.append(f"{st.name}: 撞到门禁给的 "
                                   f"{STEP_MEM_LIMIT // 1024**3} GB 内存上限")
                return res
            if st.breaks_contract:
                res.contract.append(st.why())
            if st.crashed:
                res.outcome = Outcome.CRASH
                res.reasons.append(st.why())
                return res
            if st.refused:
                if st.name == "validate" and "VALIDATE" in masked:
                    # 输入本来就不合 schema，输出同样不合 —— 这是输入的既有属性，
                    # 不是技能拒绝了它，也不是技能弄坏了它。继续走保真度判定。
                    res.reasons.append("ⓘ 输入本来就不合 ECMA-376 schema，"
                                       "输出同样（既有属性，不计拒绝）")
                    continue
                res.outcome = Outcome.REFUSED
                res.reasons.append(st.why())
                return res
        out = td / EDITED[kind]
        if not out.is_file():
            res.outcome = Outcome.CORRUPT
            res.reasons.append("每一步都 rc=0，但产物根本没落地")
            return res
        # 输入自带违规的那几条检查屏蔽掉 —— 它们不是我们弄坏的。
        live = LEGALITY[kind] - masked
        findings = l2_findings(out, live)
        findings += l2_findings(out, FIDELITY[kind], baseline=src,
                                touched=[1] if kind == "pdf" else None)
        if masked:
            res.reasons.append(f"ⓘ 输入自带违规，已屏蔽 {'/'.join(sorted(masked))}")
        gate = [f for f in findings if GATE_TIMEOUT_MARK in f]
        real = [f for f in findings if GATE_TIMEOUT_MARK not in f]
        if real:
            res.outcome = Outcome.CORRUPT
            res.reasons = real
        elif gate:
            res.outcome = Outcome.GATE_LIMIT
            res.reasons = gate
    return res


# ══ 语料 ═══════════════════════════════════════════════════════════════════════

def fetcher_mod():
    spec = importlib.util.spec_from_file_location("l3fetch", FETCHER)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)              # type: ignore[union-attr]
    return m


def collect_corpus(root: Path, doc: dict) -> tuple[dict[str, list[Path]], list[str]]:
    """返回 {kind: [路径]} 与问题列表。**空遍历长得和通过一模一样** ⇒ 这里必查。"""
    problems: list[str] = []
    pub = root / "public"
    if not pub.is_dir():
        return {k: [] for k in KINDS}, [f"缓存目录不存在：{pub}"]
    by_kind: dict[str, list[Path]] = {k: [] for k in KINDS}
    for e in doc["files"]:
        p = pub / e["source"] / e["path"]
        if not p.is_file():
            problems.append(f"清单里的文件缺失：{e['source']}/{e['path']}")
            continue
        by_kind[e["kind"]].append(p)
    # 本地私有语料（用户丢过来的真实文档）：不进清单、不进 CI，但要跑、要单列。
    local: dict[str, list[Path]] = {k: [] for k in KINDS}
    for k in KINDS:
        d = root / "local" / k
        if d.is_dir():
            local[k] = sorted(p for p in d.rglob(f"*.{k}") if p.is_file())
            by_kind[k] += local[k]
    for k in KINDS:
        if not by_kind[k]:
            problems.append(f"{k}: 一个文件都没扫到 —— 空遍历不算通过")
        elif len(by_kind[k]) < CORPUS_MIN[k]:
            problems.append(f"{k}: 只有 {len(by_kind[k])} 份，059 §5 L3 要求 ≥ {CORPUS_MIN[k]}")
    return by_kind, problems


def materialize_baseline(td: Path) -> tuple[Path | None, str]:
    """把旧 doc-edit 脚本从 git 取到临时目录。**绝不落回工作树。**"""
    dest = td / "baseline-doc-edit"
    dest.mkdir(parents=True, exist_ok=True)
    for name in ("docx_read", "docx_edit", "xlsx_read", "xlsx_edit"):
        # ⚠️ encoding 必须显式给。`text=True` 不带 encoding 时用宿主的 locale 编码，
        # Windows 上就是 cp1252，而旧 doc-edit 的脚本里有中文注释 ⇒ 解码失败。
        # 而这个失败**不会以异常的形式交给你**：解码发生在 subprocess 的读取线程里，
        # 那个线程抛异常只会把 traceback 打到 stderr，run() 照常返回，
        # 只是 `stdout` 变成 **None** —— 于是下一行 write_text(None) 才炸，
        # 报的是 `TypeError: data must be str`，离真因十万八千里。
        # CI 第一跑（run 30911346681）就是这么红的。
        r = subprocess.run(["git", "show", f"{BASELINE_REF}:{BASELINE_DIR}/{name}.py"],
                           cwd=str(REPO_ROOT), capture_output=True,
                           text=True, encoding="utf-8", errors="replace")
        if r.returncode != 0:
            return None, (f"取不到 {BASELINE_REF}:{BASELINE_DIR}/{name}.py —— "
                          f"浅克隆？CI 里要 fetch-depth: 0。"
                          f"{(r.stderr or '').strip()[:120]}")
        if not r.stdout:
            # 上面那条已经堵死了已知的成因，但「rc=0 却没有内容」必须是一句话，
            # 不是一个 TypeError —— 沉默地写出一个空基线比失败更糟。
            return None, (f"{BASELINE_REF}:{BASELINE_DIR}/{name}.py 退出码为 0 却没有内容")
        (dest / f"{name}.py").write_text(r.stdout, encoding="utf-8")
    return dest, ""


# ══ 刻画 ═══════════════════════════════════════════════════════════════════════

CJK_RE = re.compile(r"[一-鿿぀-ヿ가-힯]")


def _zparts(p: Path) -> dict[str, bytes]:
    try:
        with zipfile.ZipFile(p) as z:
            return {n: z.read(n) for n in z.namelist()}
    except Exception:                                # noqa: BLE001
        return {}


def _tags(xml: bytes, tag: str) -> str:
    return " ".join(re.findall(rf"<{tag}[^>]*>([^<]*)</{tag}>",
                               xml.decode("utf-8", "replace")))


def characterize(by_kind: dict[str, list[Path]]) -> dict[str, collections.Counter]:
    """语料到底覆盖了什么 —— 在下任何结论之前先摆出来。

    ⚠️ 这个函数第一版自己就有假阳性：xlsx 的「含 CJK」用全部 XML 拼起来数，
    结果把 theme1.xml 里的东亚字体名（宋体 / ＭＳ Ｐゴシック）算成了正文，
    105 份里报 88 份，真值是 1 份。**只看 sharedStrings 与 sheet 的 <t>。**
    """
    prof: dict[str, collections.Counter] = {k: collections.Counter() for k in KINDS}
    for p in by_kind["docx"]:
        d = _zparts(p)
        if not d:
            prof["docx"]["<打不开>"] += 1
            continue
        doc = d.get("word/document.xml", b"")
        allx = b"".join(v for k, v in d.items() if k.startswith("word/") and k.endswith(".xml"))
        if re.search(rb"<w:(ins|del)[ />]", allx):
            prof["docx"]["跟踪修订"] += 1
        if re.search(rb"<w:(rPrChange|pPrChange|tblPrChange|sectPrChange)[ />]", allx):
            prof["docx"]["格式修订"] += 1
        if any(k.startswith("word/comments") for k in d):
            prof["docx"]["批注"] += 1
        if CJK_RE.search(_tags(doc, "w:t")):
            prof["docx"]["正文含 CJK"] += 1
        if b"<w:tbl>" in doc:
            prof["docx"]["表格"] += 1
        if any("header" in k or "footer" in k for k in d):
            prof["docx"]["页眉/页脚"] += 1
        if "word/numbering.xml" in d:
            prof["docx"]["多级编号"] += 1
        if any(k.startswith("word/media/") for k in d):
            prof["docx"]["图片"] += 1
        if b"<w:fldChar" in doc or b"<w:instrText" in doc:
            prof["docx"]["域(TOC/引用)"] += 1
        if b"<w:hyperlink" in doc:
            prof["docx"]["超链接"] += 1
    for p in by_kind["xlsx"]:
        d = _zparts(p)
        if not d:
            prof["xlsx"]["<打不开>"] += 1
            continue
        names = set(d)
        blob = b"".join(v for k, v in d.items() if k.endswith(".xml"))
        app = d.get("docProps/app.xml", b"").decode("utf-8", "replace")
        m = re.search(r"<Application>([^<]*)</Application>", app)
        if m and "Excel" in m.group(1) and "Openpyxl" not in m.group(1):
            prof["xlsx"]["Excel 亲手存的"] += 1
        txt = _tags(d.get("xl/sharedStrings.xml", b""), "t")
        rows = 0
        cross = False
        for k, v in d.items():
            if re.match(r"xl/worksheets/sheet\d+\.xml$", k):
                txt += _tags(v, "t")
                rr = [int(x) for x in re.findall(rb'<row r="(\d+)"', v)]
                rows = max(rows, max(rr) if rr else 0)
                if re.search(rb"<f[ >][^<]*!", v):
                    cross = True
        if CJK_RE.search(txt):
            prof["xlsx"]["正文含 CJK"] += 1
        if cross:
            prof["xlsx"]["跨表公式"] += 1
        if rows > 1000:
            prof["xlsx"][">1000 行"] += 1
        if rows > 10000:
            prof["xlsx"][">1 万行"] += 1
        if any(k.startswith("xl/charts/") for k in names):
            prof["xlsx"]["图表"] += 1
        if b"<mergeCell " in blob:
            prof["xlsx"]["合并单元格"] += 1
        if b"<conditionalFormatting" in blob:
            prof["xlsx"]["条件格式"] += 1
        if any(k.startswith("xl/pivotCache") for k in names):
            prof["xlsx"]["透视表"] += 1
        if any("vbaProject" in k for k in names):
            prof["xlsx"]["宏"] += 1
        if "xl/metadata.xml" in names:
            prof["xlsx"]["metadata part"] += 1
    for p in by_kind["pdf"]:
        b = p.read_bytes()
        if b"/Encrypt" in b:
            prof["pdf"]["加密"] += 1
        if b"/AcroForm" in b:
            prof["pdf"]["AcroForm 表单"] += 1
        if b"/Rotate" in b:
            prof["pdf"]["含旋转页"] += 1
        if not b.startswith(b"%PDF-"):
            prof["pdf"]["文件头不是 %PDF"] += 1
        try:
            import io
            import contextlib
            from pdfminer.high_level import extract_text
            with contextlib.redirect_stderr(io.StringIO()):
                t = extract_text(str(p), maxpages=3) or ""
            if len(t.strip()) < 20:
                prof["pdf"]["无文字层(≈扫描件)"] += 1
            if CJK_RE.search(t):
                prof["pdf"]["含 CJK"] += 1
        except Exception:                            # noqa: BLE001
            prof["pdf"]["文字抽取失败"] += 1
    return prof


def print_profile(by_kind: dict[str, list[Path]], prof: dict[str, collections.Counter]) -> None:
    print("\n══ 语料刻画（下任何结论之前先看这张表）══")
    for k in KINDS:
        n = len(by_kind[k])
        print(f"\n  {k}  {n} 份")
        if not n:
            continue
        for name, v in prof[k].most_common():
            print(f"     {name:<20}{v:>4} 份  {100 * v / n:>3.0f}%")
        print(f"     {'（未列出的 = 0 份）':<20}")


# ══ 汇总 ═══════════════════════════════════════════════════════════════════════

class Tally:
    def __init__(self) -> None:
        self.by: collections.Counter = collections.Counter()
        self.items: list[FileResult] = []

    def add(self, r: FileResult) -> None:
        self.by[r.outcome] += 1
        self.items.append(r)

    @property
    def total(self) -> int:
        return sum(self.by.values())

    @property
    def usable(self) -> int:
        """分母：输入合法、该臂跑得了的那些。"""
        return (self.total - self.by[Outcome.INPUT_BAD] - self.by[Outcome.NOT_RUN]
                - self.by[Outcome.GATE_LIMIT])

    @property
    def completed(self) -> int:
        return self.by[Outcome.OK] + self.by[Outcome.CORRUPT]

    def crash_rate(self) -> float | None:
        return None if not self.usable else self.by[Outcome.CRASH] / self.usable

    def corrupt_rate(self) -> float | None:
        return None if not self.completed else self.by[Outcome.CORRUPT] / self.completed

    def refuse_rate(self) -> float | None:
        return None if not self.usable else self.by[Outcome.REFUSED] / self.usable


def pct(v: float | None) -> str:
    return "  n/a " if v is None else f"{100 * v:5.1f}%"


def run_arm(arm: str, by_kind: dict[str, list[Path]], base: Path | None,
            input_verdicts: dict[Path, list[str]], masks: dict[Path, set[str]],
            jobs: int) -> dict[str, Tally]:
    from concurrent.futures import ThreadPoolExecutor
    out: dict[str, Tally] = {}
    for kind in KINDS:
        t = Tally()
        files = by_kind[kind]
        print(f"\n── {arm} / {kind}：{len(files)} 份（{jobs} 并行）──", flush=True)
        # executor.map 保序 ⇒ 进度符号与逐条报告的次序**与文件次序一致**，
        # 换台机器跑出来的报告可以逐行 diff。
        with ThreadPoolExecutor(max_workers=jobs) as ex:
            results = ex.map(
                lambda p: run_one(kind, p, arm, base,
                                  input_verdicts.get(p, []), masks.get(p, set())),
                files)
            for i, r in enumerate(results, 1):
                t.add(r)
                mark = {Outcome.OK: "·", Outcome.CORRUPT: "D", Outcome.CRASH: "X",
                        Outcome.REFUSED: "r", Outcome.INPUT_BAD: "i",
                        Outcome.NOT_RUN: "-", Outcome.GATE_LIMIT: "?"}[r.outcome]
                sys.stdout.write(mark)
                if i % 80 == 0:
                    sys.stdout.write("\n")
                sys.stdout.flush()
        sys.stdout.write("\n")
        out[kind] = t
        # 逐条具名打印：崩溃 / 损坏 全列；拒绝与输入坏的列出全部名字（沉默 = 通过）
        for label, oc in (("崩溃", Outcome.CRASH), ("损坏", Outcome.CORRUPT)):
            hits = [r for r in t.items if r.outcome == oc]
            if hits:
                print(f"   ⚠️ {label} {len(hits)} 份：")
                for r in hits:
                    print(f"      {r.src.name}")
                    for why in r.reasons[:4]:
                        print(f"         {why}")
        for label, oc in (("拒绝(非0退出，无 traceback)", Outcome.REFUSED),
                          ("输入本来就坏", Outcome.INPUT_BAD),
                          ("⚠️ 门禁自己判不了（我的局限，不是技能的缺陷）", Outcome.GATE_LIMIT)):
            hits = [r for r in t.items if r.outcome == oc]
            if hits:
                print(f"   ⓘ {label} {len(hits)} 份：")
                for r in hits:
                    why = r.reasons[0] if r.reasons else ""
                    print(f"      {r.src.name:<44} {why[:110]}")
        bad = [r for r in t.items if r.contract]
        if bad and arm == "new":
            print(f"   ⚠️ 契约违规（rc ∉ {{0,2}} 且无 traceback）{len(bad)} 份 —— "
                  f"新技能自己声明的是「exit 2 = 可操作的错误」：")
            for r in bad:
                print(f"      {r.src.name:<44} {r.contract[0][:110]}")
    return out


def print_summary(arms: dict[str, dict[str, Tally]]) -> None:
    print("\n══ 汇总 ══")
    print(f"  {'臂':<10}{'类型':<6}{'总数':>5}{'输入坏':>7}{'判不了':>7}{'可用':>6}"
          f"{'崩溃':>6}{'拒绝':>6}{'损坏':>6}   {'崩溃率':>7}{'损坏率':>7}{'拒绝率':>7}")
    for arm, per in arms.items():
        for kind in KINDS:
            t = per[kind]
            if not t.total:
                continue
            print(f"  {arm:<10}{kind:<6}{t.total:>5}{t.by[Outcome.INPUT_BAD]:>7}"
                  f"{t.by[Outcome.GATE_LIMIT]:>7}{t.usable:>6}"
                  f"{t.by[Outcome.CRASH]:>6}{t.by[Outcome.REFUSED]:>6}"
                  f"{t.by[Outcome.CORRUPT]:>6}   "
                  f"{pct(t.crash_rate()):>7}{pct(t.corrupt_rate()):>7}{pct(t.refuse_rate()):>7}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--baseline", action="store_true", help="加跑旧 doc-edit 对照臂")
    ap.add_argument("--characterize", action="store_true", help="只打印语料刻画表")
    ap.add_argument("--selftest", action="store_true", help="门禁自己的正负控制")
    ap.add_argument("--require-corpus", action="store_true",
                    help="语料缺失时判红而不是跳过（CI 用；沉默与通过长得一样）")
    ap.add_argument("--kind", choices=KINDS, action="append", help="只跑某类")
    ap.add_argument("--limit", type=int, help="每类只跑前 N 份（调试用，会在结果里标出）")
    ap.add_argument("--jobs", type=int, default=0,
                    help="并行度（默认 = CPU 数，上限 8）")
    ap.add_argument("--verbose", action="store_true")
    ap.add_argument("--l2-worker", action="store_true",
                    help=argparse.SUPPRESS)   # 内部：在子进程里跑 L2，带超时
    a = ap.parse_args()
    if not a.jobs:
        import os as _os
        a.jobs = max(1, min(8, (_os.cpu_count() or 4)))

    if a.l2_worker:
        return l2_worker()
    if a.selftest:
        return selftest()

    fm = fetcher_mod()
    root = fm.corpus_root()
    doc = fm.load_manifest()
    print(f"语料缓存根：{root}")
    print(f"清单：{len(doc['files'])} 份 / {len(doc['sources'])} 个源 "
          f"（{', '.join(s['id'] + '@' + s['commit'][:8] for s in doc['sources'])}）")

    by_kind, problems = collect_corpus(root, doc)
    if problems:
        print("\n⚠️ 语料有问题：")
        for p in problems[:15]:
            print("   -", p)
        if len(problems) > 15:
            print(f"   … 另有 {len(problems) - 15} 条")
        if a.require_corpus:
            print("\n红：--require-corpus 下语料不完整即失败。先跑 scripts/fetch-l3-corpus.py。")
            return 1
        print("\n跳过 L3（语料不完整）。⚠️ 跳过 ≠ 通过。CI 请传 --require-corpus。")
        return 0

    local_n = {k: sum(1 for p in by_kind[k] if (root / "local") in p.parents
                      or str(p).startswith(str(root / "local"))) for k in KINDS}
    for k in KINDS:
        if local_n[k]:
            print(f"ⓘ {k}: 其中 {local_n[k]} 份来自本地私有语料 "
                  f"（不入 git、**不进 CI** —— 这部分结论只在本机成立）")

    if a.kind:
        by_kind = {k: (v if k in a.kind else []) for k, v in by_kind.items()}
    if a.limit:
        print(f"⚠️ --limit {a.limit}：**这不是全语料回归**，结论不得当作 L3 结果引用。")
        by_kind = {k: v[:a.limit] for k, v in by_kind.items()}

    prof = characterize(by_kind)
    print_profile(by_kind, prof)
    if a.characterize:
        return 0

    print("\n── 输入门控 ──")
    print("   ① 结构性：不是 zip / 缺必需 part / XML 不良构 ⇒ 扣出分母（具名打印）")
    print("      **刻意不用被测技能站着的那个库**（pdf 用 pdfminer 而非 pypdf，"
          "OOXML 只做 zip+lxml 结构检查而非 openpyxl 加载）")
    print("   ② 输入自带的 L2 违规 ⇒ 不扣分母，只屏蔽那一条检查对输出的判定", flush=True)
    from concurrent.futures import ThreadPoolExecutor
    verdicts: dict[Path, list[str]] = {}
    masks: dict[Path, set[str]] = {}
    total_files = sum(len(v) for v in by_kind.values())
    for k in KINDS:
        with ThreadPoolExecutor(max_workers=a.jobs) as ex:
            for p, v in zip(by_kind[k], ex.map(lambda p: input_is_bad(k, p), by_kind[k])):
                if v:
                    verdicts[p] = v
        live = [p for p in by_kind[k] if p not in verdicts]
        with ThreadPoolExecutor(max_workers=a.jobs) as ex:
            for p, m in zip(live, ex.map(lambda p: input_legality(k, p), live)):
                if m:
                    masks[p] = m
    print(f"   ① 结构性不合法：{len(verdicts)} 份 / {total_files}")
    for p, v in sorted(verdicts.items()):
        print(f"      {p.name:<44} {v[0][:110]}")
    print(f"   ② 输入自带 L2 违规：{len(masks)} 份 —— "
          f"{collections.Counter(c for m in masks.values() for c in m).most_common()}")

    arms: dict[str, dict[str, Tally]] = {}
    t0 = time.time()
    arms["new"] = run_arm("new", by_kind, None, verdicts, masks, a.jobs)

    if a.baseline:
        with tempfile.TemporaryDirectory(prefix="l3-base-") as td:
            base, err = materialize_baseline(Path(td))
            if base is None:
                print(f"\n⚠️ 基线臂取不到：{err}")
                print("   ⚠️ 这不是通过 —— 没有基线就没有「≤ 当前基线」这个判据。")
                if a.require_corpus:
                    return 1
            else:
                print(f"\nⓘ 基线臂 = {BASELINE_REF}:{BASELINE_DIR}（旧 doc-edit，"
                      f"S6 删除前的最后一版），落在临时目录，绝不入工作树")
                arms["baseline"] = run_arm("baseline", by_kind, base, verdicts,
                                           masks, a.jobs)

    print_summary(arms)
    print(f"\n耗时 {time.time() - t0:.0f}s")

    print("\n── 两条臂的完整命令行（不做黑箱对比）──")
    for k in KINDS:
        for arm, fn in (
                ("new", lambda kk: ring_new(kk, Path("<IN>"), Path("<TMP>"), "<SHEET1>")),
                ("baseline", lambda kk: ring_baseline(
                    kk, Path("<IN>"), Path("<TMP>"), Path("<BASE>"), "<SHEET1>"))):
            for name, cmd in fn(k):
                short = [c.replace(str(SKILLS), "skills/builtin").replace(sys.executable, "py")
                         for c in cmd]
                print(f"   {arm:<9}{k:<5}{name:<9}{' '.join(short)}")

    # ── 判据 ──────────────────────────────────────────────────────────────
    print("\n── 判据（059 §5「L3 崩溃率 / 损坏率 ≤ 当前基线」）──")
    red: list[str] = []
    for kind in KINDS:
        nt = arms["new"][kind]
        bt = arms.get("baseline", {}).get(kind)
        if nt.by[Outcome.CORRUPT]:
            red.append(f"{kind}: 新技能损坏 {nt.by[Outcome.CORRUPT]} 份 —— "
                       f"损坏是最严重的一类，任何一份都要单独判定它挡不挡合并")
        if bt is None or not bt.usable:
            # ⚠️ 第一版这里只打印绝对值就 `continue` 了，于是 pdf 崩溃 12/73 = 16.4%
            # 照样打印「绿。」并 exit 0 —— **沉默与通过长得一样**，而且是门禁自己犯的。
            # 没有基线 ⇒ 「≤ 基线」这个判据根本无法求值 ⇒ 不许当通过。
            unknown = [r for r in nt.items if r.outcome == Outcome.CRASH
                       and r.src.name not in KNOWN_CRASHES]
            print(f"   {kind}: **没有基线可比**"
                  + ("（pdf 的前身是零脚本的上游 Apache 版，本来就没有）"
                     if kind == "pdf" else "（基线臂没跑）")
                  + f" ⇒ 判据退化成绝对值：崩溃 {nt.by[Outcome.CRASH]}"
                    f"（其中已具名认领 {nt.by[Outcome.CRASH] - len(unknown)}）"
                    f" / 损坏 {nt.by[Outcome.CORRUPT]}")
            for r in nt.items:
                if r.outcome == Outcome.CRASH and r.src.name in KNOWN_CRASHES:
                    print(f"      ⓘ 认领：{r.src.name} —— {KNOWN_CRASHES[r.src.name]}")
            if unknown:
                red.append(f"{kind}: {len(unknown)} 份崩溃没有被具名认领 —— "
                           f"没有基线时任何未认领的崩溃都判红（"
                           f"{', '.join(r.src.name for r in unknown[:5])}）")
            continue
        for label, nv, bv in (("崩溃率", nt.crash_rate(), bt.crash_rate()),
                              ("损坏率", nt.corrupt_rate(), bt.corrupt_rate())):
            if nv is None or bv is None:
                print(f"   {kind} {label}: 分母为 0，判不了")
                continue
            ok = nv <= bv + 1e-9
            print(f"   {kind} {label}: 新 {pct(nv)} vs 旧 {pct(bv)}  "
                  f"{'✅ ≤ 基线' if ok else '❌ 高于基线'}")
            if not ok:
                red.append(f"{kind} {label} 高于基线（{pct(nv)} > {pct(bv)}）")

    print("\n⚠️ L3 验不了的：排版好不好、财务表符不符合行业审美、PDF 版面可不可交付 —— "
          "那是 L4，只能人工判。不拿门禁全绿冒充「更优」。")

    if red:
        print("\n红：")
        for r in red:
            print("   -", r)
        return 1
    print("\n绿。")
    return 0


# ══ L3 抓到的 pdf 回归（C21 / C22）═════════════════════════════════════════════

PDF_SCRIPTS = SKILLS / "pdf" / "scripts"


def _patched(skill: str, work: Path, edits: list[tuple[str, str]], name: str) -> Path:
    """某个技能脚本树的一份副本，逐条打补丁。锚点命中 ≠1 直接 SystemExit。

    抄 test-pptx-edit-skill.py 的 `patched()`：一条**没打上**的控制臂，
    和一条打上了但什么也没改变的控制臂，从外面看一模一样。
    """
    dest = work / f"patched-{skill}-{name}"
    if dest.exists():
        shutil.rmtree(dest)
    shutil.copytree(SKILLS / skill / "scripts", dest,
                    ignore=shutil.ignore_patterns("__pycache__"))
    for old, new in edits:
        hits = 0
        for py in sorted(dest.glob("*.py")):
            text = py.read_text(encoding="utf-8")
            if old in text:
                hits += text.count(old)
                py.write_text(text.replace(old, new), encoding="utf-8")
        if hits != 1:
            raise SystemExit(f"控制臂 {name!r}：锚点命中 {hits} 次（应为 1）"
                             f"—— 这条控制没有复刻那个缺陷")
    return dest


def _crlf_control(work: Path) -> tuple[bool, str]:
    """两条臂：有 / 无 `core.autocrlf=false`，跑在一个强制 autocrlf=true 的宿主上。

    返回 (通过?, 说明)。**不联网** —— origin 指向本地仓库，复现的是 git 自己的行为，
    不是我对它的猜测。
    """
    import os
    fm = fetcher_mod()
    home = work / "crlf-home"
    home.mkdir(parents=True, exist_ok=True)
    (home / "gitconfig").write_text("[core]\n\tautocrlf = true\n", encoding="utf-8")

    origin = work / "crlf-origin"
    (origin / "sub").mkdir(parents=True, exist_ok=True)
    (origin / "LICENSE").write_bytes(b"MIT License\nline two\nline three\n")
    (origin / "sub" / "a.docx").write_bytes(b"PK\x03\x04binary\x00\nnot text\n")
    for cmd in (["git", "init", "-q"],
                ["git", "config", "user.email", "l3@example.invalid"],
                ["git", "config", "user.name", "l3"],
                ["git", "add", "-A"],
                ["git", "commit", "-q", "-m", "seed"]):
        fm.run(cmd, cwd=origin)
    sha = subprocess.run(["git", "rev-parse", "HEAD"], cwd=str(origin),
                         capture_output=True, text=True,
                         encoding="utf-8", errors="replace").stdout.strip()

    saved = os.environ.get("GIT_CONFIG_GLOBAL")
    os.environ["GIT_CONFIG_GLOBAL"] = str(home / "gitconfig")
    arms: dict[str, bytes] = {}
    try:
        for name, keep_fix in (("fixed", True), ("control", False)):
            dest = work / f"crlf-{name}"
            dest.mkdir(parents=True, exist_ok=True)
            fm.run(["git", "init", "-q"], cwd=dest)
            fm.run(["git", "remote", "add", "origin", origin.as_uri()], cwd=dest)
            fm.run(["git", "config", "core.sparseCheckout", "true"], cwd=dest)
            if keep_fix:                       # ← 被测的就是这两行
                fm.run(["git", "config", "core.autocrlf", "false"], cwd=dest)
                fm.run(["git", "config", "core.eol", "lf"], cwd=dest)
            fm.run(["git", "sparse-checkout", "init", "--cone"], cwd=dest)
            fm.run(["git", "sparse-checkout", "set", "--cone", "sub"], cwd=dest)
            fm.run(["git", "fetch", "-q", "--depth", "1", "origin", sha], cwd=dest)
            fm.run(["git", "checkout", "-q", "FETCH_HEAD"], cwd=dest)
            arms[name] = (dest / "LICENSE").read_bytes()
    finally:
        if saved is None:
            os.environ.pop("GIT_CONFIG_GLOBAL", None)
        else:
            os.environ["GIT_CONFIG_GLOBAL"] = saved

    want = (origin / "LICENSE").read_bytes()
    fixed_ok = arms["fixed"] == want
    differs = arms["control"] != want
    # 生产代码里那两行确实在（不然这条控制测的是别人）。
    present = all(x in (REPO_ROOT / "scripts" / "fetch-l3-corpus.py")
                  .read_text(encoding="utf-8")
                  for x in ('"core.autocrlf", "false"', '"core.eol", "lf"'))
    note = (f"修复臂逐字节等于上游={fixed_ok}，生产代码里有那两行={present}；"
            + ("控制臂被 autocrlf 改写了（两臂分得开）" if differs else
               "⚠️ **控制臂与修复臂结果相同** —— 这台宿主的 git 没改写换行，"
               "这条控制这次什么也没证明，按红处理"))
    return fixed_ok and differs and present, note


FAR_ROW = 1_048_576          # Excel 的最大行号
FAR_COLS = 20


def _far_row_book(path: Path) -> None:
    """两行真数据 + 一个落在最大行号上的单元格 —— 野外真实存在的形状。

    ⚠️ 第一版这个夹具只改了 `<dimension>` 字符串，控制臂当场判绿 —— 因为**非
    read_only 的 Worksheet 根本不看 dimension**，它的 max_row 是从实际单元格算的。
    真实文件（calamine issue_174，145 KB）里是**确实有一个 r="1048576" 的 row**。
    按实测重建：max_row 就变成 1048576，乘以列数就是千万级单元格，而出厂实现
    要走它**两遍**。断言在正确实现上判红，说明我对被测对象的描述错了。
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    head = [f"c{i}" for i in range(FAR_COLS)]
    ws.append(head)
    ws.append(list(range(FAR_COLS)))
    wb.save(path)
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        data = {n: z.read(n) for n in names}
    far = (f'<row r="{FAR_ROW}"><c r="A{FAR_ROW}" t="inlineStr">'
           f"<is><t>far</t></is></c></row>").encode()
    for n in list(data):
        if re.match(r"xl/worksheets/sheet\d+\.xml$", n):
            data[n] = data[n].replace(b"</sheetData>", far + b"</sheetData>")
            data[n] = re.sub(rb'<dimension ref="[^"]*"/>',
                             f'<dimension ref="A1:XFD{FAR_ROW}"/>'.encode(),
                             data[n], count=1)
    with zipfile.ZipFile(path, "w") as w:
        for n in names:
            w.writestr(n, data[n])


def _indirect_meta_pdf(path: Path) -> None:
    """一份把 /Producer 存成**间接引用**的 PDF —— 野外常见，本仓库夹具一份都没有。"""
    from pypdf import PdfWriter
    from pypdf.generic import NameObject, TextStringObject
    w = PdfWriter()
    w.add_blank_page(200, 200)
    ref = w._add_object(TextStringObject("L3 indirect producer"))
    info = w._info.get_object() if hasattr(w._info, "get_object") else w._info
    info[NameObject("/Producer")] = ref
    with path.open("wb") as fh:
        w.write(fh)


def _pdf_regressions(work: Path) -> list[tuple[str, str, list[str]]]:
    out: list[tuple[str, str, list[str]]] = []
    src = work / "indirect-meta.pdf"
    _indirect_meta_pdf(src)

    # ── C21：/Producer 是间接引用时，pdf_info 必须产出可用的 JSON。
    #    修复前：`Object of type IndirectObject is not JSON serializable`，裸 traceback，
    #    而且是**agent 对一份文档跑的第一条命令**。L3 语料里 3 份真实文档命中
    #    （两份美国 NICS 背景调查统计、一份学区预算）。
    bad = []
    fixed = run_step("info", [sys.executable, str(PDF_SCRIPTS / "pdf_info.py"),
                              "--in", str(src), "--out", str(work / "i.json")])
    if fixed.rc != 0 or fixed.crashed:
        bad.append(f"修复后仍失败：{fixed.why()}")
    else:
        meta = json.loads((work / "i.json").read_text(encoding="utf-8"))["metadata"]
        if meta.get("producer") != "L3 indirect producer":
            bad.append(f"间接引用没被解开：producer={meta.get('producer')!r}")
    # 控制臂 = 把 _json_safe 原样撤回（复刻发过货的实现，不是随便一种破坏）
    ctrl = _patched("pdf", work, [(
        '    return {k.lstrip("/").lower(): _json_safe(v) for k, v in raw.items()}',
        '    return {k.lstrip("/").lower(): v for k, v in raw.items()}')], "meta")
    c = run_step("info", [sys.executable, str(ctrl / "pdf_info.py"),
                          "--in", str(src), "--out", str(work / "c.json")])
    if not c.crashed:
        bad.append(f"控制臂没崩（rc={c.rc}）—— 两条臂分不开就不是控制臂")
    out.append(("C21", f"/Producer 间接引用：修复后 rc={fixed.rc}，"
                       f"撤回修复后 rc={c.rc} 裸 traceback={TRACEBACK_MARK in c.err}", bad))

    # ── C22：pypdf 自己的异常族必须落成一句话 + exit 2，而不是一墙 Python。
    #    pypdf 是**惰性解析**的，所以 open_reader() 在构造处的兜底只盖住了一小半；
    #    真正的破损在走页面树时才炸。L3 语料里 9 份真实/畸形 PDF 走到这条路上。
    #    激励用注入（合成不出稳定复现延迟抛出的最小 PDF —— pypdf 对小文件恢复力太强），
    #    两条臂的差别就是 `run()` 里那个 `_is_pypdf_error` 分支在不在。
    inject = ('def describe(src: Path, password: str | None) -> dict:\n',
              'def describe(src: Path, password: str | None) -> dict:\n'
              '    from pypdf.errors import PdfReadError\n'
              '    raise PdfReadError("Invalid object in /Pages")\n')
    bad2 = []
    with_fix = _patched("pdf", work, [inject], "pypdf-fixed")
    a = run_step("info", [sys.executable, str(with_fix / "pdf_info.py"), "--in", str(src)])
    if a.rc != 2 or a.crashed:
        bad2.append(f"修复后没落成 exit 2：rc={a.rc} traceback={TRACEBACK_MARK in a.err}")
    without = _patched("pdf", work, [inject, (
        """        if _is_pypdf_error(e):
            print(f"error: this file cannot be parsed as a PDF — "
                  f"{type(e).__name__}: {e}", file=sys.stderr)
            return 2
""", "")], "pypdf-prefix")
    b = run_step("info", [sys.executable, str(without / "pdf_info.py"), "--in", str(src)])
    if not b.crashed or b.rc != 1:
        bad2.append(f"控制臂没退回裸 traceback：rc={b.rc} —— 两条臂分不开")
    out.append(("C22", f"pypdf 异常落成一句话：修复后 rc={a.rc}，"
                       f"撤回分支后 rc={b.rc} 裸 traceback={TRACEBACK_MARK in b.err}", bad2))
    return out


# ══ 自检（正负控制）═════════════════════════════════════════════════════════════
# 「控制臂分不出两种实现 = 不是控制臂」—— 每条控制臂都复刻一个**真实存在过的**
# 错误实现或真实会发生的输入，不是随便一种破坏。

def selftest() -> int:
    passed: list[str] = []
    failed: list[str] = []

    def expect(cid: str, cond: bool, msg: str) -> None:
        (passed if cond else failed).append(f"{cid}: {msg}")
        print(f"   {'PASS' if cond else 'FAIL'}  {cid}  {msg}", flush=True)

    def guarded(cid: str, fn) -> None:
        """一条控制抛异常时记 FAIL 并继续。

        ⚠️ 第一版没有这层：Windows 上 C9 抛 TypeError，**后面 17 条控制一条都没跑**，
        CI 日志里只剩一个与真因无关的 traceback。一个崩掉的 harness 比一个判红的
        harness 告诉你的少得多。
        """
        try:
            fn()
        except Exception as e:                       # noqa: BLE001 - 故意的边界
            import traceback
            failed.append(f"{cid}: 控制自身抛异常 {type(e).__name__}: {e}")
            print(f"   FAIL  {cid}  控制自身抛异常 {type(e).__name__}: {e}", flush=True)
            traceback.print_exc()

    print("══ L3 门禁自检 ══")
    td = Path(tempfile.mkdtemp(prefix="l3-selftest-"))
    try:
        fx = SKILLS / "docx" / "fixtures" / "report.docx"
        xf = SKILLS / "xlsx" / "fixtures" / "book.xlsx"

        # ── C1 正样本：健康输入 + 真实入口 = ok
        r = run_one("docx", fx, "new", None, [])
        expect("C1", r.outcome == Outcome.OK,
               f"健康 docx 过环 → {r.outcome} {r.reasons[:1]}")

        # ── C2 崩溃探测器**真的会响**：注入一个必抛裸 traceback 的入口。
        #     （不验这条，「崩溃率 0」可能只是探测器坏了。）
        boom = td / "boom.py"
        boom.write_text("raise RuntimeError('injected')\n", encoding="utf-8")
        st = run_step("boom", [sys.executable, str(boom)])
        expect("C2", st.crashed and TRACEBACK_MARK in st.err,
               f"裸 traceback 被判成崩溃（rc={st.rc}）")

        # ── C3 exit 2 **不**被判成崩溃：这是技能自己写下的契约，
        #     喂一个改名成 .docx 的文本文件（野外最常见的「不是这个格式」）。
        fake = td / "notreally.docx"
        fake.write_bytes(b"this is plainly not a zip\n")
        st = run_step("read", ring_new("docx", fake, td)[0][1])
        expect("C3", st.rc == 2 and not st.crashed and TRACEBACK_MARK not in st.err,
               f"改名的文本文件 → exit 2 无 traceback（rc={st.rc}）")

        # ── C4 输入门控真的拦得住，且拦的是**输入**不是输出
        expect("C4", bool(input_is_bad("docx", fake)) and not input_is_bad("docx", fx),
               "坏输入判 input_bad、好输入放行")

        # ── C5 损坏探测器 ①：丢一个 part。
        #     这复刻的是真实缺陷 —— 059 §5·补.1 实测 `load→save` 丢 xl/metadata.xml
        #     （904B 真实动态数组元数据），L2 的 F1 就是为它设计的。
        good = td / "good.docx"
        r0 = run_step("edit", [sys.executable, str(SKILLS / "docx/scripts/docx_edit.py"),
                               "--in", str(fx), "--out", str(good),
                               "--append-paragraph", PROBE])
        dropped = td / "dropped.docx"
        with zipfile.ZipFile(good) as z:
            # 受害 part 必须是 L2 眼里**非惰性**的 —— 丢一个空壳 rels 不是数据丢失
            # （§5·补.1 那对正负样本），拿它当控制臂就分不出两种实现。
            victim = next(n for n in z.namelist()
                          if n.startswith("customXml/")
                          and not l2().part_is_inert(n, z.read(n)))
            with zipfile.ZipFile(dropped, "w") as w:
                for n in z.namelist():
                    if n != victim:
                        w.writestr(n, z.read(n))
        f_clean = l2_findings(good, FIDELITY["docx"], baseline=fx)
        f_drop = l2_findings(dropped, FIDELITY["docx"], baseline=fx)
        expect("C5", r0.rc == 0 and not f_clean and bool(f_drop),
               f"丢掉 {victim} → F1/F2 打红（干净 {len(f_clean)} 条 / 丢件 {len(f_drop)} 条）")

        # ── C6 损坏探测器 ②：把 w:pPr 挪到 w:r 后面（ECMA-376 元素序违规，
        #     野外真实存在，D4 就是为它写的）。
        bad_order = td / "order.docx"
        with zipfile.ZipFile(good) as z:
            names = z.namelist()
            data = {n: z.read(n) for n in names}
        doc = data["word/document.xml"]
        m = re.search(rb"(<w:p\b[^>]*>)(<w:pPr>.*?</w:pPr>)(<w:r\b.*?</w:r>)", doc, re.S)
        if m:
            data["word/document.xml"] = doc.replace(
                m.group(0), m.group(1) + m.group(3) + m.group(2), 1)
        with zipfile.ZipFile(bad_order, "w") as w:
            for n in names:
                w.writestr(n, data[n])
        f_ord = l2_findings(bad_order, LEGALITY["docx"])
        expect("C6", m is not None and bool(f_ord),
               f"w:pPr 挪到 w:r 之后 → D1/D4/D5 打红（{len(f_ord)} 条）")

        # ── C7 空遍历必须判红（空遍历长得和通过一模一样）
        empty = td / "empty-cache"
        (empty / "public").mkdir(parents=True)
        fake_doc = {"files": [], "sources": []}
        _bk, probs = collect_corpus(empty, fake_doc)
        expect("C7", len(probs) >= 3, f"空语料 → {len(probs)} 条问题（三类各一条）")

        # ── C8 语料量不足也要判红，而不是「有几份跑几份」
        one = {"files": [{"source": "x", "path": "a.docx", "kind": "docx"}], "sources": []}
        (empty / "public" / "x").mkdir(parents=True, exist_ok=True)
        shutil.copy2(fx, empty / "public" / "x" / "a.docx")
        _bk2, probs2 = collect_corpus(empty, one)
        expect("C8", any("≥ 20" in p for p in probs2),
               f"docx 只有 1 份 → 报「要求 ≥ 20」（{len(probs2)} 条问题）")

        # ── C28 L2 worker 自设的内存上限：撞上限要归成「门禁的局限」，
        #     绝不能记成「技能崩了」。控制臂 = 把上限压到 64MB（必然撞上）。
        #     ⚠️ **两个平台走同一条代码路径**，只有期望值不同 —— 第一版按平台提前
        #     return，于是「设得上上限」那一支在 macOS 上一行都不执行，
        #     里面一个 `st.stdout`（字段其实叫 `out`）的笔误直到 CI 才现形。
        #     不留从未执行过的平台分支，连控制自己也不许留。
        def c28():
            import os as _os
            capable = not self_cap_memory().startswith("nocap")
            # 8MB 而不是 64MB：CI 第五轮实测 **64MB 不够小** —— worker 打印了
            # `cap=64MB` 之后照样干完了活，因为这条路径上的库（openpyxl 是纯 Python、
            # fitz 只在 pdf 检查里才 import）加起来没超过 64MB。
            # 8MB 一定低于解释器自身已经映射的地址空间 ⇒ 之后任何新映射都必然失败。
            # **这个数字是量出来的，不是拍的。**
            env = dict(_os.environ, ULTRAWORK_L3_MEMCAP_MB="8")
            r = subprocess.run(
                [sys.executable, str(Path(__file__).resolve()), "--l2-worker"],
                input=json.dumps({"path": str(SKILLS / "xlsx/fixtures/book.xlsx"),
                                  "only": ["X1"], "baseline": None, "touched": []}),
                capture_output=True, text=True, encoding="utf-8",
                errors="replace", env=env)
            st = Step("l2", [], r.returncode, r.stdout, r.stderr, 0)
            tail = " | ".join(x.strip() for x in (r.stderr or "").splitlines()
                              if x.strip())[-200:]
            if capable:
                # 设得上上限 ⇒ 必须撞上、必须归成门禁局限、必须没有产出
                ok = (not st.out) and st.hit_mem_cap and not st.crashed \
                    and not st.refused
                want = "撞上限并归门禁局限"
            else:
                # 设不上上限 ⇒ 如实说这台机器没有内存边界，且 worker 应当正常干完活
                ok = bool(st.out) and not st.hit_mem_cap and st.rc == 0
                want = "本平台无内存上限，worker 正常完成"
            expect("C28", ok,
                   f"[{sys.platform}] 可设上限={capable} 上限=8MB 期望「{want}」→ "
                   f"rc={r.returncode} hit_mem_cap={st.hit_mem_cap} "
                   f"crashed={st.crashed} refused={st.refused} "
                   f"有产出={bool(st.out)}；stderr: {tail or '(空)'}")
        guarded("C28", c28)

        # ── C27 `text=True` 不给 encoding 是个**静默**陷阱，不是报错。
        #     解码发生在 subprocess 的读取线程里：那个线程抛异常只把 traceback 打到
        #     stderr，run() 照常返回，只是 stdout 变成 **None**。Windows 的 locale
        #     编码是 cp1252，而旧 doc-edit 的脚本带中文注释 ⇒ CI 第一跑就死在这。
        #     两条臂用同一段 UTF-8 中文输出，只差 encoding 参数。
        def c27():
            # ⚠️ 这条控制的**正确期望集依赖平台**，实测量出来的，不是猜的：
            #   subprocess 的读取线程（`_readerthread`）是 **Windows 专有**的，
            #   POSIX 用 selector 在主线程解码。所以同一个解码失败：
            #     Windows → 线程里抛、traceback 打到 stderr、run() 照常返回 stdout=None
            #     POSIX   → 直接抛 UnicodeDecodeError 给调用方
            #   前者是**静默**的，后者不是 —— CI 第一跑正是死在前者上。
            # 另一半在两个平台都成立且更阴：cp1252 对绝大多数字节**有**映射，
            # 所以随便一句中文只会被静默解成**乱码**（内容悄悄错了，不报错）。
            # 探针的字是实测挑的：「不」= E4 B8 8D，命中 cp1252 未定义字节 0x8D；
            # CI 那次撞上的是 0x90。
            # ⚠️ 探针必须**绕过子进程自己的文本编码**直接写原始字节。
            # 第一版用 print()，在 Windows 上子进程按 cp1252 编码 stdout，打第一个
            # 中文字就 UnicodeEncodeError 死掉 ⇒ 管道上**一个字节都没有**，
            # 父进程拿到 ''，解码路径压根没被走到 —— 一条测编码陷阱的控制，
            # 自己被同一个编码陷阱废掉了（S6 那个缺陷的形状，长在控制身上）。
            # 写 sys.stdout.buffer 之后，管道上的字节在三平台完全一致，
            # 唯一变量才是**父进程怎么解码**，也正是要测的东西。
            probe = td / "cjk_out.py"
            probe.write_text(
                "import sys; sys.stdout.buffer.write('不可解码探针'.encode('utf-8'))\n",
                encoding="utf-8")
            plain = td / "mojibake_out.py"
            # 乱码臂的字也是**实测挑的**：它的每个字节都必须能被 cp1252 映射，
            # 否则这一臂也会变成「解不动」那一臂，两种坏法就分不开了。
            # （第一版随手写「中文乱码探针」，「码」= E7 A0 81 命中未定义字节 0x81，
            #   控制当场自爆 —— 断言判红三次都是我对被测对象的描述错了。）
            plain.write_text(
                "import sys; sys.stdout.buffer.write('乮垺敱知'.encode('utf-8'))\n",
                encoding="utf-8")

            def run_as(f, enc, **kw):
                return subprocess.run([sys.executable, str(f)], capture_output=True,
                                      text=True, encoding=enc, **kw)

            # ① 两平台共有：可映射字节 → 内容悄悄错了，不抛异常
            moji = run_as(plain, "cp1252")
            common = (moji.stdout is not None
                      and "乮垺敱知" not in moji.stdout)
            # ② 平台相关：未定义字节
            try:
                none_arm = run_as(probe, "cp1252")
                branch = f"stdout={none_arm.stdout!r}（不抛异常）"
                platform_ok = none_arm.returncode == 0 and none_arm.stdout is None
                observed = "silent-None"
            except UnicodeDecodeError as e:
                branch = f"抛 UnicodeDecodeError: {str(e)[:48]}"
                platform_ok = True
                observed = "raises"
            expected = "silent-None" if sys.platform == "win32" else "raises"
            # ③ 修复本身：utf-8 一定拿得到原文
            good = run_as(probe, "utf-8", errors="replace")
            expect("C27",
                   common and platform_ok and observed == expected
                   and good.stdout is not None and "不可解码探针" in good.stdout,
                   f"[{sys.platform}] 未定义字节 → 量到 {observed}（该平台应为 "
                   f"{expected}）：{branch}；可映射字节 → "
                   f"{(moji.stdout or '').strip()[:12]!r} 内容悄悄错了；utf-8 → 正确")
        guarded("C27", c27)

        # ── C9 基线臂真的取得到，且与新臂**不是同一个东西**
        base, err = materialize_baseline(td)
        if base is None:
            expect("C9", False, f"基线取不到：{err}")
        else:
            nb = {c for _n, cmd in ring_new("docx", fx, td) for c in cmd}
            bb = {c for _n, cmd in ring_baseline("docx", fx, td, base) for c in cmd}
            old_src = (base / "docx_edit.py").read_text(encoding="utf-8")
            new_src = (SKILLS / "docx/scripts/docx_edit.py").read_text(encoding="utf-8")
            expect("C9", nb != bb and "import docx" in old_src and "import docx" not in new_src,
                   "基线臂命令行与新臂不同，且旧实现用 python-docx / 新实现不用")

        # ── C10 崩溃规则**与退出码约定无关**，两条臂才是同一把尺。
        #     这条是我第一版真写错的地方：按 rc ∉ {0,2} 判，旧 doc-edit 对打不开的
        #     文件 `print('Error opening …'); return 1` 会被整片记成崩溃 ——
        #     比出来的差距是我造的。
        tb = Step("t", [], 1, "", TRACEBACK_MARK, 0)
        clean1 = Step("t", [], 1, "", "Error opening x.docx: not a zip", 0)   # 旧实现的形状
        clean2 = Step("t", [], 2, "", "error: not a Word document", 0)        # 新实现的形状
        sig = Step("t", [], -9, "", "timeout", 0)
        # Windows 上没有负退出码：硬崩溃是 NTSTATUS 大正数。只写 rc<0 的话，
        # Windows 的段错误会被判成「拒绝」= 把崩溃记成正确行为。
        winseg = Step("t", [], 0xC0000005, "", "", 0)      # 访问违例
        expect("C10", tb.crashed and sig.crashed and winseg.crashed
               and not clean1.crashed and clean1.refused
               and not clean2.crashed and clean2.refused,
               f"裸 traceback / 信号 / Windows NTSTATUS({0xC0000005}) = 崩溃；"
               f"旧的 rc=1 干净报错与新的 rc=2 都只算拒绝")

        # ── C11 契约违规单独记一笔，且只有新臂才有意义
        expect("C11", clean1.breaks_contract and not clean2.breaks_contract
               and not tb.breaks_contract,
               "rc=1 无 traceback = 契约违规；rc=2 不是；裸 traceback 归崩溃不重复记")

        # ── C12 两条臂写的是**同一张表**。
        #     旧 --append-row 第一个 token 是表名，新的是逗号串 + --sheet。不对齐的话
        #     旧臂整片报「needs SHEET then at least one value」—— 我的测量错，不是它的缺陷。
        sh = first_sheet(xf)
        nb = " ".join(c for _n, cmd in ring_new("xlsx", xf, td, sh) for c in cmd)
        bb = " ".join(c for _n, cmd in ring_baseline("xlsx", xf, td, td, sh) for c in cmd)
        expect("C12", bool(sh) and f"--sheet {sh}" in nb and f"--append-row {sh} " in bb,
               f"两条臂都写 {sh!r}（新 --sheet / 旧 --append-row 首参）")

        # ── C13 xlsx 环也真的能跑（不是只有 docx 被验过）
        r = run_one("xlsx", xf, "new", None, [])
        expect("C13", r.outcome == Outcome.OK,
               f"健康 xlsx 过环 → {r.outcome} {r.reasons[:1]}")

        # ── C14 pdf 环也真的能跑
        pf = SKILLS / "pdf" / "fixtures" / "report-cjk.pdf"
        r = run_one("pdf", pf, "new", None, [])
        expect("C14", r.outcome == Outcome.OK,
               f"健康 pdf 过环 → {r.outcome} {r.reasons[:1]}")

        # ── C31 结果协议扛得住「别人也在往 stdout 打字」
        #     真实来源：pymupdf 1.28 起 `import fitz`（L2 的 P5 就是这么导的）会往
        #     **stdout** 打一行弃用警告。2026-08-16 的 CI 三平台同时红在 C14 上，
        #     而本机 pymupdf 1.27 一个字不打 ⇒ 本机全绿。这条控制臂**不绑定某个库、
        #     也不绑定某个版本**：用 sitecustomize 让子进程在启动时就往 stdout 打字，
        #     等价于「某个被 import 的东西打了字」。修复前它必红（实测：C14 那句
        #     `rc=0: nocap(...)` 的形状），修复后必绿。
        import os as _os
        noise_dir = td / "stdout-noise"
        noise_dir.mkdir(exist_ok=True)
        (noise_dir / "sitecustomize.py").write_text(
            "import sys\n"
            "print('warning: some library is chatty on import', file=sys.stdout)\n",
            encoding="utf-8")
        old_pp = _os.environ.get("PYTHONPATH")
        _os.environ["PYTHONPATH"] = (str(noise_dir) + _os.pathsep + old_pp) if old_pp \
            else str(noise_dir)
        try:
            noisy = run_one("pdf", pf, "new", None, [])
        finally:
            if old_pp is None:
                _os.environ.pop("PYTHONPATH", None)
            else:
                _os.environ["PYTHONPATH"] = old_pp
        expect("C31", noisy.outcome == Outcome.OK,
               f"子进程 stdout 被别人打了字 → 仍然 {noisy.outcome} {noisy.reasons[:1]}")

        # ── C15 刻画函数不再把 theme 里的字体名当成正文 CJK。
        #     这是它第一版真犯过的错：105 份报「84% 含 CJK」，真值 1 份 —— 因为
        #     Office 的 theme1.xml 里有 <a:font typeface="宋体"/>。
        #     控制臂 = 一份**只有 theme 有中文、正文一个中文都没有**的 workbook：
        #     正确实现报 0，我那个错误实现（拼全部 XML 再搜）报 1。两臂必须分得开。
        theme_only = td / "theme-only-cjk.xlsx"
        with zipfile.ZipFile(xf) as z:
            names = z.namelist()
            data = {n: z.read(n) for n in names}
        for n in list(data):
            if n.endswith(".xml") and CJK_RE.search(data[n].decode("utf-8", "replace")):
                data[n] = CJK_RE.sub("X", data[n].decode("utf-8", "replace")).encode("utf-8")
        theme_part = next((n for n in names if "theme" in n), None)
        if theme_part:
            data[theme_part] = data[theme_part].replace(
                b"</a:theme>", '<!-- <a:latin typeface="宋体"/> --></a:theme>'.encode())
        with zipfile.ZipFile(theme_only, "w") as w:
            for n in names:
                w.writestr(n, data[n])
        prof = characterize({"docx": [], "xlsx": [theme_only], "pdf": []})
        naive = CJK_RE.search(b"".join(
            v for k, v in _zparts(theme_only).items() if k.endswith(".xml")
        ).decode("utf-8", "replace"))
        expect("C15", theme_part is not None and prof["xlsx"]["正文含 CJK"] == 0 and bool(naive),
               f"只有 theme 有中文的 workbook：正确实现报 "
               f"{prof['xlsx']['正文含 CJK']}，错误实现（拼全部 XML）会报 1")

        # ── C16 探针含中文：编码缺陷在纯 ASCII 探针下不现形（S6 的 Windows 缺陷就这样）
        expect("C16", bool(CJK_RE.search(PROBE)), f"探针 {PROBE!r} 含 CJK")

        # ── C17 超时**真的**被抓成崩溃，不是「跑完了」。
        #     这条不是假设：语料里那份 1,048,576 行的 workbook 让 in-process 的 L2
        #     吃到 4.7 GB 且不收敛，第一版会挂死整轮。
        global STEP_TIMEOUT
        saved, STEP_TIMEOUT = STEP_TIMEOUT, 1
        try:
            st = run_step("slow", [sys.executable, "-c", "import time; time.sleep(5)"])
        finally:
            STEP_TIMEOUT = saved
        expect("C17", st.crashed and st.rc == -9, f"5s 的活 / 1s 预算 → 崩溃（rc={st.rc}）")

        # ── C19 输入门**与被测实现无关**：一份 openpyxl 打不开、但结构完好的
        #     workbook 必须留在分母里。（拿 openpyxl 当输入门 = 让被测对象划自己的分母。）
        headless = td / "headless.xlsx"
        with zipfile.ZipFile(xf) as z:
            keep = [n for n in z.namelist()
                    if not re.match(r"xl/worksheets/sheet\d+\.xml$", n)]
            with zipfile.ZipFile(headless, "w") as w:
                for n in keep:
                    w.writestr(n, z.read(n))
        expect("C19", input_is_bad("xlsx", headless) == [] and first_sheet(headless) is None,
               "缺工作表 part：结构门放行（留在分母），而 openpyxl 确实打不开")

        # ── C20 输入自带的违规不算我们弄坏的。
        #     用的缺陷形状是 D5 自己 docstring 点名的野外形状：`<w:del>` 里放 `<w:t>`
        #     而不是 `<w:delText>`。选它而不是 C6 那份元素序违规，是因为**技能会拒绝
        #     元素序违规的输入**（第一版这么写，两条臂都是 refused，控制臂分不出两种
        #     实现）—— 断言在正确实现上判红，说明我对被测对象的描述错了。
        d5bad = td / "d5-input-defect.docx"
        with zipfile.ZipFile(fx) as z:
            names = z.namelist()
            data = {n: z.read(n) for n in names}
        doc = data["word/document.xml"]
        mrun = re.search(rb"(<w:r\b[^>]*>.*?</w:r>)", doc, re.S)
        data["word/document.xml"] = doc.replace(
            mrun.group(1),
            b'<w:del w:id="9001" w:author="L3" w:date="2026-08-04T00:00:00Z">'
            + mrun.group(1) + b"</w:del>", 1)
        with zipfile.ZipFile(d5bad, "w") as w:
            for n in names:
                w.writestr(n, data[n])
        m_in = input_legality("docx", d5bad)
        r_masked = run_one("docx", d5bad, "new", None, [], m_in)
        r_unmasked = run_one("docx", d5bad, "new", None, [], set())
        expect("C20", m_in == {"D5"} and r_masked.outcome == Outcome.OK
               and r_unmasked.outcome == Outcome.CORRUPT,
               f"输入自带 {sorted(m_in)}：屏蔽后 {r_masked.outcome} / 不屏蔽 "
               f"{r_unmasked.outcome}（两臂分得开才算控制臂）")

        # ── C23 每个入口自己声明过的退出码不算契约违规。
        #     docx_validate 用 rc=1 表示「有 schema 违规」（明细走 stdout）——
        #     第一版没这张表，3 份真实 Word 文档被我贴上「契约违规」。
        expect("C23", not Step("validate", [], 1, "", "", 0).breaks_contract
               and Step("read", [], 1, "", "", 0).breaks_contract,
               "validate 的 rc=1 是它的约定；read 的 rc=1 不是")

        # ── C24 输入本来就不合 ECMA-376 时，输出同样不合**不算技能拒绝了它**。
        #     缺陷形状取自 L3 真实语料：`<w:tbl>` 缺必需的 `<w:tblPr>`（45 份公开
        #     Word 夹具里 3 份如此，XSD 判得对，不是假阳性）。
        noprops = td / "tbl-without-tblPr.docx"
        with zipfile.ZipFile(fx) as z:
            names = z.namelist()
            data = {n: z.read(n) for n in names}
        doc2 = data["word/document.xml"]
        data["word/document.xml"] = re.sub(rb"<w:tblPr>.*?</w:tblPr>", b"", doc2, count=1,
                                           flags=re.S)
        with zipfile.ZipFile(noprops, "w") as w:
            for n in names:
                w.writestr(n, data[n])
        m24 = input_legality("docx", noprops)
        r_on = run_one("docx", noprops, "new", None, [], m24)
        r_off = run_one("docx", noprops, "new", None, [], m24 - {"VALIDATE"})
        expect("C24", "VALIDATE" in m24 and r_on.outcome != Outcome.REFUSED
               and r_off.outcome == Outcome.REFUSED,
               f"输入缺 w:tblPr：门控开 {r_on.outcome} / 关 {r_off.outcome}")

        # ── C30 环里的步骤在 Linux 上有内存上界，撞上界与超时同等记崩溃。
        #     ⚠️ 两平台走**同一条**代码路径（都调 run_step(bound_memory=True)），
        #     只有期望值按「这台机器包得上 ulimit 吗」分叉 —— C28 的教训：
        #     按平台提前 return 会让那一支在本机一行都不执行。
        def c30():
            hog = td / "ringhog.py"
            hog.write_text(
                "b = []\n"
                "while True:\n"
                "    b.append(bytearray(64 * 1024 * 1024))\n", encoding="utf-8")
            saved, gl = STEP_TIMEOUT, globals()
            gl["STEP_TIMEOUT"] = 60
            try:
                st = run_step("edit", [sys.executable, str(hog)], bound_memory=True)
            finally:
                gl["STEP_TIMEOUT"] = saved
            linux = sys.platform.startswith("linux")
            if linux:
                ok = st.mem_bounded and st.hit_ring_mem_bound and st.crashed \
                    and not st.hit_mem_cap
                want = "撞上界并记崩溃（与超时同类）"
            else:
                # 包不上就如实说没有上界；此时它只会撞 60s 超时（也算崩溃）
                ok = (not st.mem_bounded) and st.crashed and not st.hit_ring_mem_bound
                want = "本平台没有内存上界，只靠超时兜底"
            expect("C30", ok,
                   f"[{sys.platform}] 期望「{want}」→ 包上界={st.mem_bounded} "
                   f"撞上界={st.hit_ring_mem_bound} 崩溃={st.crashed} rc={st.rc}；"
                   f"{st.why()[:80]}")
        guarded("C30", c30)

        # ── C29 输入门必须与平台无关：一个把条目名存成反斜杠的（违规但真实存在的）
        #     档案，在 Windows 与 POSIX 上 `zipfile.namelist()` 给的名字是不同的
        #     （CPython 的 ZipInfo.__init__ 会把 os.sep 换成 "/"）。语料里
        #     issue_530.xlsx 就是这样，害得两台机器的分母都不一样。
        def c29():
            bs = td / "backslash-entries.xlsx"
            with zipfile.ZipFile(xf) as z:
                items = [(n, z.read(n)) for n in z.namelist()]
            with zipfile.ZipFile(bs, "w") as w:
                for n, data in items:
                    w.writestr(n.replace("/", "\\"), data)   # 违规写法，野外真有
            with zipfile.ZipFile(bs) as z:
                naive = "xl/workbook.xml" in set(z.namelist())   # 平台相关
                fixed = "xl/workbook.xml" in set(zip_names(z))   # 归一化后
            verdict = input_is_bad("xlsx", bs)
            expect("C29", fixed and not verdict and naive == (sys.platform == "win32"),
                   f"[{sys.platform}] 反斜杠条目名：裸 namelist 命中={naive}"
                   f"（平台相关，Windows 应为 True）· 归一化后命中={fixed} · "
                   f"输入门判定={verdict or '放行'}")
        guarded("C29", c29)

        # ── C26 取语料的检出必须**逐字节等于上游**，与宿主的 git 换行设置无关。
        #     这是新门禁上 CI 第一跑就被抓到的、只在 Windows 上犯的错：Windows 的 git
        #     默认 `core.autocrlf=true`，把 LICENSE 的 LF 换成 CRLF ⇒ sha256 与在
        #     macOS 上建清单时记的值对不上，脚本报「许可变了」并退出 1。
        #     控制臂 = 撤掉那两行 config（= CI 上红掉的那版）。不联网：用一个本地
        #     git 仓库复现 git 自己的行为，而不是复现我对它的猜测。
        expect("C26", *_crlf_control(td))

        # ── C25 一张**声称**自己是整表大小、实际只有两行的 workbook 必须秒回，
        #     且报出的是真实范围。L3 实测：这样的 145 KB 文件让 xlsx_read 十分钟
        #     没返回，并把它报成 1048576 行 × 16384 列。
        #     控制臂 = 把无界的两遍扫描原样撤回（复刻发过货的实现）。
        liar = td / "far-row.xlsx"
        _far_row_book(liar)
        t0 = time.time()
        good = run_step("read", [sys.executable, str(SKILLS / "xlsx/scripts/xlsx_read.py"),
                                 "--in", str(liar), "--out", str(td / "liar.json")])
        secs = time.time() - t0
        rows = cols = -1
        if good.rc == 0:
            sh = json.loads((td / "liar.json").read_text(encoding="utf-8"))["sheets"][0]
            rows, cols = sh["rows"], sh["columns"]
        # 控制臂 = 出厂到 2026-08-04 为止那版的**逐字**实现：两遍无界扫描。
        old = _patched("xlsx", td, [(
            """        formulas = uncalculated = scanned = 0
        cjk = truncated = False
        rows = cols = 0
        for row in ws.iter_rows():""",
            """        formulas = uncalculated = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
                    if vws[cell.coordinate].value is None:
                        uncalculated += 1
        cjk = any(has_cjk(str(c.value)) for r in ws.iter_rows() for c in r
                  if isinstance(c.value, str))
        rows, cols, scanned, truncated = ws.max_row, ws.max_column, 0, False
        for row in ():""")], "unbounded")
        saved2, STEP_TIMEOUT_LOCAL = STEP_TIMEOUT, 20
        globals()["STEP_TIMEOUT"] = STEP_TIMEOUT_LOCAL
        try:
            ctl = run_step("read", [sys.executable, str(old / "xlsx_read.py"),
                                    "--in", str(liar)])
        finally:
            globals()["STEP_TIMEOUT"] = saved2
        expect("C25", good.rc == 0 and rows == 2 and cols == FAR_COLS and secs < 60
               and ctl.rc == -9,
               f"最大行号上有一个单元格：修复后 {secs:.1f}s 报 {rows}行×{cols}列"
               f"（真实范围，不是 {FAR_ROW}）；出厂版 20s 预算内跑不完（rc={ctl.rc}）")

        # ── C21 / C22：L3 在真实语料上抓到的两个 pdf 缺陷，钉在这里防复发。
        #     两条都是**只有别人产的文档才暴露得出来**的形状，手编夹具一次都没碰到。
        for cid, msg, findings in _pdf_regressions(td):
            expect(cid, not findings, msg + ("" if not findings else f" —— {findings[0]}"))

        # ── C18 「门禁自己判不了」与「技能弄坏了」必须分开，否则我的局限会冒充成缺陷。
        #     走真实的子进程失败路径（喂一个不存在的路径，worker 会抛）。
        ghost = l2_findings(td / "no-such-file.docx", {"D1"})
        expect("C18", len(ghost) == 1 and GATE_TIMEOUT_MARK in ghost[0],
               f"L2 子进程失败 → 标成 gate_limit 而不是 corrupt（{ghost[0][:70] if ghost else '无'}）")

    finally:
        shutil.rmtree(td, ignore_errors=True)

    print(f"\n{len(passed)} passed / {len(failed)} failed")
    return 0 if not failed else 1


if __name__ == "__main__":
    sys.exit(main())
