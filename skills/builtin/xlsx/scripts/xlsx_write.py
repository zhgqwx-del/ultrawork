#!/usr/bin/env python3
"""X2 — write cells and append rows. X15 — column widths that count CJK properly.

This edits `xl/worksheets/sheetN.xml` inside the package and leaves every other
byte alone. It does NOT go through `openpyxl.load_workbook(...) → save()`, and the
reason is measured, not stylistic: that round trip on the repo's own `sample.xlsx`
drops `xl/metadata.xml` (904 bytes of dynamic-array metadata) and the
`<ignoredErrors>` element inside `sheet1.xml`, for a no-op edit. Charts, pivot
caches, macros, threaded comments and cell metadata are all in the same position —
present in the input, understood by nothing in the writer, gone from the output.

    python3 xlsx_write.py --in book.xlsx --out out.xlsx --set 利润表!B4=1240
    python3 xlsx_write.py --in book.xlsx --out out.xlsx \
            --set B4=1240 --set-formula D4=C4-B4 --sheet 利润表
    python3 xlsx_write.py --in book.xlsx --out out.xlsx --append-row 现金,1200,=B2*2
    python3 xlsx_write.py --in book.xlsx --out out.xlsx --autofit          # X15

`--autofit` is not cosmetic. A column holding 营业收入合计 at the 8.43 default is
displayed truncated, and a numeric column too narrow shows `####` — the value is
in the file and invisible on screen. Widths are computed in DISPLAY width, where a
Han character is two units; using `len()` produces a column that is half the size
it needs to be, which is the state every workbook the old doc-edit skill touched
was left in.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from office.formula import (  # noqa: E402
    build_graph, error_tokens_in, missing_sheets, references,
)
from office.package import Package  # noqa: E402
from office.sheet import Workbook, col_to_index, index_to_col, parse_ref  # noqa: E402
from office.validate import check_package  # noqa: E402
from xlsxcommon import (  # noqa: E402
    XlsxError, displayed_text, emit, ensure_distinct, fail, has_cjk, needed_width,
    run,
)


def split_target(spec: str, default_sheet: str | None) -> tuple[str | None, str, str]:
    """'利润表!B4=1240' -> (sheet, ref, value); '=' inside the value is kept."""
    if "=" not in spec:
        fail(f"--set expects REF=VALUE, got {spec!r}")
    target, _, value = spec.partition("=")
    sheet = default_sheet
    if "!" in target:
        sheet, _, target = target.partition("!")
        sheet = sheet.strip("'")
    return sheet, target.strip(), value


def coerce(text: str):
    """Turn the command-line string into the type the cell should hold.

    A quoted value stays text no matter what it looks like — the escape hatch for
    order codes, phone numbers and anything else where 007 must not become 7.
    """
    if len(text) >= 2 and text[0] == text[-1] and text[0] in "\"'":
        return text[1:-1]
    if text == "":
        return None
    low = text.strip().lower()
    if low in ("true", "false"):
        return low == "true"
    try:
        return int(text)
    except ValueError:
        pass
    try:
        return float(text)
    except ValueError:
        return text


def check_formula(expr: str, sheets: set[str], ref: str) -> None:
    """Refuse a formula that is already broken, before it reaches the file (X10).

    Both of these are stored without complaint by every library in this stack and
    only become visible when Excel opens the workbook — which is the worst possible
    moment, because by then nobody remembers which step wrote the cell.
    """
    gone = missing_sheets(expr, sheets)
    if gone:
        fail(f"{ref} refers to sheet {', '.join(repr(g) for g in gone)}, which this "
             f"workbook does not have (sheets: {', '.join(sorted(sheets))}). Writing "
             f"it would store a formula that becomes #REF! the moment Excel opens "
             f"the file")
    baked = error_tokens_in(expr)
    if baked:
        fail(f"{ref}: the formula text itself contains {', '.join(baked)} — that is "
             f"an error value, not an expression")


def autofit(src: Path, wb: Workbook, only_sheet: str | None,
            cjk_only: bool) -> dict:
    """Widen every column that is too narrow to show what is in it.

    Values are read with openpyxl from the untouched input on disk — reading never
    rewrites a package — and the widths are then written surgically into the sheet
    XML this run is editing.

    Returns `{"changes": [...], "measured": [...]}`. **`measured` is not decoration.**
    An empty `changes` and a run where `--autofit` did nothing at all produce the
    same report, and a caller cannot tell them apart: measured 2026-08-16 (059
    §三十二) a model ran `--autofit`, got `changes: []`, ran it again with
    `--autofit-scope all`, got `changes: []`, and then went around the skill and set
    five widths by hand — to numbers larger than its own reading of this file's
    formula said were needed, which pushed a column onto a second page. A report
    that shows what was measured is what makes "already wide enough" an answer
    instead of a silence.

    Numbers are measured through their number format, and REGARDLESS of
    `--autofit-scope`: `###` is a display failure that has nothing to do with CJK.
    """
    import openpyxl
    from contextlib import closing
    book = openpyxl.load_workbook(src, data_only=False)
    values = openpyxl.load_workbook(src, data_only=True)
    changes: list[dict] = []
    measured: list[dict] = []
    with closing(book), closing(values):
        for ws in book.worksheets:
            if only_sheet and ws.title != only_sheet:
                continue
            if ws.title not in wb.sheets:
                continue
            # A title merged across A1:F1 is displayed across all six columns, so it
            # must not drive column A's width — Excel's own autofit ignores merged
            # cells for the same reason. A merge inside ONE column (a vertical merge)
            # buys no extra room and still counts.
            spanned = {c.coordinate for rng in ws.merged_cells.ranges
                       if rng.max_col > rng.min_col
                       for row in ws[rng.coord] for c in row}
            vws = values[ws.title]
            widest: dict[int, tuple[float, str, str]] = {}
            unrendered: dict[int, int] = {}
            for row, vrow in zip(ws.iter_rows(), vws.iter_rows(values_only=True)):
                for cell, cached in zip(row, vrow):
                    if cell.coordinate in spanned:
                        continue
                    v = cell.value
                    is_formula = isinstance(v, str) and v.startswith("=")
                    if isinstance(v, str) and not is_formula:
                        # Text drives the width only within the chosen scope.
                        if cjk_only and not has_cjk(v):
                            continue
                        shown, kind = v, "text"
                    else:
                        # A formula's own text is never displayed — what is displayed
                        # is its RESULT, through the cell's number format.
                        shown = displayed_text(cached, cell.number_format)
                        if shown is None:
                            unrendered[cell.column] = unrendered.get(cell.column, 0) + 1
                            continue
                        if not shown:
                            continue
                        kind = "number"
                    want = needed_width(shown)
                    if want > widest.get(cell.column, (0, "", ""))[0]:
                        widest[cell.column] = (want, shown, kind)
            target = wb.sheet(ws.title)
            for col in sorted(set(widest) | set(unrendered)):
                want, sample, kind = widest.get(col, (0.0, "", "none"))
                current = current_width(target, col)
                note = {"sheet": ws.title, "column": index_to_col(col),
                        "widest": sample[:24], "widest_is": kind,
                        "needs": want, "current": current}
                if unrendered.get(col):
                    # Saying "n cells were not measured" is the difference between a
                    # width that is right and one that merely was not contradicted.
                    note["unmeasured_cells"] = unrendered[col]
                if want and (current is None or current + 1e-6 < want):
                    changes.append({"sheet": ws.title, **target.set_column_width(
                        col, want,
                        reason=f"{sample[:16]!r} needs {want:g} display units"),
                        "was": current})
                    note["verdict"] = "widened"
                else:
                    note["verdict"] = "already wide enough"
                measured.append(note)
    return {"changes": changes, "measured": measured}


# Total referenced cells the precise dependency walk is allowed to expand. Chosen
# from measurement, not taste: at this size the walk costs well under a second, and
# the shape that blows past it (long chains of widening SUM ranges) is exactly the
# one where expanding is hopeless.
MAX_DEPENDENCY_SPAN = 500_000


def invalidate_stale_caches(src: Path, wb: Workbook,
                            edited: dict[str, str | None]) -> dict:
    """Clear the cached result of every formula cell this edit just made untrue.

    A cached `<v>` is what every consumer that is not a spreadsheet application
    reads — this skill's own reader and its own PDF renderer included. Leaving one
    in place after changing what it was computed from does not produce a stale
    number with a warning beside it; it produces a WRONG number that reports itself
    as calculated.

    Measured 2026-08-16 on 利润表.xlsx (059 §三十): `--set B2=1350` and the file
    still said 毛利 488.2 / 营业利润 47.6, `xlsx_read` marked both
    `uncalculated: false`, and the PDF preview put that table on the page —
    LibreOffice does NOT honour `fullCalcOnLoad` on export, so `calcPr` rescues
    nobody here. With the caches cleared the same render produces 553.6 and 113.0,
    because a formula cell with no cached result IS computed at load.

    `edited` maps "Sheet!REF" to the formula text now in that cell, or None if it
    now holds a constant. Only cells that actually carry a cached value are touched,
    so a workbook a library wrote (no caches at all) costs nothing and changes
    nothing.
    """
    import openpyxl
    from contextlib import closing

    formulas: dict[str, dict[str, str]] = {}
    cached: set[str] = set()
    fbook = openpyxl.load_workbook(src, read_only=True, data_only=False)
    vbook = openpyxl.load_workbook(src, read_only=True, data_only=True)
    with closing(fbook), closing(vbook):
        for name in fbook.sheetnames:
            per: dict[str, str] = {}
            # Zipped rather than indexed: random access on a read_only worksheet
            # re-reads the sheet, which is the memory behaviour X12 exists to keep out.
            for frow, vrow in zip(fbook[name].iter_rows(),
                                  vbook[name].iter_rows(values_only=True)):
                for cell, value in zip(frow, vrow):
                    expr = cell.value
                    if isinstance(expr, str) and expr.startswith("="):
                        per[cell.coordinate] = expr
                        if value is not None:
                            cached.add(f"{name}!{cell.coordinate}")
            formulas[name] = per

    # The graph has to describe the workbook AS WRITTEN, not as read: a cell that
    # this run turned into a formula has dependents of its own, and one that stopped
    # being a formula no longer reads anything.
    for node, expr in edited.items():
        sheet, _, ref = node.partition("!")
        per = formulas.setdefault(sheet, {})
        if expr is None:
            per.pop(ref, None)
        else:
            per[ref] = "=" + expr.lstrip("=")

    # The precise walk needs every reference expanded to single cells, and that is
    # quadratic on a shape real workbooks have: 10,000 rows of =SUM(A1:A<row>) is
    # ~50M edges. Measured 2026-08-16 — the first version of this function ran for
    # over nine minutes on exactly that and had to be killed. So price the expansion
    # BEFORE doing it, from cell counts alone, and when it is too big fall back to
    # clearing every cached formula: lossy, never wrong, and O(formulas).
    span = sum(r.cell_count for per in formulas.values() for f in per.values()
               for r in references(f))
    if span > MAX_DEPENDENCY_SPAN:
        stale, coarse = sorted(cached), (
            f"{span} referenced cell(s) is past the {MAX_DEPENDENCY_SPAN} this walk "
            f"is bounded to, so every cached formula was cleared instead of only the "
            f"dependents — no number left behind is wrong, but numbers that were "
            f"still true were dropped too")
        truncated = []
    else:
        graph, stats = build_graph(formulas)
        readers: dict[str, set[str]] = {}
        for node, deps in graph.items():
            for dep in deps:
                readers.setdefault(dep, set()).add(node)
        seen: set[str] = set()
        queue = list(edited)
        while queue:
            for reader in readers.get(queue.pop(), ()):
                if reader not in seen and reader not in edited:
                    seen.add(reader)
                    queue.append(reader)
        stale, coarse, truncated = sorted(n for n in seen if n in cached), None, \
            stats["truncated"]

    by_sheet: dict[str, set[str]] = {}
    for node in stale:
        sheet, _, ref = node.partition("!")
        if sheet in wb.sheets:
            by_sheet.setdefault(sheet, set()).add(ref)
    # One traversal per sheet. Clearing cell by cell rescans the row list every time
    # and is quadratic on exactly the workbooks where the sweep has most to do.
    cleared = {f"{sheet}!{ref}"
               for sheet, refs in by_sheet.items()
               for ref in wb.sheet(sheet).clear_cached(refs)}
    # Report what the sheet actually gave up, not what the walk hoped to clear.
    return {"cells": sorted(cleared), "truncated": truncated, "coarse": coarse}


def current_width(worksheet, column: int) -> float | None:
    from office.xmlorder import q
    cols = worksheet.root.find(q("cols"))
    if cols is None:
        return None
    for col in cols.findall(q("col")):
        if int(col.get("min")) <= column <= int(col.get("max")):
            w = col.get("width")
            return float(w) if w is not None else None
    return None


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in", dest="src", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    ap.add_argument("--sheet", help="default sheet for refs that do not name one")
    ap.add_argument("--set", dest="sets", action="append", default=[],
                    metavar="REF=VALUE")
    ap.add_argument("--set-formula", dest="formulas", action="append", default=[],
                    metavar="REF=EXPR")
    ap.add_argument("--append-row", dest="rows", action="append", default=[],
                    metavar="V1,V2,...",
                    help="comma-separated; a value starting with = is a formula")
    ap.add_argument("--autofit", action="store_true",
                    help="give text columns an explicit width (X15)")
    ap.add_argument("--autofit-scope", choices=("cjk", "all"), default="cjk",
                    help="cjk (default) widens only columns holding CJK text")
    ap.add_argument("--report", type=Path, help="write the change log here")
    args = ap.parse_args()

    def entry():
        replaced = ensure_distinct(args.src, args.out)
        if not (args.sets or args.formulas or args.rows or args.autofit):
            fail("nothing to do: pass --set / --set-formula / --append-row / --autofit")
        pkg = Package.open(args.src)
        wb = Workbook(pkg)
        changes: list[dict] = []
        # "Sheet!REF" -> the formula now in that cell, or None for a constant. Feeds
        # the staleness sweep below; a cell nobody records here keeps a cached value
        # that no longer matches what it was computed from.
        edited: dict[str, str | None] = {}

        for spec in args.sets:
            sheet, ref, value = split_target(spec, args.sheet)
            ws = wb.sheet(sheet)
            changes.append({"sheet": ws.name, "op": "set",
                            **ws.set_cell(ref, coerce(value))})
            edited[f"{ws.name}!{ref.upper()}"] = None
        for spec in args.formulas:
            sheet, ref, expr = split_target(spec, args.sheet)
            if not expr.strip():
                fail(f"--set-formula {spec!r} has an empty expression")
            check_formula(expr, set(wb.sheets), ref)
            ws = wb.sheet(sheet)
            changes.append({"sheet": ws.name, "op": "set-formula",
                            **ws.set_cell(ref, expr, formula=True)})
            edited[f"{ws.name}!{ref.upper()}"] = expr
        for spec in args.rows:
            values, formulas = [], set()
            for i, raw in enumerate(spec.split(","), start=1):
                raw = raw.strip()
                if raw.startswith("="):
                    formulas.add(i)
                    values.append(raw[1:])
                else:
                    values.append(coerce(raw))
            ws = wb.sheet(args.sheet)
            appended = ws.append_row(values, formulas)
            changes.append({"sheet": ws.name, "op": "append-row", **appended})
            for i, value in enumerate(values, start=1):
                if value is None:  # append_row skips these, so nothing was written
                    continue
                ref = f"{index_to_col(i)}{appended['row']}"
                edited[f"{ws.name}!{ref}"] = value if i in formulas else None

        fit = autofit(args.src, wb, args.sheet, args.autofit_scope == "cjk") \
            if args.autofit else {"changes": [], "measured": []}
        stale = invalidate_stale_caches(args.src, wb, edited) if edited else \
            {"cells": [], "truncated": [], "coarse": None}

        saved = wb.save(args.out)
        # The package is re-opened from disk, not inspected in memory: the claim is
        # about the FILE, and an in-memory check would agree with itself even if
        # saving wrote something else.
        #
        # Only damage this edit INTRODUCED is fatal. A file that arrived with a
        # dangling relationship is still a file the user wants edited — refusing it
        # would make the skill useless on exactly the documents that need help — but
        # an edit that creates new damage must write nothing at all.
        was = set(check_package(Package.open(args.src)))
        now = check_package(Package.open(args.out))
        introduced = [f for f in now if f not in was]
        if introduced:
            args.out.unlink(missing_ok=True)
            fail("this edit would have introduced package damage, so nothing was "
                 "written: " + "; ".join(introduced[:3]))
        lost = sorted(set(pkg_names(args.src)) - set(pkg_names(args.out)))
        report = {"in": args.src.name, "out": args.out.name,
                  "replaced_existing": replaced,
                  "changes": changes, "widths": fit["changes"],
                  "widths_measured": fit["measured"],
                  "sheets_written": saved["sheets_written"],
                  "caches_dropped": saved["caches_dropped"],
                  "caches_invalidated": stale["cells"],
                  "parts_in": len(pkg_names(args.src)),
                  "parts_out": len(pkg_names(args.out)),
                  "parts_lost": [p for p in lost if p not in saved["caches_dropped"]],
                  "pre_existing_package_findings": sorted(was)}
        if stale["cells"]:
            report["caches_invalidated_note"] = (
                f"{len(stale['cells'])} formula cell(s) were computed from a cell "
                f"this edit changed, so their stored results no longer match the "
                f"file and have been cleared rather than left to be read as current. "
                f"Run xlsx_recalc.py to put real numbers back in")
        if stale["truncated"]:
            # An incomplete sweep that reports nothing is the defect this whole
            # function exists to remove, one level up.
            report["caches_invalidated_incomplete"] = stale["truncated"]
        if stale["coarse"]:
            report["caches_invalidated_coarse"] = stale["coarse"]
        if args.autofit and not fit["changes"]:
            # The one sentence that separates "I measured, they all fit" from "the
            # flag did nothing" — which is what an empty `changes` looked like.
            report["widths_note"] = (
                f"{len(fit['measured'])} column(s) measured, none too narrow — the "
                f"widths already in the file are at or above what the content needs. "
                f"Widening past that is not an improvement: it pushes columns onto "
                f"the next page when the sheet is printed or converted")
        emit(report, args.report, "changes", "widths", "widths_measured",
             "caches_invalidated")

    return run(entry)


def pkg_names(path: Path) -> list[str]:
    import zipfile
    with zipfile.ZipFile(path) as z:
        return [n for n in z.namelist() if not n.endswith("/")]


if __name__ == "__main__":
    raise SystemExit(main())
