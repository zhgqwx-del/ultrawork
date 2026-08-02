#!/usr/bin/env python3
"""X11 — CSV / JSON in and out. X12 — read large workbooks without loading them.

Reading ALWAYS goes through openpyxl's `read_only` mode, which is not an
optimisation flag but the difference between "works" and "eats the machine".
Measured with `tracemalloc`, reading a three-column sheet:

    rows       normal      read_only     ratio
     10,000    12.9 MB        1.7 MB      7.5x
     50,000    62.3 MB        5.1 MB     12.2x
    250,000   326.6 MB       22.0 MB     14.9x

Normal mode costs ~1.3 KB per row and that figure does not improve with size —
it builds a Python object for every cell and keeps them all. A 500k-row export
would need most of a gigabyte for a file that is 8 MB on disk.

    python3 xlsx_convert.py --in book.xlsx --to csv  --out ./out
    python3 xlsx_convert.py --in book.xlsx --to jsonl --out rows.jsonl --sheet 明细
    python3 xlsx_convert.py --in book.xlsx --stats --sheet 明细
    python3 xlsx_convert.py --from data.csv --out new.xlsx --sheet 明细

CSV is written `utf-8-sig` (with a BOM). Without it, Chinese in a CSV opens as
mojibake in Excel — and Excel is where a CSV export almost always goes.

Values, not formulas: an export reads what a spreadsheet would DISPLAY. A file no
one has calculated has no values at all, so that case is reported rather than
exported as a sheet full of blanks (`xlsx_recalc.py` is the fix).
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from xlsxcommon import XlsxError, emit, ensure_distinct, fail, needed_width, run  # noqa: E402

FORMATS = ("csv", "json", "jsonl")
STDOUT_ROWS = 5


def open_streaming(path: Path):
    """read_only + data_only: the bounded-memory reader."""
    import openpyxl
    if not path.is_file():
        fail(f"no such file: {path}")
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        fail(f"cannot open {path.name} as a workbook: {type(e).__name__}: {e}")


def sheet_rows(wb, name: str):
    if name not in wb.sheetnames:
        fail(f"no sheet named {name!r} (have: {', '.join(wb.sheetnames)})")
    return wb[name].iter_rows(values_only=True)


def uncalculated_count(path: Path, sheet: str) -> int:
    """Formula cells with no cached result — the reason an export can be all blanks."""
    import openpyxl
    from contextlib import closing
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    vals = openpyxl.load_workbook(path, read_only=True, data_only=True)
    n = 0
    with closing(wb), closing(vals):
        if sheet not in wb.sheetnames:
            return 0
        for frow, vrow in zip(wb[sheet].iter_rows(values_only=True),
                              vals[sheet].iter_rows(values_only=True)):
            for f, v in zip(frow, vrow):
                if isinstance(f, str) and f.startswith("=") and v is None:
                    n += 1
    return n


def export(src: Path, fmt: str, out: Path, sheet: str | None,
           headers: bool, header_row: int) -> dict:
    from contextlib import closing
    wb = open_streaming(src)
    written: list[dict] = []
    sample: list = []
    with closing(wb):
        names = [sheet] if sheet else list(wb.sheetnames)
        if fmt == "csv" and len(names) > 1:
            out.mkdir(parents=True, exist_ok=True)
        for name in names:
            rows = sheet_rows(wb, name)
            target = (out / f"{name}.csv") if (fmt == "csv" and len(names) > 1) else out
            target.parent.mkdir(parents=True, exist_ok=True)
            count = 0
            if fmt == "csv":
                # utf-8-sig: Excel reads a BOM-less UTF-8 CSV as the local codepage
                # and Chinese arrives as mojibake. newline="" is required or every
                # row gets a blank line after it on Windows.
                with target.open("w", encoding="utf-8-sig", newline="") as fh:
                    w = csv.writer(fh)
                    for row in rows:
                        w.writerow(["" if v is None else v for v in row])
                        count += 1
                        if len(sample) < STDOUT_ROWS:
                            sample.append(list(row))
            else:
                head: list[str] | None = None
                with target.open("w", encoding="utf-8") as fh:
                    if fmt == "json":
                        fh.write("[\n")
                    first = True
                    for index, row in enumerate(rows, 1):
                        # Everything above the header row is dropped. Real sheets
                        # routinely carry a merged title above the headers, and
                        # taking row 1 on faith turns the title into a column name.
                        if index < header_row:
                            continue
                        if headers and head is None:
                            head = [str(v) if v is not None else f"col{i}"
                                    for i, v in enumerate(row, 1)]
                            continue
                        obj = dict(zip(head, row)) if head else \
                            {f"col{i}": v for i, v in enumerate(row, 1)}
                        text = json.dumps(obj, ensure_ascii=False, default=str)
                        if fmt == "jsonl":
                            fh.write(text + "\n")
                        else:
                            fh.write(("" if first else ",\n") + "  " + text)
                        first = False
                        count += 1
                        if len(sample) < STDOUT_ROWS:
                            sample.append(obj)
                    if fmt == "json":
                        fh.write("\n]\n")
            written.append({"sheet": name, "rows": count, "path": str(target)})
    return {"written": written, "sample": sample}


def stats(src: Path, sheet: str | None, header_row: int = 1) -> dict:
    """Running aggregates, computed without holding the sheet."""
    from contextlib import closing
    wb = open_streaming(src)
    out: list[dict] = []
    with closing(wb):
        for name in ([sheet] if sheet else list(wb.sheetnames)):
            rows = sheet_rows(wb, name)
            head: list | None = None
            per: dict[int, dict] = {}
            total = 0
            for index, row in enumerate(rows, 1):
                if index < header_row:
                    continue
                if head is None:
                    head = list(row)
                    continue
                total += 1
                for i, v in enumerate(row):
                    if isinstance(v, bool) or not isinstance(v, (int, float)):
                        continue
                    s = per.setdefault(i, {"count": 0, "sum": 0.0,
                                           "min": None, "max": None})
                    s["count"] += 1
                    s["sum"] += float(v)
                    s["min"] = v if s["min"] is None else min(s["min"], v)
                    s["max"] = v if s["max"] is None else max(s["max"], v)
            out.append({"sheet": name, "rows": total,
                        "columns": [{"header": (head[i] if head and i < len(head)
                                                and head[i] is not None
                                                else f"col{i + 1}"),
                                     **v,
                                     "mean": v["sum"] / v["count"] if v["count"] else None}
                                    for i, v in sorted(per.items())]})
    return {"stats": out}


def import_rows(src: Path, out: Path, sheet: str, autofit: bool) -> dict:
    import openpyxl
    if not src.is_file():
        fail(f"no such file: {src}")
    if src.suffix.lower() == ".csv":
        with src.open(encoding="utf-8-sig", newline="") as fh:
            rows = [list(r) for r in csv.reader(fh)]
    else:
        try:
            data = json.loads(src.read_text(encoding="utf-8"))
        except json.JSONDecodeError as e:
            fail(f"{src.name} is not valid JSON: {e}")
        if isinstance(data, dict):
            data = [data]
        if not isinstance(data, list) or not data:
            fail(f"{src.name} must be a non-empty list of objects (or one object)")
        if not all(isinstance(o, dict) for o in data):
            fail(f"{src.name}: every element must be an object")
        keys: list[str] = []
        for o in data:
            for k in o:
                if k not in keys:
                    keys.append(k)
        rows = [keys] + [[o.get(k) for k in keys] for o in data]

    # write_only: the import mirror of read_only. Building a normal Workbook holds
    # every cell in memory, which is the same wall the export side hits.
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(sheet)
    widest: dict[int, float] = {}
    for row in rows:
        ws.append([coerce_cell(v) for v in row])
        if autofit:
            for i, v in enumerate(row, 1):
                if isinstance(v, str) and v:
                    widest[i] = max(widest.get(i, 0.0), needed_width(v))
    if autofit and widest:
        from openpyxl.utils import get_column_letter
        for i, w in widest.items():
            ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(out)
    wb.close()
    return {"sheet": sheet, "rows": len(rows), "widths_set": len(widest) if autofit else 0}


def coerce_cell(v):
    """CSV gives strings; a column of numbers should arrive as numbers."""
    if not isinstance(v, str):
        return v
    if v == "":
        return None
    low = v.strip().lower()
    if low in ("true", "false"):
        return low == "true"
    try:
        return int(v)
    except ValueError:
        pass
    try:
        return float(v)
    except ValueError:
        return v


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in", dest="src", type=Path)
    ap.add_argument("--from", dest="source", type=Path,
                    help="a .csv or .json file to build a workbook FROM")
    ap.add_argument("--to", choices=FORMATS)
    ap.add_argument("--out", type=Path)
    ap.add_argument("--sheet")
    ap.add_argument("--stats", action="store_true",
                    help="streaming aggregates instead of an export")
    ap.add_argument("--no-headers", action="store_true",
                    help="the header row is data, not column names (json/jsonl)")
    ap.add_argument("--header-row", type=int, default=1, metavar="N",
                    help="1-based row holding the column names; rows above it are "
                         "dropped (a merged title above the headers is common)")
    ap.add_argument("--autofit", action="store_true",
                    help="on import, set CJK-aware column widths")
    ap.add_argument("--report", type=Path)
    args = ap.parse_args()

    def entry():
        if args.source:
            if not args.out:
                fail("--from needs --out to say where the workbook goes")
            ensure_distinct(args.source, args.out)
            report = import_rows(args.source, args.out,
                                 args.sheet or "Sheet1", args.autofit)
            emit({"from": args.source.name, "out": args.out.name, **report},
                 args.report)
            return
        if not args.src:
            fail("pass --in (to export) or --from (to import)")
        if args.stats:
            emit({"file": args.src.name,
                  **stats(args.src, args.sheet, args.header_row)},
                 args.report, "stats")
            return
        if not args.to or not args.out:
            fail("exporting needs --to csv|json|jsonl and --out")
        ensure_distinct(args.src, args.out)
        blank = uncalculated_count(args.src, args.sheet) if args.sheet else 0
        if args.header_row < 1:
            fail(f"--header-row {args.header_row} must be 1 or more (rows are 1-based)")
        report = export(args.src, args.to, args.out, args.sheet,
                        headers=not args.no_headers, header_row=args.header_row)
        payload = {"file": args.src.name, "format": args.to,
                   "read_mode": "read_only (bounded memory)", **report}
        if blank:
            # Not a failure — but an export full of blanks with no explanation is
            # how someone concludes the data is gone.
            payload["warning"] = (
                f"{blank} formula cell(s) on {args.sheet} have no cached value, so "
                f"they export as empty. Run xlsx_recalc.py first if you want numbers")
        emit(payload, args.report, "written", "sample")

    return run(entry)


if __name__ == "__main__":
    raise SystemExit(main())
