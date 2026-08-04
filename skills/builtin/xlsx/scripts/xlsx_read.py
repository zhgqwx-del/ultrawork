#!/usr/bin/env python3
"""X1 — read cells, sheets and ranges out of a workbook.

Reads two views of every cell because a spreadsheet has two, and confusing them is
the classic mistake:

    value   — what the last application to calculate the file left behind (the
              CACHED result; absent in a file written by a library)
    formula — the expression itself

`openpyxl.load_workbook(data_only=True)` gives the first and `data_only=False` the
second, and neither alone is the truth. A file this skill wrote has formulas and no
cached values at all, so a reader that only asks for values reports an empty sheet
and every downstream conclusion is drawn from nothing.

    python3 xlsx_read.py --in book.xlsx                       # sheet inventory
    python3 xlsx_read.py --in book.xlsx --sheet 利润表 --range A1:D12
    python3 xlsx_read.py --in book.xlsx --cells B4 C4 --out cells.json
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from xlsxcommon import XlsxError, emit, fail, has_cjk, run  # noqa: E402

MAX_SCAN_CELLS = 200_000
# The summary path gets its own, larger budget. 200k is sized for a window a caller
# asked to see cell-by-cell; the inventory only counts, and a legitimate
# 100k-row x 10-column sheet must not come back truncated. Measured on this machine:
# 200k cells ~0.5s, so 2M is a few seconds — bounded, which is the whole point.
MAX_INVENTORY_CELLS = 2_000_000


def load_pair(path: Path):
    import openpyxl
    if not path.is_file():
        fail(f"no such file: {path}")
    try:
        formulas = openpyxl.load_workbook(path, data_only=False)
        values = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:  # noqa: BLE001 - openpyxl raises several unrelated types
        fail(f"cannot open {path.name} as a workbook: {type(e).__name__}: {e}")
    return formulas, values


def cell_entry(fws, vws, ref: str) -> dict:
    from openpyxl.utils.cell import coordinate_from_string
    try:
        coordinate_from_string(ref)
    except Exception as e:  # noqa: BLE001
        fail(f"{ref!r} is not a cell reference (expected e.g. B4): {e}")
    f, v = fws[ref].value, vws[ref].value
    is_formula = isinstance(f, str) and f.startswith("=")
    return {
        "ref": ref,
        "value": v,
        "formula": f if is_formula else None,
        # An agent that sees `value: null` on a formula cell needs to be told the
        # difference between "the answer is empty" and "nobody has calculated this
        # file yet" — the second is normal for anything a library wrote.
        "uncalculated": bool(is_formula and v is None),
    }


def sheet_inventory(fwb, vwb) -> list[dict]:
    """Per-sheet summary, with the same cell budget --range already had.

    ⚠️ MAX_SCAN_CELLS used to guard read_range() and nothing else, so the summary
    path — the one that runs when no --range is given, i.e. the default — had no
    bound at all. It scanned whatever `<dimension>` claimed.

    That is not a theoretical hole. `<dimension ref="A1:XFD1048576"/>` on a sheet
    holding five real rows is something non-Excel writers emit, and openpyxl
    believes it: max_row 1048576 x max_column 16384 = 17 BILLION cells, walked
    twice. Found by the L3 real-corpus run (059 §六·补九) on a **145 KB** file from
    calamine's issue corpus, i.e. a workbook a real user actually hit: xlsx_read
    had not returned after ten minutes. A file that small hanging forever is worse
    than one that fails, because nothing about it looks like it should.

    Two things change here, and the second one is a wrong ANSWER, not just a hang:
      - one pass instead of two, under a hard cell budget that says so out loud when
        it stops. A truncated scan that looks like a complete one is the failure this
        whole file is written to avoid.
      - `rows`/`columns` are now counted from cells that actually hold a value,
        not read off `<dimension>`. That same 145 KB workbook was REPORTED as
        1048576 rows x 16384 columns. It has two.

    (`reset_dimensions()` looks like the obvious fix and is not available here:
    openpyxl 3.1.5 defines it on ReadOnlyWorksheet only, and this entry point loads
    two full workbooks because it needs the formula view AND the cached-value view
    of every cell. Checked rather than assumed — a guarded call that silently never
    runs would have left the docstring claiming a fix that was not happening.)
    """
    out = []
    for ws in fwb.worksheets:
        vws = vwb[ws.title]
        formulas = uncalculated = scanned = 0
        cjk = truncated = False
        rows = cols = 0
        for row in ws.iter_rows():
            for cell in row:
                scanned += 1
                if cell.value is None:
                    continue
                rows = max(rows, cell.row or 0)
                cols = max(cols, cell.column or 0)
                if isinstance(cell.value, str):
                    if not cjk and has_cjk(cell.value):
                        cjk = True
                    if cell.value.startswith("="):
                        formulas += 1
                        if vws[cell.coordinate].value is None:
                            uncalculated += 1
            if scanned > MAX_INVENTORY_CELLS:
                truncated = True
                break
        entry = {
            "name": ws.title, "state": ws.sheet_state,
            "rows": rows, "columns": cols,
            "dimensions": ws.dimensions,
            "formulas": formulas, "uncalculated_formulas": uncalculated,
            "has_cjk": cjk,
        }
        if truncated:
            # Never silent: a partial count that reads like a total is the one
            # outcome worth failing over.
            entry["scan_truncated"] = {
                "limit": MAX_INVENTORY_CELLS,
                "note": f"stopped after {MAX_INVENTORY_CELLS} cells; formulas, "
                        f"uncalculated_formulas, has_cjk, rows and columns are "
                        f"lower bounds, not totals. Use --range for a window.",
            }
        out.append(entry)
    return out


def read_range(fws, vws, ref: str) -> list[dict]:
    from openpyxl.utils.cell import range_boundaries
    try:
        c0, r0, c1, r1 = range_boundaries(ref)
    except Exception as e:  # noqa: BLE001
        fail(f"{ref!r} is not a range (expected e.g. A1:D12): {e}")
    if None in (c0, r0, c1, r1):
        fail(f"{ref!r} is an open-ended range; give both corners, e.g. A1:D12")
    span = (c1 - c0 + 1) * (r1 - r0 + 1)
    if span > MAX_SCAN_CELLS:
        fail(f"range {ref} covers {span} cells; ask for a smaller window "
             f"(limit {MAX_SCAN_CELLS}) or use --sheet without --range for a summary")
    from openpyxl.utils import get_column_letter
    return [cell_entry(fws, vws, f"{get_column_letter(c)}{r}")
            for r in range(r0, r1 + 1) for c in range(c0, c1 + 1)]


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in", dest="src", required=True, type=Path)
    ap.add_argument("--sheet", help="sheet name; default = the first one")
    ap.add_argument("--range", dest="area", help="e.g. A1:D12")
    ap.add_argument("--cells", nargs="+", metavar="REF", help="e.g. B4 C4")
    ap.add_argument("--out", type=Path, help="write the full report here")
    args = ap.parse_args()

    def entry():
        fwb, vwb = load_pair(args.src)
        sheets = sheet_inventory(fwb, vwb)
        payload = {"file": args.src.name, "sheets": sheets}
        if args.area or args.cells:
            name = args.sheet or fwb.sheetnames[0]
            if name not in fwb.sheetnames:
                fail(f"no sheet named {name!r} (have: {', '.join(fwb.sheetnames)})")
            fws, vws = fwb[name], vwb[name]
            cells = read_range(fws, vws, args.area) if args.area else \
                [cell_entry(fws, vws, r) for r in args.cells]
            payload["sheet"] = name
            payload["cells"] = cells
        elif args.sheet:
            if args.sheet not in fwb.sheetnames:
                fail(f"no sheet named {args.sheet!r} (have: {', '.join(fwb.sheetnames)})")
            payload["sheet"] = args.sheet
        emit(payload, args.out, "cells", "sheets")
        fwb.close()
        vwb.close()

    return run(entry)


if __name__ == "__main__":
    raise SystemExit(main())
